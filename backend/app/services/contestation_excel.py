"""
Gerador de Relatório Excel de Contestação — Gestora Smart
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERDE     = "1E9B6B"
VERDE_ESC = "157a54"
VERDE_CL  = "dcfce7"
LARANJA   = "f97316"
LARANJA_CL= "fff7ed"
AZUL_CL   = "dbeafe"
AZUL      = "1d4ed8"
VERMELHO  = "dc2626"
VERM_CL   = "fee2e2"
CINZA     = "f8fafc"
CINZA_ESC = "1e293b"
BRANCO    = "FFFFFF"
BORDA     = "e2e8f0"
AMARELO_CL= "fef9c3"
AMARELO_DK= "a16207"

MESES = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def _fill(h):  return PatternFill("solid", fgColor=h)
def _font(bold=False, color=CINZA_ESC, size=9): return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border(): s = Side(style="thin", color=BORDA); return Border(left=s, right=s, top=s, bottom=s)
def _bottom(): return Border(bottom=Side(style="thin", color=BORDA))

BRL = '#,##0.00'
INT = '#,##0'

def _header(ws, row, col, val, bg=VERDE, fg=BRANCO, size=8):
    c = ws.cell(row, col, val)
    c.fill  = _fill(bg); c.font = _font(True, fg, size)
    c.alignment = _align("center"); c.border = _border()
    return c

def _cell(ws, row, col, val, bg=BRANCO, bold=False, align="left", fmt=None, color=CINZA_ESC):
    c = ws.cell(row, col, val)
    c.fill = _fill(bg); c.font = _font(bold, color); c.alignment = _align(align); c.border = _bottom()
    if fmt: c.number_format = fmt
    return c

def _title_row(ws, text, period, color=VERDE):
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"GESTORA SMART — {text.upper()}  |  {period}"
    c.fill  = _fill(color); c.font = Font(bold=True, color=BRANCO, size=13, name="Calibri")
    c.alignment = _align("left"); ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6


def generate_contestation_excel(cycle, items, credits=None) -> io.BytesIO:
    """
    Gera o relatório Excel de contestação.
    items: lista de ContestationItem do banco
    """
    wb = Workbook()
    wb.remove(wb.active)
    periodo = f"{cycle.month:02d}/{cycle.year}"
    mes_ext = f"{MESES[cycle.month]} {cycle.year}"

    # Separa itens por tipo e status
    def filter_items(type_val, status_list=None):
        result = [i for i in items if i.type == type_val]
        if status_list:
            result = [i for i in result if i.status in status_list]
        return result

    from app.models.contestation import ContestationItemType as CIT, ContestationItemStatus as CIS

    acima       = filter_items(CIT.VALOR_ACIMA_CONTRATO,   [CIS.DETECTADO, CIS.CONTESTAR, CIS.ENVIADO, CIS.ACEITO])
    pcte        = filter_items(CIT.PCTE_ADICIONAL_INDEVIDO,[CIS.DETECTADO, CIS.CONTESTAR, CIS.ENVIADO, CIS.ACEITO])
    nao_ident   = filter_items(CIT.LINHA_NAO_IDENTIFICADA, [CIS.DETECTADO, CIS.CONTESTAR, CIS.ENVIADO, CIS.ACEITO])
    transferen  = filter_items(CIT.TRANSFERENCIA)
    cs_items    = filter_items(CIT.CS)

    # ── ABA 1: RESUMO ────────────────────────────────────────
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False
    _title_row(ws, f"Relatório de Contestação — {mes_ext}", periodo)

    ws.merge_cells("A3:G3")
    ws["A3"] = "Categoria"
    ws["A3"].fill = _fill(CINZA_ESC); ws["A3"].font = _font(True, BRANCO, 9); ws["A3"].alignment = _align("center")

    def resume_row(ws, row, label, qtd, valor, bg=BRANCO, bold=False):
        for col, (val, fmt, aln) in enumerate([
            (label, None, "left"), (qtd, INT, "center"), (valor, BRL, "right")
        ], 1):
            c = _cell(ws, row, col, val, bg=bg, bold=bold, align=aln, fmt=fmt)
        ws.row_dimensions[row].height = 16

    summary_data = [
        ("Valor acima do contrato",      len(acima),      sum(i.valor_diferenca or 0 for i in acima),    VERM_CL),
        ("Pacote adicional indevido",     len(pcte),       sum(i.valor_faturado or 0 for i in pcte),     AMARELO_CL),
        ("Linhas não identificadas",      len(nao_ident),  sum(i.valor_faturado or 0 for i in nao_ident), VERM_CL),
        ("Transferências (verificar)",    len(transferen), sum(i.valor_faturado or 0 for i in transferen), AZUL_CL),
        ("CS — Licenças (informativo)",   len(cs_items),   cycle.valor_cs or 0,                           CINZA),
    ]
    total_contestado = sum(v for _, _, v, _ in summary_data[:3])

    for col, h in enumerate(["Categoria", "Qtd Itens", "Valor (R$)"], 1):
        _header(ws, 4, col, h)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    for i, (label, qtd, valor, bg) in enumerate(summary_data, 5):
        resume_row(ws, i, label, qtd, round(valor, 2), bg=bg)

    resume_row(ws, 10, "TOTAL A CONTESTAR", sum(q for _, q, _, _ in summary_data[:3]),
               round(total_contestado, 2), bg=CINZA_ESC, bold=True)
    ws["A10"].font = _font(True, BRANCO, 9); ws["B10"].font = _font(True, BRANCO, 9); ws["C10"].font = _font(True, BRANCO, 9)

    # Créditos recebidos
    if credits:
        ws["A12"] = "Histórico de Créditos Recebidos"
        ws["A12"].font = _font(True, VERDE, 10)
        for col, h in enumerate(["Referência", "Valor Contestado", "Valor Recebido", "Data", "Observação"], 1):
            _header(ws, 13, col, h, bg=VERDE)
        for i, cr in enumerate(credits, 14):
            _cell(ws, i, 1, f"{cr.ref_month:02d}/{cr.ref_year}")
            _cell(ws, i, 2, cr.valor_contestado, fmt=BRL, align="right")
            _cell(ws, i, 3, cr.valor_recebido, fmt=BRL, align="right", bold=True, color=VERDE)
            _cell(ws, i, 4, cr.data_recebimento.strftime("%d/%m/%Y") if cr.data_recebimento else "")
            _cell(ws, i, 5, cr.observacao or "")

    # ── ABA 2: ITENS PARA CONTESTAR ──────────────────────────
    COLS_CONTEST = [
        ("MSISDN",           "msisdn",          18, None,  "left"),
        ("ICCID",            "iccid",            22, None,  "left"),
        ("Operadora",        "operadora",        18, None,  "left"),
        ("Pedido",           "id_pedido",        14, None,  "center"),
        ("Pacote Fornecedor","pacote_forn",       30, None,  "left"),
        ("Status Forn.",     "status_forn",       16, None,  "center"),
        ("Dias Forn.",       "dias_forn",         10, INT,   "center"),
        ("Custo Contrato",   "valor_contrato",    14, BRL,   "right"),
        ("Valor Esperado",   "valor_esperado",    14, BRL,   "right"),
        ("Valor Produto",    "valor_produto",     14, BRL,   "right"),
        ("Valor Faturado",   "valor_faturado",    14, BRL,   "right"),
        ("Diferença",        "valor_diferenca",   14, BRL,   "right"),
        ("Observação",       "observacao",        50, None,  "left"),
        ("Status",           "status",            14, None,  "center"),
    ]

    def write_items_sheet(ws_name, item_list, bg_header=VERDE, title=None):
        ws2 = wb.create_sheet(ws_name)
        ws2.sheet_view.showGridLines = False
        _title_row(ws2, title or ws_name, periodo, bg_header)
        ws2.freeze_panes = "A4"

        for col, (header, _, width, _, _) in enumerate(COLS_CONTEST, 1):
            _header(ws2, 3, col, header, bg=bg_header)
            ws2.column_dimensions[get_column_letter(col)].width = width

        for r, item in enumerate(item_list, 4):
            bg = CINZA if r % 2 == 0 else BRANCO
            for col, (_, attr, _, fmt, align) in enumerate(COLS_CONTEST, 1):
                val = getattr(item, attr, None)
                bold_col = (attr == "valor_diferenca")
                color_col = VERMELHO if (attr == "valor_diferenca" and (val or 0) > 0) else CINZA_ESC
                if attr == "status":
                    STATUS_BG = {
                        "detectado": CINZA,     "contestar": VERDE_CL,
                        "ignorar":   AMARELO_CL,"enviado":   AZUL_CL,
                        "aceito":    VERDE_CL,  "rejeitado": VERM_CL,
                    }
                    STATUS_LABEL = {
                        "detectado":"Detectado","contestar":"Contestar",
                        "ignorar":"Ignorar","enviado":"Enviado",
                        "aceito":"Aceito","rejeitado":"Rejeitado",
                    }
                    c2 = ws2.cell(r, col, STATUS_LABEL.get(str(val), str(val or "")))
                    c2.fill = _fill(STATUS_BG.get(str(val), CINZA))
                    c2.font = _font(True, CINZA_ESC, 8); c2.alignment = _align("center"); c2.border = _bottom()
                else:
                    c2 = _cell(ws2, r, col, val, bg=bg, bold=bold_col, align=align, fmt=fmt, color=color_col)
            ws2.row_dimensions[r].height = 15

        # Total
        total_row = len(item_list) + 4
        ws2.merge_cells(f"A{total_row}:K{total_row}")
        tc = ws2.cell(total_row, 1, "TOTAL")
        tc.fill = _fill(CINZA_ESC); tc.font = _font(True, BRANCO)
        tv = ws2.cell(total_row, 12,
                      round(sum((getattr(i, "valor_diferenca", None) or getattr(i, "valor_faturado", 0) or 0) for i in item_list), 2))
        tv.fill = _fill(CINZA_ESC); tv.font = _font(True, BRANCO); tv.number_format = BRL; tv.alignment = _align("right")
        ws2.row_dimensions[total_row].height = 18

    write_items_sheet("Valor Acima Contrato",  acima,     VERDE,   "Itens com Valor Acima do Contrato")
    write_items_sheet("Pcte.Adicional Indevido",pcte,    LARANJA,  "Pacote Adicional em Operadora Não-Claro")
    write_items_sheet("Linhas Não Identificadas",nao_ident,VERMELHO if nao_ident else CINZA_ESC, "Linhas Fora do Inventário")
    write_items_sheet("Transferências",        transferen, "4472C4", "Transferências — Verificar Manualmente")

    # ── ABA CS ───────────────────────────────────────────────
    ws_cs = wb.create_sheet("CS — Informativo")
    ws_cs.sheet_view.showGridLines = False
    _title_row(ws_cs, "CS — Licenças do Analista (Informativo)", periodo, CINZA_ESC)
    ws_cs["A3"] = f"Valor total CS em {mes_ext}:"
    ws_cs["A3"].font = _font(True, CINZA_ESC, 10)
    ws_cs["B3"] = cycle.valor_cs or 0
    ws_cs["B3"].font = _font(True, VERDE, 14); ws_cs["B3"].number_format = BRL
    ws_cs["A4"] = "Aba informativa. Este valor não entra na contestação."
    ws_cs["A4"].font = _font(False, "94a3b8", 9)
    ws_cs.column_dimensions["A"].width = 40; ws_cs.column_dimensions["B"].width = 20

    # Salva
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
