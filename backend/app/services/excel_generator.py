"""
Gerador de Excel — Gestora Smart
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MESES = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def _fmt_date(d):
    if d is None:
        return None
    try:
        return d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)[:10]
    except Exception:
        return str(d)[:10]


# ════════════════════════════════════════════════════════════════
#  FORMATO MODELO — 32 colunas (A–AF) · exato do Pasta4.xlsx
# ════════════════════════════════════════════════════════════════

HEADERS_32 = [
    "Nome do pedido",                # A   1
    "ID do pedido",                  # B   2
    "Data de ativação",              # C   3
    "Preço de ativação",             # D   4
    "Nome do contrato",              # E   5
    "ID do contrato",                # F   6
    "Data de cancelamento",          # G   7
    "ID_CPF/CNPJ",                   # H   8
    "CPF/CNPJ",                      # I   9
    "Nome do cliente",               # J   10
    "Franquia (MB)",                 # K   11
    "ICCID",                         # L   12
    "Preço do MB Excedente",         # M   13
    "MSISDN",                        # N   14
    "Operadora",                     # O   15
    "Data fim da pré-ativação",      # P   16
    "Mensalidade",                   # Q   17
    "Crédito adicionado no Simcard", # R   18  ← MB (corrigido pelo sistema)
    "Status",                        # S   19
    "Operadora específica",          # T   20
    "Data de término da suspensão",  # U   21
    "Data de início da suspensão",   # V   22
    "Consumo total (KB)",            # W   23
    # ── calculadas pelo sistema (fundo amarelo) ──────────────────
    "Reajuste",                      # X   24
    "Reajuste 2025",                 # Y   25
    "Excedente",                     # Z   26
    "Multa Cancelamento",            # AA  27
    "SMS",                           # AB  28
    "Dias",                          # AC  29
    "Mensalidade",                   # AD  30
    "Ativação",                      # AE  31
    "TOTAL",                         # AF  32
]

# Colunas calculadas (24–32) → fundo amarelo
_CALC_COLS = set(range(24, 33))

# Colunas monetárias (BRL)
_MONEY_COLS = {4, 13, 17, 25, 26, 27, 28, 30, 31, 32}

# Colunas com número inteiro / sem casas
_INT_COLS = {29}

# Colunas percentual
_PCT_COLS = {24}

HDR_FILL  = PatternFill("solid", fgColor="FFD9D9D9")
CALC_FILL = PatternFill("solid", fgColor="FFFFFF00")   # amarelo
HDR_FONT  = Font(bold=True, size=9, name="Calibri")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DAT_FONT  = Font(size=9, name="Calibri")
DAT_ALIGN = Alignment(horizontal="left", vertical="center")
NUM_ALIGN = Alignment(horizontal="right", vertical="center")
BRL       = 'R$ #,##0.00'
PCT       = '0.00%'


def _g(line, attr, default=None):
    if hasattr(line, attr):
        return getattr(line, attr, default)
    try:
        return line[attr]
    except Exception:
        return default


def _row_from_line(line) -> list:
    id_smart = _g(line, "id_smart") or ""
    cpf_cnpj = id_smart.replace("ss_", "").replace("SS_", "")

    cred_kb    = _g(line, "credito_simcard_kb") or 0
    # Crédito corrigido: converte KB → MB (valor que o sistema usa para calcular excedente)
    cred_mb    = round(cred_kb / 1024, 4) if cred_kb else None
    franquia   = _g(line, "franquia_mb")   # Franquia do plano (vem da base, em MB)

    return [
        _g(line, "nome_pedido"),                            # A   1  Nome do pedido
        _g(line, "id_pedido"),                              # B   2  ID do pedido
        _fmt_date(_g(line, "data_ativacao")),               # C   3  Data de ativação
        _g(line, "preco_ativacao"),                         # D   4  Preço de ativação
        _g(line, "nome_contrato"),                          # E   5  Nome do contrato
        _g(line, "id_contrato"),                            # F   6  ID do contrato
        _fmt_date(_g(line, "data_cancelamento")),           # G   7  Data de cancelamento
        id_smart,                                           # H   8  ID_CPF/CNPJ
        cpf_cnpj,                                           # I   9  CPF/CNPJ
        _g(line, "client_nome"),                            # J   10 Nome do cliente
        franquia,                                           # K   11 Franquia (MB)
        _g(line, "iccid"),                                  # L   12 ICCID
        _g(line, "preco_mb_excedente"),                     # M   13 Preço do MB Excedente
        _g(line, "msisdn"),                                 # N   14 MSISDN
        _g(line, "operadora"),                              # O   15 Operadora
        _fmt_date(_g(line, "data_fim_pre_ativacao")),       # P   16 Data fim da pré-ativação
        _g(line, "mensalidade_base"),                       # Q   17 Mensalidade
        cred_mb,                                            # R   18 Crédito Simcard (MB — corrigido)
        _g(line, "status"),                                 # S   19 Status
        _g(line, "operadora_especifica"),                   # T   20 Operadora específica
        _fmt_date(_g(line, "data_fim_suspensao")),          # U   21 Data de término da suspensão
        _fmt_date(_g(line, "data_inicio_suspensao")),       # V   22 Data de início da suspensão
        _g(line, "consumo_total_kb"),                       # W   23 Consumo total (KB)
        _g(line, "reajuste_pct"),                           # X   24 Reajuste (%)
        _g(line, "mensalidade_reaj"),                       # Y   25 Reajuste 2025
        _g(line, "excedente_cobrado"),                      # Z   26 Excedente
        _g(line, "multa_cobrada"),                          # AA  27 Multa Cancelamento
        _g(line, "sms_cobrado"),                            # AB  28 SMS
        _g(line, "dias"),                                   # AC  29 Dias
        _g(line, "mensalidade_cobrada"),                    # AD  30 Mensalidade (calculada)
        _g(line, "ativacao_cobrada"),                       # AE  31 Ativação
        _g(line, "total_linha"),                            # AF  32 TOTAL
    ]


def generate_faturamento_excel(cycle, lines) -> io.BytesIO:
    """
    Gera Excel com 32 colunas no formato modelo.
    `lines` pode ser lista de ORM objects, RowMapping ou iterável de cursor.
    Usa write_only + singletons de estilo para suportar 692k+ linhas sem timeout.
    """
    from openpyxl.cell import WriteOnlyCell

    buf = io.BytesIO()
    wb  = Workbook(write_only=True)
    ws  = wb.create_sheet(f"{MESES[cycle.month]} {cycle.year}")

    # Singletons de estilo — evita criar milhões de objetos redundantes
    _fn   = Font(size=9, name="Calibri")
    _al_l = Alignment(horizontal="left",  vertical="center")
    _al_r = Alignment(horizontal="right", vertical="center")
    _yel  = PatternFill("solid", fgColor="FFFFF0CC")

    # Cabeçalho
    hdr_cells = []
    for col_idx, h in enumerate(HEADERS_32, 1):
        c = WriteOnlyCell(ws, value=h)
        c.fill      = CALC_FILL if col_idx in _CALC_COLS else HDR_FILL
        c.font      = HDR_FONT
        c.alignment = HDR_ALIGN
        hdr_cells.append(c)
    ws.append(hdr_cells)
    ws.row_dimensions[1].height = 30

    # Dados — WriteOnlyCell só para colunas que precisam de formatação numérica/monetária.
    # Células simples usam valor direto (sem objeto) — reduz criações de 22M → ~7M.
    for line in lines:
        row_vals  = _row_from_line(line)
        row_cells = []
        for col_idx, val in enumerate(row_vals, 1):
            is_calc  = col_idx in _CALC_COLS
            is_money = col_idx in _MONEY_COLS
            is_pct   = col_idx in _PCT_COLS
            is_right = col_idx in _INT_COLS or col_idx == 29

            if is_money or is_pct or is_calc or is_right:
                c = WriteOnlyCell(ws, value=val)
                c.font = _fn
                if is_calc:
                    c.fill = _yel
                if is_money and val is not None:
                    c.number_format = BRL
                    c.alignment     = _al_r
                elif is_pct and val is not None:
                    c.number_format = PCT
                    c.alignment     = _al_r
                elif is_right:
                    c.alignment = _al_r
                else:
                    c.alignment = _al_l
                row_cells.append(c)
            else:
                row_cells.append(val)   # valor direto — sem objeto Cell
        ws.append(row_cells)

    # Larguras das 32 colunas (A–AF)
    widths = [
        22,  # A   Nome do pedido
        18,  # B   ID do pedido
        14,  # C   Data de ativação
        14,  # D   Preço de ativação
        22,  # E   Nome do contrato
        18,  # F   ID do contrato
        14,  # G   Data de cancelamento
        22,  # H   ID_CPF/CNPJ
        18,  # I   CPF/CNPJ
        30,  # J   Nome do cliente
        12,  # K   Franquia (MB)
        24,  # L   ICCID
        18,  # M   Preço do MB Excedente
        18,  # N   MSISDN
        16,  # O   Operadora
        18,  # P   Data fim pré-ativação
        14,  # Q   Mensalidade
        14,  # R   Crédito Simcard (MB)
        20,  # S   Status
        20,  # T   Operadora específica
        22,  # U   Data término suspensão
        22,  # V   Data início suspensão
        16,  # W   Consumo total (KB)
        10,  # X   Reajuste
        14,  # Y   Reajuste 2025
        12,  # Z   Excedente
        16,  # AA  Multa Cancelamento
        10,  # AB  SMS
        8,   # AC  Dias
        14,  # AD  Mensalidade (calc)
        12,  # AE  Ativação
        14,  # AF  TOTAL
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(buf)
    buf.seek(0)
    return buf


def generate_client_lines_excel(cycle, lines, summary, client_name: str = "") -> io.BytesIO:
    """Faturamento Cliente — delega ao gerador de 29 colunas."""
    nome = client_name or ""
    if "|" in nome:
        nome = nome.split("|", 1)[0].strip()

    class _Wrapped:
        def __init__(self, ln):
            self._ln = ln
            self.client_nome = nome
        def __getattr__(self, attr):
            return getattr(self._ln, attr)

    return generate_faturamento_excel(cycle, [_Wrapped(ln) for ln in lines])


def generate_client_excel_fast(cycle, lines) -> io.BytesIO:
    """
    Versão otimizada para clientes com muitas linhas usando xlsxwriter.
    ~15x mais rápido que openpyxl para 30k+ linhas.
    """
    import xlsxwriter

    buf = io.BytesIO()
    wb  = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_numbers': True})
    ws  = wb.add_worksheet(f"{MESES[cycle.month]} {cycle.year}")

    # Formatos
    hdr_gray = wb.add_format({
        'bold': True, 'font_name': 'Calibri', 'font_size': 9,
        'bg_color': '#D9D9D9', 'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
        'border': 1,
    })
    hdr_yel = wb.add_format({
        'bold': True, 'font_name': 'Calibri', 'font_size': 9,
        'bg_color': '#FFFF00', 'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
        'border': 1,
    })
    fmt_brl  = wb.add_format({'font_name': 'Calibri', 'font_size': 9, 'num_format': 'R$ #,##0.00', 'align': 'right'})
    fmt_pct  = wb.add_format({'font_name': 'Calibri', 'font_size': 9, 'num_format': '0.00%',       'align': 'right'})
    fmt_int  = wb.add_format({'font_name': 'Calibri', 'font_size': 9, 'align': 'right'})
    fmt_def  = wb.add_format({'font_name': 'Calibri', 'font_size': 9})

    # Mapeamento coluna → formato (0-indexed)
    _WIDTHS = [22,18,14,14,22,18,14,22,18,30,12,24,18,18,16,18,14,14,20,20,22,22,16,10,14,12,16,10,8,14,12,14]
    for i, w in enumerate(_WIDTHS):
        col = i
        ws.set_column(col, col, w)

    # Cabeçalho (linha 0)
    for col_idx, h in enumerate(HEADERS_32):
        fmt = hdr_yel if (col_idx + 1) in _CALC_COLS else hdr_gray
        ws.write(0, col_idx, h, fmt)
    ws.set_row(0, 30)

    # Formato por coluna (0-indexed) para dados
    _COL_FMT = []
    for i in range(1, 33):
        if i in _MONEY_COLS:
            _COL_FMT.append(fmt_brl)
        elif i in _PCT_COLS:
            _COL_FMT.append(fmt_pct)
        elif i in _INT_COLS or i == 29:
            _COL_FMT.append(fmt_int)
        else:
            _COL_FMT.append(fmt_def)

    # Dados (a partir da linha 1)
    for row_idx, line in enumerate(lines, 1):
        vals = _row_from_line(line)
        for col_idx, (val, fmt) in enumerate(zip(vals, _COL_FMT)):
            if val is None:
                ws.write_blank(row_idx, col_idx, None, fmt)
            elif isinstance(val, str):
                ws.write_string(row_idx, col_idx, val, fmt)
            else:
                ws.write_number(row_idx, col_idx, float(val) if val is not None else 0, fmt)

    wb.close()
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════════
#  Resumo por cliente (uma linha por cliente)
# ════════════════════════════════════════════════════════════════

NAVY      = "FF0E2841"
VERDE     = "FF3CB54A"
VERDE_ESC = "FF196B24"
CINZA_Z   = "FFFAFAFA"
BRANCO    = "FFFFFFFF"
PRETO     = "FF0E2841"
BORDA_COR = "FFCCCCCC"
BRL_FMT   = 'R$ #,##0.00'
INT_FMT   = '#,##0'


def _fill(h):
    return PatternFill("solid", fgColor=h)

def _bdr():
    s = Side(style="thin", color=BORDA_COR)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, color=PRETO, size=9, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _col_w(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _hdr(ws, row, col, text, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=BRANCO, size=9)
    c.alignment = _al(align); c.border = _bdr()
    ws.row_dimensions[row].height = 18
    return c

def _dat(ws, row, col, val, bg_hex=BRANCO, align="left", fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _fill(bg_hex); c.font = _font(bold=bold, color=PRETO, size=9)
    c.alignment = _al(align); c.border = _bdr()
    if fmt:
        c.number_format = fmt
    ws.row_dimensions[row].height = 14
    return c


def generate_billing_excel(cycle, summaries, breakdown_rows=None) -> io.BytesIO:
    """Resumo por cliente — uma linha por cliente."""
    buf = io.BytesIO()
    wb  = Workbook()
    mes = f"{MESES[cycle.month]} {cycle.year}"

    ws = wb.active
    ws.title = "DIN"
    ws.sheet_view.showGridLines = False

    COLS = ["ID SMART", "CNPJ / CPF", "Qtd Ativas", "Cancelamentos",
            "Suspensões", "Mensalidade", "Ativação", "Excedente",
            "Multa", "SMS", "Frete", "Mensageria", "Ajustes", "TOTAL FINAL", "Vencimento"]
    NC = len(COLS)

    def _title_row(ws, row, text, ncols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = _fill(NAVY); c.font = _font(bold=True, color=BRANCO, size=13)
        c.alignment = _al("left"); ws.row_dimensions[row].height = 26

    def _sub_row(ws, row, text, ncols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = _fill(VERDE); c.font = _font(bold=True, color=BRANCO, size=9)
        c.alignment = _al("left"); ws.row_dimensions[row].height = 16

    total_geral = sum(s.total_final or 0 for s in summaries)
    _title_row(ws, 1, f"GESTORA SMART — FATURAMENTO {mes.upper()}", NC)
    _sub_row(ws, 2, f"  Total: R$ {total_geral:,.2f}   |   {len(summaries)} clientes   |   Período: {cycle.month:02d}/{cycle.year}", NC)

    for col, h in enumerate(COLS, 1):
        _hdr(ws, 3, col, h, align="right" if col > 2 else "left")

    summaries_sorted = sorted(summaries, key=lambda s: s.total_final or 0, reverse=True)

    for i, s in enumerate(summaries_sorted):
        row = 4 + i
        bg  = CINZA_Z if i % 2 == 1 else BRANCO
        cnpj = (s.id_smart or "").replace("ss_", "").replace("SS_", "")
        due  = ""
        if s.due_date:
            try:
                due = s.due_date.strftime("%d/%m/%Y") if hasattr(s.due_date, "strftime") else str(s.due_date)[:10]
            except Exception:
                due = str(s.due_date)[:10]

        vals   = [s.id_smart or "", cnpj, s.qtd_linhas_ativas or 0, s.qtd_cancelamentos or 0,
                  s.qtd_suspensoes or 0, s.total_mensalidade or 0, s.total_ativacao or 0,
                  s.total_excedente or 0, s.total_multa or 0, s.total_sms or 0,
                  s.total_frete or 0, s.total_mensageria or 0, s.total_ajustes or 0,
                  s.total_final or 0, due]
        fmts   = [None, None, INT_FMT, INT_FMT, INT_FMT, BRL_FMT, BRL_FMT, BRL_FMT,
                  BRL_FMT, BRL_FMT, BRL_FMT, BRL_FMT, BRL_FMT, BRL_FMT, None]
        aligns = ["left","left","right","right","right","right","right","right",
                  "right","right","right","right","right","right","center"]

        for col, (val, fmt, al) in enumerate(zip(vals, fmts, aligns), 1):
            _dat(ws, row, col, val, bg_hex=bg, align=al, fmt=fmt)

    tot_row = 4 + len(summaries)
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)
    c = ws.cell(row=tot_row, column=1, value="TOTAL GERAL")
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=BRANCO, size=10)
    c.alignment = _al("center"); c.border = _bdr()

    tot_attrs = {6:"total_mensalidade", 7:"total_ativacao", 8:"total_excedente",
                 9:"total_multa", 10:"total_sms", 11:"total_frete",
                 12:"total_mensageria", 13:"total_ajustes", 14:"total_final"}
    for col, attr in tot_attrs.items():
        val = sum(getattr(s, attr, 0) or 0 for s in summaries)
        c2 = ws.cell(row=tot_row, column=col, value=val)
        c2.fill = _fill(VERDE_ESC if col == 14 else NAVY)
        c2.font = _font(bold=True, color=BRANCO, size=9)
        c2.alignment = _al("right"); c2.number_format = BRL_FMT; c2.border = _bdr()
    ws.row_dimensions[tot_row].height = 18

    ws.freeze_panes = "A4"
    _col_w(ws, [22, 18, 10, 13, 11, 14, 12, 12, 10, 10, 10, 12, 11, 15, 13])

    wb.save(buf)
    buf.seek(0)
    return buf
