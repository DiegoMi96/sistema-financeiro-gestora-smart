"""
BillingEngineService — motor de faturamento.
Recebe caminhos de arquivo em disco e retorna DataFrames.
"""
import gc
import os
import pandas as pd
import numpy as np
import io
import re
import math
import calendar
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")


def _sanitize_id(x):
    if pd.isna(x): return None
    d = re.sub(r"\D", "", str(x))
    return f"ss_{d}" if d else None


def _kb_to_mb(v):
    if not v or v == 0: return 0.0
    x = float(v)
    while x >= 1024: x /= 1024
    return x


def _roundup2(x): return math.ceil(round(x * 100, 6)) / 100


def _normalizar_status(s):
    if not isinstance(s, str): return s
    # Qualquer "Aguardando..." → Ativo (mês cheio)
    if s.startswith("Aguardando"): return "Ativo"
    return {"Cancelado": "Cancelamento"}.get(s, s)


# ── Corte de elegibilidade do reajuste anual ──────────────────────────────────
# Só recebe o % da planilha "Base de Reajuste" quem tem "Data de ativação"
# ATÉ dezembro do ano abaixo (ou sem data de ativação); clientes ativados
# depois ficam com reajuste 0% neste ciclo. Confirmado com o Diego
# (01/09/2026): reajuste 2025 considerou até dez/2024 (REAJUSTE_ANO_CORTE=2025,
# já rodado); reajuste 2026 passa a considerar até dez/2026 — precisa mudar
# este número de novo antes do reajuste de 2027.
REAJUSTE_ANO_CORTE = 2027  # exclusivo: ano de ativação < este valor recebe reajuste


# ── IDs/CPFs excluídos do faturamento — fallback hardcoded ────────────────────
# Usado quando o banco não tem configuração salva ainda.
CNPJ_EXCLUIDOS_DEFAULT = {
    "22222222222",
    "24152616000146",
    "24283777000179",
    "10119014475",
    "42604215810",
    "56388853072",
    "60687193834",
    "77268784104",
}


class BillingEngineService:

    def __init__(self, year: int, month: int,
                 cnpj_excluidos: set | None = None,
                 mensageria_valor: float | None = None,
                 cnpj_sem_arredondamento: set | None = None):
        self.year            = year
        self.month           = month
        self.mes_ref         = datetime(year, month, 1)
        self.total_dias      = calendar.monthrange(year, month)[1]
        self.cnpj_excluidos  = cnpj_excluidos if cnpj_excluidos is not None else CNPJ_EXCLUIDOS_DEFAULT
        self.mensageria_valor = mensageria_valor if mensageria_valor is not None else 9.90
        # Clientes cujo Mensalidade cobrada usa arredondamento padrão (ROUND)
        # em vez do ROUNDUP normal do sistema — configurável em Configurações.
        self.cnpj_sem_arredondamento = cnpj_sem_arredondamento or set()

    def _is_excluido(self, cnpj_str: str) -> bool:
        digits = re.sub(r"\D", "", str(cnpj_str))
        return digits in self.cnpj_excluidos

    def _sem_arredondamento(self, id_smart: str) -> bool:
        digits = re.sub(r"\D", "", str(id_smart))
        return digits in self.cnpj_sem_arredondamento

    # ── API de processamento em streaming (baixa RAM) ─────────────────────────

    def _excel_to_csv(self, source, csv_path: str) -> None:
        """
        Lê o Excel base linha a linha (openpyxl read_only) e grava como CSV
        em disco — sem nunca montar a planilha inteira em memória.

        Trocado de calamine pra streaming em 04/08/2026: a base desse mês
        tem XML interno de ~1,5 GB (58 colunas — a maioria telemetria não
        usada pelo motor, tipo GPS/última conexão/hostname) e travava o
        processo mesmo com >3,8 GB de RAM disponível (confirmado testando
        localmente com o arquivo real — calamine monta a planilha inteira
        de uma vez, então o pico de memória escala com o tamanho do XML,
        não só com o número de linhas). O servidor de produção só tem
        1,9 GB de RAM no total. openpyxl read_only processa uma linha por
        vez — bem mais lento que calamine, mas memória praticamente
        constante (não escala com o tamanho do arquivo).
        """
        from openpyxl import load_workbook
        import csv as _csv

        file_input = io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
        print("📂 Lendo base (streaming, linha a linha)...", flush=True)

        wb = load_workbook(file_input, read_only=True, data_only=True)
        ws = wb["Inventário"] if "Inventário" in wb.sheetnames else wb[wb.sheetnames[0]]
        if "Inventário" not in wb.sheetnames:
            print("⚠️ Aba 'Inventário' não encontrada, usando primeira aba", flush=True)

        rows_iter = ws.iter_rows(values_only=True)
        header = list(next(rows_iter))

        n = 0
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = _csv.writer(f)
            writer.writerow(header)
            batch = []
            for row in rows_iter:
                batch.append(row)
                n += 1
                if len(batch) >= 50_000:
                    writer.writerows(batch)
                    batch = []
                    print(f"📂 {n:,} linhas convertidas...", flush=True)
            if batch:
                writer.writerows(batch)

        wb.close()
        gc.collect()
        print(f"📂 {n:,} linhas gravadas em CSV — processamento em chunks a seguir", flush=True)

    def _prepare_base_chunk(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aplica conversões de tipo a um chunk raw lido do CSV (substitui _load_base para streaming)."""
        _date_cols = [
            "Data de ativação", "Data de início do bloqueio de rede", "Data de cancelamento",
            "Data de início da suspensão", "Data de término da suspensão", "Data fim da pré-ativação",
        ]
        _num_cols = ["Mensalidade", "Preço de ativação", "Preço do MB Excedente", "Crédito adicionado no Simcard"]
        _all_cols = [
            "Nome do cliente", "CPF/CNPJ", "MSISDN", "ICCID", "Status",
            "Nome do pedido", "ID do pedido", "Nome do contrato", "ID do contrato",
            "Reajuste", "Franquia (MB)", "Operadora", "Operadora específica", "Consumo total (KB)",
        ] + _date_cols + _num_cols

        for c in _all_cols:
            if c not in df.columns:
                df[c] = None
        for c in _date_cols:
            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
        for c in _num_cols:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

        df["ID_CPF/CNPJ"] = df["CPF/CNPJ"].apply(_sanitize_id)
        df["Status"]      = df["Status"].apply(_normalizar_status)
        return df

    def setup(self, file_paths: dict, base_bytes: bytes | None = None) -> dict:
        """
        Carrega todos os arquivos de referência (pequenos) e converte a base Excel → CSV em disco.
        Pico de RAM: ~830 MB durante a leitura da base; cai para ~0 após gravar o CSV.
        Retorna ref_data com csv_path e todos os mapas de referência.
        """
        def _read(key) -> bytes | None:
            path = file_paths.get(key)
            if not path:
                return None
            with open(path, "rb") as fh:
                return fh.read()

        print("📂 Carregando arquivos de referência...", flush=True)
        reajuste_map        = self._load_reajuste(_read("reajuste"))   if file_paths.get("reajuste")   else {}
        cancel_prop, ativ_prop = self._load_atencao(_read("atencao")) if file_paths.get("atencao")    else (set(), set())
        df_cancel           = self._load_cancelamentos(_read("cancelamentos"))
        df_fretes_raw       = self._load_fretes(_read("fretes"))       if file_paths.get("fretes")     else pd.DataFrame()
        venc_map            = self._load_vencimentos(_read("vencimentos"))
        sms_map             = self._load_sms(_read("sms"))             if file_paths.get("sms")        else {}
        df_men_raw          = self._load_mensageria(_read("mensageria")) if file_paths.get("mensageria") else pd.DataFrame()

        df_fretes_fmt = self._montar_fretes(df_fretes_raw)
        df_men_fmt    = self._montar_mensageria(df_men_raw) if not df_men_raw.empty else pd.DataFrame()

        multa_map = (
            df_cancel[["ICCID", "VALOR DA MULTA"]]
            .dropna(subset=["ICCID"])
            .assign(ICCID=lambda x: x["ICCID"].astype(str).str.strip())
            .set_index("ICCID")["VALOR DA MULTA"].to_dict()
        ) if not df_cancel.empty else {}

        # Converte base Excel → CSV em disco (pico 830 MB, depois liberado)
        base_path = file_paths.get("base")
        csv_path  = base_path + "_stream.csv"
        source    = base_bytes if base_bytes is not None else base_path
        self._excel_to_csv(source, csv_path)

        # Scan rápido para ICCIDs de Cancelamento (deduplicação entre chunks)
        cancel_iccids_base: set = set()
        try:
            for sc in pd.read_csv(csv_path, usecols=["ICCID", "Status"],
                                  chunksize=50_000, dtype=str, low_memory=False):
                sc["Status"] = sc["Status"].fillna("").apply(_normalizar_status)
                icc = sc.loc[sc["Status"] == "Cancelamento", "ICCID"].astype(str).str.strip()
                cancel_iccids_base.update(v for v in icc if v and v not in ("", "nan"))
        except Exception as exc:
            print(f"⚠️ Scan dedup falhou (ignorado): {exc}", flush=True)
        print(f"📂 ICCIDs Cancelamento para dedup: {len(cancel_iccids_base)}", flush=True)

        return {
            "csv_path":           csv_path,
            "reajuste_map":       reajuste_map,
            "cancel_prop":        cancel_prop,
            "ativ_prop":          ativ_prop,
            "df_cancel":          df_cancel,
            "df_fretes":          df_fretes_fmt,
            "df_mensageria":      df_men_fmt,
            "venc_map":           venc_map,
            "sms_map":            sms_map,
            "multa_map":          multa_map,
            "cancel_iccids_base": cancel_iccids_base,
        }

    def process_chunk(self, df_raw: pd.DataFrame, ref: dict) -> pd.DataFrame:
        """
        Aplica conversões de tipo, dedup e cálculos de faturamento a um chunk CSV.
        Retorna apenas linhas principais (não anuidade).
        """
        df = self._prepare_base_chunk(df_raw)

        # Dedup: remove Ativo se o mesmo ICCID aparece como Cancelamento em qualquer chunk
        cancel_iccids_base = ref.get("cancel_iccids_base", set())
        if cancel_iccids_base and "ICCID" in df.columns:
            icc_str  = df["ICCID"].astype(str).str.strip()
            mask_dup = (df["Status"] == "Ativo") & icc_str.isin(cancel_iccids_base)
            n_dup    = int(mask_dup.sum())
            if n_dup:
                print(f"⚠️ {n_dup} linha(s) Ativo duplicada(s) — removidas", flush=True)
                df = df[~mask_dup].copy()

        mask_excluido = df["CPF/CNPJ"].astype(str).str.strip().apply(self._is_excluido)
        def _starts_anuidade(s): return str(s).upper().startswith("ANUIDADE")
        mask_anuidade = (
            df["Nome do cliente"].astype(str).apply(_starts_anuidade)
            | df.get("Nome do pedido",  pd.Series("", index=df.index)).astype(str).apply(_starts_anuidade)
            | df.get("Nome do contrato", pd.Series("", index=df.index)).astype(str).apply(_starts_anuidade)
        )

        df_principal = df[~mask_excluido & ~mask_anuidade].copy()
        del df

        df_p = self._calcular(
            df_principal,
            ref["reajuste_map"], ref["cancel_prop"], ref["ativ_prop"],
            ref["df_cancel"],    ref["sms_map"],
        )
        del df_principal
        return df_p

    # ── API legada (mantida para compatibilidade / uso local) ──────────────────

    def run(self, file_paths: dict, base_bytes: bytes | None = None) -> dict:
        """
        file_paths: caminhos em disco para cada arquivo.
        base_bytes: bytes da base pré-lidos no request handler (evita I/O lento na VM Docker).
                    Se fornecido, calamine processa diretamente em memória.
        """
        def _read(key) -> bytes | None:
            path = file_paths.get(key)
            if not path:
                return None
            with open(path, "rb") as f:
                return f.read()

        # ── Carregar ───────────────────────────────────────────
        base_path = file_paths.get("base")
        base_size = os.path.getsize(base_path) / 1024 / 1024 if base_path else 0
        print(f"📂 Carregando base ({base_size:.1f} MB)...", flush=True)
        df_base       = self._load_base(base_bytes if base_bytes is not None else base_path)
        base_bytes    = None  # libera bytes crus imediatamente após carregar em DataFrame
        gc.collect()
        print(f"📂 Base: {len(df_base)} linhas", flush=True)

        # Deduplicar: ICCID que aparece como Ativo e Cancelamento → manter só Cancelamento
        _iccid_col = df_base.get("ICCID")
        if _iccid_col is not None:
            _iccid_str = _iccid_col.astype(str).str.strip()
            _has_iccid = _iccid_str.str.len() > 0
            _iccids_cancel = set(_iccid_str[_has_iccid & (df_base["Status"] == "Cancelamento")])
            if _iccids_cancel:
                _mask_ativo_dup = (df_base["Status"] == "Ativo") & _has_iccid & _iccid_str.isin(_iccids_cancel)
                n_dup = int(_mask_ativo_dup.sum())
                if n_dup:
                    print(f"⚠️ {n_dup} linha(s) Ativo duplicada(s) com Cancelamento — removidas", flush=True)
                    df_base = df_base[~_mask_ativo_dup].copy()
        reajuste_map  = self._load_reajuste(_read("reajuste")) if file_paths.get("reajuste") else {}
        print(f"📂 Reajuste: {len(reajuste_map)} entradas", flush=True)
        cancel_prop, ativ_prop = self._load_atencao(_read("atencao")) if file_paths.get("atencao") else (set(), set())
        print(f"📂 Atenção: {len(cancel_prop)} cancel, {len(ativ_prop)} ativ", flush=True)
        df_cancel     = self._load_cancelamentos(_read("cancelamentos"))
        print(f"📂 Cancelamentos: {len(df_cancel)} linhas", flush=True)
        df_fretes     = self._load_fretes(_read("fretes")) if file_paths.get("fretes") else pd.DataFrame()
        print(f"📂 Fretes: {len(df_fretes)} linhas", flush=True)
        venc_map      = self._load_vencimentos(_read("vencimentos"))
        print(f"📂 Vencimentos: {len(venc_map)} entradas", flush=True)
        sms_map       = self._load_sms(_read("sms")) if file_paths.get("sms") else {}
        print(f"📂 SMS: {len(sms_map)} entradas", flush=True)

        # ── Separar grupos ─────────────────────────────────────
        mask_excluido = df_base["CPF/CNPJ"].astype(str).str.strip().apply(self._is_excluido)
        def _starts_anuidade(s): return str(s).upper().startswith("ANUIDADE")
        mask_anuidade = (
            df_base["Nome do cliente"].astype(str).apply(_starts_anuidade)
            | df_base.get("Nome do pedido",  pd.Series("", index=df_base.index)).astype(str).apply(_starts_anuidade)
            | df_base.get("Nome do contrato", pd.Series("", index=df_base.index)).astype(str).apply(_starts_anuidade)
        )
        df_principal  = df_base[~mask_excluido & ~mask_anuidade].copy()
        df_anuidade   = df_base[mask_anuidade & ~mask_excluido].copy()

        # da_map extraído ANTES de liberar df_base
        _iccids = df_base.get("ICCID", pd.Series(dtype=str)).astype(str).str.strip()
        _das    = pd.to_datetime(df_base.get("Data de ativação", pd.Series(dtype=str)), errors="coerce")
        da_map  = {icc: da for icc, da in zip(_iccids, _das) if icc and icc != "nan"}
        del df_base, _iccids, _das
        gc.collect()

        # ── Calcular — libera cada fonte logo após calcular ─────
        df_p = self._calcular(df_principal, reajuste_map, cancel_prop, ativ_prop, df_cancel, sms_map)
        del df_principal
        gc.collect()

        df_a = self._calcular(df_anuidade,  reajuste_map, cancel_prop, ativ_prop, df_cancel, sms_map)
        del df_anuidade
        gc.collect()

        # ── Linhas adicionais ───────────────────────────────────
        df_cr = self._montar_cancelamentos(df_cancel, reajuste_map, cancel_prop, sms_map, da_map)
        df_fr = self._montar_fretes(df_fretes)

        # ── Mensageria ─────────────────────────────────────────
        if file_paths.get("mensageria"):
            df_men_raw = self._load_mensageria(_read("mensageria"))
            df_men     = self._montar_mensageria(df_men_raw)
        else:
            df_men = pd.DataFrame()

        # ── Boletos ────────────────────────────────────────────
        # df_cr não entra em all_rows: as linhas de Cancelamento já estão em df_p
        # (vindas da base com Status="Cancelamento"), a multa é aplicada via multa_map.
        # Incluir df_cr causaria duplicidade de linhas e valores.
        all_rows = pd.concat([df_p, df_fr, df_men], ignore_index=True)
        df_bill  = all_rows[
            all_rows["ID_CPF/CNPJ"].notna()
            & ~all_rows["CPF/CNPJ"].astype(str).str.strip().apply(self._is_excluido)
            & ~all_rows.get("Nome do cliente", pd.Series(dtype=str)).astype(str).str.upper().str.startswith("ANUIDADE", na=False)
        ]
        boletos = df_bill.groupby("ID_CPF/CNPJ").agg(
            valor=("_total", "sum")
        ).reset_index().rename(columns={"ID_CPF/CNPJ": "Cliente"})
        boletos["valor"]      = boletos["valor"].round(2)
        boletos["vencimento"] = boletos["Cliente"].map(venc_map)
        boletos = boletos.sort_values("valor", ascending=False)

        return {
            "df_inventario":  df_p,
            "df_anuidade":    df_a,
            "df_cancel_rows": df_cr,
            "df_fretes":      df_fr,
            "df_mensageria":  df_men,
            "boletos":        boletos,
        }

    # ── Loaders ───────────────────────────────────────────────

    def _load_base(self, source) -> pd.DataFrame:
        """source: bytes (preferencial, sem I/O) ou str (caminho em disco).
        Usa python-calamine (Rust) — 10-50x mais rápido que openpyxl para 700k linhas.
        """
        _cols = [
            "Nome do cliente", "CPF/CNPJ", "MSISDN", "ICCID", "Status",
            "Nome do pedido", "ID do pedido",
            "Nome do contrato", "ID do contrato",
            "Data de ativação",
            "Data de início do bloqueio de rede",
            "Data de cancelamento",
            "Data de início da suspensão",
            "Data de término da suspensão",
            "Data fim da pré-ativação",
            "Mensalidade", "Preço de ativação",
            "Preço do MB Excedente", "Crédito adicionado no Simcard",
            "Reajuste",
            "Franquia (MB)", "Operadora", "Operadora específica",
            "Consumo total (KB)",
        ]

        file_input = io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source

        print("📂 Carregando base com calamine...", flush=True)
        try:
            df = pd.read_excel(file_input, sheet_name="Inventário", engine="calamine")
        except Exception:
            if hasattr(file_input, "seek"):
                file_input.seek(0)
            df = pd.read_excel(file_input, sheet_name=0, engine="calamine")
            print(f"⚠️ Aba 'Inventário' não encontrada, usando primeira aba", flush=True)

        print(f"📂 {len(df):,} linhas carregadas", flush=True)

        # Garante que todas as colunas esperadas existam
        for c in _cols:
            if c not in df.columns:
                df[c] = None

        df = df[_cols].copy()

        for c in ["Data de ativação", "Data de início do bloqueio de rede", "Data de cancelamento",
                  "Data de início da suspensão", "Data de término da suspensão", "Data fim da pré-ativação"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
        for c in ["Mensalidade", "Preço de ativação", "Preço do MB Excedente", "Crédito adicionado no Simcard"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

        df["ID_CPF/CNPJ"] = df["CPF/CNPJ"].apply(_sanitize_id)
        df["Status"]      = df["Status"].apply(_normalizar_status)
        return df

    def _load_reajuste(self, data: bytes) -> dict:
        try:
            xl = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")
            # Tenta encontrar a aba certa por nome (flexível)
            sheet = next(
                (s for s in xl.sheet_names if "reajuste" in s.lower() or s.lower() == "planilha1"),
                xl.sheet_names[0]
            )
            df = xl.parse(sheet)
            # Coluna ID: primeira coluna que contenha ss_ ou seja a primeira
            col_id = df.columns[0]
            # Coluna Reajuste: busca por nome
            col_reaj = next((c for c in df.columns if "reajuste" in str(c).lower()), None)
            if col_reaj is None:
                return {}
            df = df[[col_id, col_reaj]].dropna(subset=[col_id])
            df[col_id]   = df[col_id].astype(str).str.strip()
            df[col_reaj] = pd.to_numeric(df[col_reaj], errors="coerce").fillna(0)
            return dict(zip(df[col_id], df[col_reaj]))
        except Exception:
            return {}

    def _load_atencao(self, data: bytes) -> tuple:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Cancelamento e Suspenção", engine="openpyxl", header=0)
        while len(df.columns) < 3:
            df[f"_c{len(df.columns)}"] = None
        df.columns = ["cnpj", "id", "status"] + [f"_c{i}" for i in range(len(df.columns)-3)]
        df = df.dropna(subset=["id"])
        df["id"]     = df["id"].astype(str).str.strip()
        df["status"] = df["status"].fillna("").astype(str)
        cancel_set = set(df[df["status"].str.contains("Cancelamento", na=False)]["id"])
        ativ_set   = set(df[df["status"].str.contains("Ativa", na=False, case=False)]["id"])
        return cancel_set, ativ_set

    def _load_cancelamentos(self, data: bytes) -> pd.DataFrame:
        # Formato antigo: abas "Cancelamento" e "Desistência"
        dfs = []
        for sheet in ["Cancelamento", "Desistência"]:
            try:
                d = pd.read_excel(io.BytesIO(data), sheet_name=sheet, engine="openpyxl", dtype={"ICCID": str})
                d["_tipo"] = sheet
                dfs.append(d)
            except Exception:
                pass
        if dfs:
            df = pd.concat(dfs, ignore_index=True)
            # Cabeçalho da planilha de cancelamentos já mudou de caixa entre
            # meses (ex.: "Valor da multa" em vez de "VALOR DA MULTA") e isso
            # derrubava o motor inteiro com um KeyError — e como o processamento
            # roda em background, o request volta 200 OK mesmo assim, então
            # ninguém via o erro (ver _run_billing_engine em billing.py). Agora
            # casa o nome da coluna ignorando maiúsculas/minúsculas e espaços.
            def _find_col(name: str):
                norm = {str(c).strip().lower(): c for c in df.columns}
                return norm.get(name.strip().lower())

            for canonical in ["VALOR DA MULTA", "Mensalidade", "Data de cancel", "ID"]:
                found = _find_col(canonical)
                if found is None:
                    raise ValueError(
                        f"Planilha de cancelamentos: coluna '{canonical}' não encontrada "
                        f"(colunas disponíveis: {list(df.columns)})"
                    )
                if found != canonical:
                    df = df.rename(columns={found: canonical})

            df["VALOR DA MULTA"] = pd.to_numeric(df["VALOR DA MULTA"], errors="coerce").fillna(0)
            df["Mensalidade"]    = pd.to_numeric(df["Mensalidade"], errors="coerce").fillna(0)
            df["Data de cancel"] = pd.to_datetime(df["Data de cancel"], errors="coerce", dayfirst=True)
            df["ID"] = df["ID"].astype(str).str.strip()
            return df

        # Formato novo: aba única (ex: "Junho-26") com colunas:
        # ID | cliente | Linha | ICCID | Operadora | Mensalidade | Data de cancel |
        # Data ativação | Vendedor | STATUS | Multa Cancelamento (label) | valor_multa (int col)
        try:
            raw = pd.read_excel(io.BytesIO(data), sheet_name=0, engine="openpyxl", dtype={"ICCID": str})
            if raw.empty:
                return pd.DataFrame()
            cols = list(raw.columns)
            # Coluna de valor numérico da multa: última coluna com nome inteiro
            multa_num_col = next((c for c in reversed(cols) if isinstance(c, int)), None)
            out = pd.DataFrame()
            out["ID"]            = raw[cols[0]].astype(str).str.strip()
            out["clientes"]      = raw[cols[1]] if len(cols) > 1 else None
            out["Linha"]         = raw["Linha"]         if "Linha"         in cols else None
            out["ICCID"]         = raw["ICCID"].astype(str).str.strip() if "ICCID" in cols else None
            out["Mensalidade"]   = pd.to_numeric(raw["Mensalidade"] if "Mensalidade" in cols else 0, errors="coerce").fillna(0)
            out["Data de cancel"]= pd.to_datetime(raw["Data de cancel"] if "Data de cancel" in cols else None, errors="coerce", dayfirst=True)
            out["Data ativação"] = pd.to_datetime(raw["Data ativação"] if "Data ativação" in cols else None, errors="coerce", dayfirst=True)
            out["VALOR DA MULTA"]= pd.to_numeric(raw[multa_num_col] if multa_num_col is not None else 0, errors="coerce").fillna(0)
            out["_tipo"]         = raw["STATUS"].fillna("Cancelamento") if "STATUS" in cols else "Cancelamento"
            out = out.dropna(subset=["ID"])
            return out
        except Exception:
            return pd.DataFrame()

    def _load_fretes(self, data: bytes) -> pd.DataFrame:
        # Formato antigo: aba "Resumo" com coluna "ID_CNPJCPF".
        try:
            df = pd.read_excel(io.BytesIO(data), sheet_name="Resumo", engine="openpyxl")
            df = df[["ID_CNPJCPF", "DESTINATARIO", "VALOR TOTAL"]].dropna(subset=["ID_CNPJCPF"])
            df.columns = ["id", "cliente", "valor"]
            df["id"]    = df["id"].astype(str).str.strip()
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0)
            return df
        except Exception:
            pass

        # Formato novo (ex.: exportação Allcom, aba "ALLCOM" — 04/08/2026):
        # sem aba "Resumo", coluna de identificação chama "CNPJ" em vez de
        # "ID_CNPJCPF". Casa por nome (ignorando maiúsculas/minúsculas) na
        # primeira aba do arquivo, em vez de exigir nome exato de aba/coluna.
        xl = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")
        raw = xl.parse(xl.sheet_names[0])
        norm = {str(c).strip().lower(): c for c in raw.columns}

        def _find(*names):
            for name in names:
                found = norm.get(name.strip().lower())
                if found is not None:
                    return found
            return None

        col_id    = _find("ID_CNPJCPF", "CNPJ", "ID_CNPJ", "CPF/CNPJ")
        col_cli   = _find("DESTINATARIO", "CLIENTE", "NOME")
        col_valor = _find("VALOR TOTAL", "VALOR")
        if col_id is None or col_valor is None:
            raise ValueError(
                f"Planilha de fretes: não encontrei as colunas de identificação/valor "
                f"(colunas disponíveis: {list(raw.columns)})"
            )

        df = pd.DataFrame()
        df["id"]     = raw[col_id].astype(str).str.strip()
        df["cliente"]= raw[col_cli] if col_cli is not None else None
        df["valor"]  = pd.to_numeric(raw[col_valor], errors="coerce").fillna(0)
        return df.dropna(subset=["id"])

    def _load_vencimentos(self, data: bytes) -> dict:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Planilha1", engine="openpyxl", header=0)
        mes_boleto = datetime(self.year if self.month < 12 else self.year+1,
                              self.month+1 if self.month < 12 else 1, 1)
        target = next((c for c in df.columns if hasattr(c, "year") and c.year == mes_boleto.year and c.month == mes_boleto.month), None)
        if not target: return {}
        result = {}
        for _, row in df.iterrows():
            id_cli = str(row[df.columns[0]]).strip()
            due    = row[target]
            if id_cli.startswith("ss_") and hasattr(due, "year"):
                result[id_cli] = due
        return result

    def _load_sms(self, data: bytes) -> dict:
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl", dtype={"MSISDN": str})
        if "MSISDN" not in df.columns: return {}
        df["MSISDN"]    = df["MSISDN"].astype(str).str.strip()
        df["QTD"]       = pd.to_numeric(df.get("QTD", 0), errors="coerce").fillna(0)
        df["VALOR UN"]  = pd.to_numeric(df.get("VALOR UN", 0.5), errors="coerce").fillna(0.5)
        df["v"]         = df["QTD"] * df["VALOR UN"]
        return df.groupby("MSISDN")["v"].sum().to_dict()

    def _load_mensageria(self, data: bytes) -> pd.DataFrame:
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl", dtype=str)
        id_col = next((c for c in df.columns if "id" in c.lower() or "cnpj" in c.lower()), df.columns[0])
        df = df[[id_col]].copy()
        df.columns = ["id_cliente"]
        return df.dropna(subset=["id_cliente"])

    # ── Cálculos ──────────────────────────────────────────────

    def _calcular(self, df, reajuste_map, cancel_prop, ativ_prop, df_cancel, sms_map):
        if df.empty: return df

        # Sem .copy() — o chamador já passou uma cópia; modifica in-place para economizar RAM

        # Pré-ativo → promover para Ativo se tiver consumo real, data_ativacao ou SMS
        _mask_pre = df["Status"] == "Pré-ativo"
        if _mask_pre.any():
            _da_tmp      = pd.to_datetime(df.get("Data de ativação"), errors="coerce")
            _con_tmp     = pd.to_numeric(df.get("Crédito adicionado no Simcard", 0), errors="coerce").fillna(0)
            _consumo_kb  = pd.to_numeric(df.get("Consumo total (KB)", 0), errors="coerce").fillna(0)
            _sms_tmp     = (
                df["MSISDN"].astype(str).str.strip().isin(sms_map)
                if sms_map and "MSISDN" in df.columns
                else pd.Series(False, index=df.index)
            )
            df.loc[_mask_pre & ((_con_tmp > 0) | (_consumo_kb > 0) | _da_tmp.notna() | _sms_tmp), "Status"] = "Ativo"

        td = self.total_dias
        mr = self.mes_ref
        mes_ts = pd.Timestamp(mr)

        # "Data de início da suspensão" costuma vir vazia na base; quando isso
        # acontece, a data real do bloqueio está em "Data de início do bloqueio
        # de rede" (mesmo evento, coluna diferente preenchida pelo sistema de
        # origem). Sem esse fallback, linha suspensa sem a 1ª coluna caía no
        # default de dias = mês inteiro em vez de 0/proporcional.
        db_susp = pd.to_datetime(df.get("Data de início da suspensão"), errors="coerce")
        db_bloq = pd.to_datetime(df.get("Data de início do bloqueio de rede"), errors="coerce")
        db  = db_susp.fillna(db_bloq)
        dc  = pd.to_datetime(df.get("Data de cancelamento"), errors="coerce")
        da  = pd.to_datetime(df.get("Data de ativação"), errors="coerce")

        dias = pd.Series(td, index=df.index, dtype=float)
        dias[df["Status"] == "Pré-ativo"] = 0
        mask_s = df["Status"] == "Suspenso"
        dias[mask_s & (db < mes_ts)] = 0
        mask_sm = mask_s & (db.dt.year == mr.year) & (db.dt.month == mr.month)
        dias[mask_sm] = db[mask_sm].dt.day.astype(float)
        mask_c = df["Status"] == "Cancelamento"
        mask_cp = mask_c & df["ID_CPF/CNPJ"].isin(cancel_prop)
        dias[mask_cp] = dc[mask_cp].dt.day.astype(float)
        mask_ap = (da.dt.year == mr.year) & (da.dt.month == mr.month) & df["ID_CPF/CNPJ"].isin(ativ_prop)
        dias[mask_ap] = (td - da[mask_ap].dt.day + 1).astype(float)

        # Reajuste: aplica apenas para linhas com data de ativação até o corte
        # vigente (REAJUSTE_ANO_CORTE, ver topo do arquivo) — ou sem data.
        per_client = df["ID_CPF/CNPJ"].map(reajuste_map).fillna(0)
        mask_elegivel_reajuste = da.isna() | (da.dt.year < REAJUSTE_ANO_CORTE)
        df["_reajuste_pct"] = per_client.where(mask_elegivel_reajuste, 0.0)
        df["_dias"]              = dias.astype(int)
        df["_mensalidade_reaj"]  = df["Mensalidade"] * (1 + df["_reajuste_pct"])
        raw = df["_mensalidade_reaj"] / td * df["_dias"]
        df["_mensalidade_cobr"]  = raw.apply(_roundup2)
        # Exceção configurável: alguns clientes cobram com arredondamento
        # padrão (pro mais próximo) em vez do ROUNDUP normal do sistema.
        if self.cnpj_sem_arredondamento:
            mask_sem_arred = df["ID_CPF/CNPJ"].apply(self._sem_arredondamento)
            if mask_sem_arred.any():
                df.loc[mask_sem_arred, "_mensalidade_cobr"] = raw[mask_sem_arred].round(2)

        mask_ativ_mes = (da.dt.year == mr.year) & (da.dt.month == mr.month)
        df["_ativacao"] = np.where(mask_ativ_mes, df.get("Preço de ativação", 0), 0.0)

        credito = pd.to_numeric(df.get("Crédito adicionado no Simcard", 0), errors="coerce").fillna(0)
        # Preço do MB Excedente tem piso de R$2,00 — qualquer valor abaixo disso
        # na base é elevado a R$2,00 antes de multiplicar pelo consumo.
        preco_mb_exc = pd.to_numeric(df.get("Preço do MB Excedente", 0), errors="coerce").fillna(0).clip(lower=2.00)
        _exc_raw = credito.apply(_kb_to_mb) * preco_mb_exc
        # Máscara usa .round(2) para definir quais linhas têm excedente real (>= R$0,01).
        # O valor armazenado é full-precision para que a soma por cliente bata com a planilha.
        mask_exc_pos = _exc_raw.round(2) > 0
        df["_excedente"] = _exc_raw.where(mask_exc_pos, 0.0)
        df.loc[mask_exc_pos, "_excedente"] = df.loc[mask_exc_pos, "_excedente"].clip(lower=1.50)

        # Strip ICCID para garantir correspondência mesmo com espaços extras
        multa_map = (
            df_cancel[["ICCID","VALOR DA MULTA"]]
            .dropna(subset=["ICCID"])
            .assign(ICCID=lambda x: x["ICCID"].astype(str).str.strip())
            .set_index("ICCID")["VALOR DA MULTA"].to_dict()
        ) if not df_cancel.empty else {}
        # Multa do arquivo de cancelamentos vai APENAS para as linhas extras df_cr (iccid='')
        # Linhas Ativo/Suspenso/Pré-ativo do inventário nunca têm multa de cancelamento
        # Apenas linhas com status Cancelamento NO PRÓPRIO inventário recebem multa_map
        if "ICCID" in df.columns and multa_map:
            mask_inv_cancel = df["Status"] == "Cancelamento"
            df["_multa"] = 0.0
            df.loc[mask_inv_cancel, "_multa"] = (
                df.loc[mask_inv_cancel, "ICCID"].astype(str).str.strip().map(multa_map).fillna(0)
            )
        else:
            df["_multa"] = 0.0
        df["_sms"]   = df["MSISDN"].astype(str).str.strip().map(sms_map).fillna(0) if sms_map and "MSISDN" in df.columns else 0.0
        df["_total"] = (df["_ativacao"] + df["_mensalidade_cobr"] + df["_sms"] + df["_multa"] + df["_excedente"]).round(2)

        # Pré-ativo: sem mensalidade/excedente/multa/SMS — mas cobra ativação se ativado neste mês
        mask_pre = df["Status"] == "Pré-ativo"
        df.loc[mask_pre, ["_excedente", "_multa", "_sms", "_mensalidade_cobr"]] = 0.0
        df.loc[mask_pre, "_total"] = df.loc[mask_pre, "_ativacao"].round(2)

        # Cancelamento da base principal é descartado — o valor oficial já vem
        # da linha extra gerada a partir do arquivo dedicado de cancelamentos
        # (_montar_cancelamentos, iccid=""). Manter as duas linhas só poluía a
        # planilha (duas entradas pro mesmo SIM, uma sempre zerada); a linha
        # do arquivo dedicado já carrega o status/datas/valores corretos.
        df = df[~mask_c].copy()

        return df

    def _montar_cancelamentos(self, df_cancel, reajuste_map, cancel_prop, sms_map=None, da_map=None):
        if df_cancel.empty: return pd.DataFrame()
        td = self.total_dias
        rows = []
        for _, r in df_cancel.iterrows():
            id_cli = str(r["ID"]).strip() if pd.notna(r["ID"]) else None
            tipo   = str(r.get("_tipo", "Cancelamento")).strip()

            # Desistência: cobra apenas multa, mensalidade = 0
            is_desistencia = "desist" in tipo.lower()

            if is_desistencia:
                mc   = 0.0
                dias = 0
                mr_  = 0.0
                reaj = 0.0
            else:
                dias = r["Data de cancel"].day if id_cli in cancel_prop and pd.notna(r["Data de cancel"]) else td
                # Busca data de ativação: primeiro usa coluna do próprio arquivo de cancelamentos
                # (formato novo com "Data ativação"), depois tenta da_map (inventário) como fallback.
                # SIMs cancelados já saem do inventário, então da_map pode não tê-los.
                da_from_row = r.get("Data ativação")
                if da_from_row is not None and pd.notna(da_from_row):
                    da = da_from_row
                else:
                    iccid_str = str(r.get("ICCID", "")).strip()
                    da = da_map.get(iccid_str) if da_map else None
                applies_reaj = da is not None and not pd.isna(da) and da.year < REAJUSTE_ANO_CORTE
                reaj = reajuste_map.get(id_cli, 0) if applies_reaj else 0.0
                mr_  = float(r["Mensalidade"]) * (1 + reaj)
                mc   = _roundup2(mr_ / td * dias) if td > 0 else 0

            multa  = float(r["VALOR DA MULTA"])
            msisdn = str(r.get("Linha", "")).strip()
            sms_c  = float(sms_map.get(msisdn, 0)) if sms_map and msisdn else 0.0
            rows.append({"ID_CPF/CNPJ": id_cli, "CPF/CNPJ": re.sub(r"\D","",id_cli.replace("ss_","")) if id_cli else None,
                         "Nome do cliente": r.get("clientes"), "ICCID": str(r.get("ICCID","")).strip(),
                         "MSISDN": msisdn, "Mensalidade": float(r["Mensalidade"]),
                         "Status": "Cancelamento",
                         "_data_ativacao": r.get("Data ativação"),
                         "_data_cancelamento": r.get("Data de cancel"),
                         "_reajuste_pct": reaj, "_dias": dias, "_mensalidade_reaj": mr_,
                         "_mensalidade_cobr": mc, "_ativacao": 0, "_excedente": 0,
                         "_multa": multa, "_sms": sms_c, "_total": round(mc + multa + sms_c, 2)})
        return pd.DataFrame(rows)

    def _montar_fretes(self, df_fretes):
        if df_fretes.empty: return pd.DataFrame()
        rows = []
        for _, r in df_fretes.iterrows():
            id_cli = _sanitize_id(r["id"])
            if not id_cli:
                continue
            rows.append({"ID_CPF/CNPJ": id_cli, "CPF/CNPJ": id_cli.replace("ss_", ""),
                         "Nome do cliente": r.get("cliente"),
                         "Status": "Frete", "_dias": 0, "_reajuste_pct": 0,
                         "_mensalidade_reaj": 0, "_mensalidade_cobr": 0,
                         "_ativacao": 0, "_excedente": 0, "_multa": 0, "_sms": 0,
                         "_total": float(r.get("valor",0))})
        return pd.DataFrame(rows)

    def _montar_mensageria(self, df_men: "pd.DataFrame") -> "pd.DataFrame":
        """Cria linhas de cobrança de mensageria (R$ 9,90 fixo por linha do arquivo)."""
        if df_men.empty: return pd.DataFrame()
        VALOR_MENS = self.mensageria_valor
        rows = []
        for _, r in df_men.iterrows():
            id_cli = str(r["id_cliente"]).strip()
            if not id_cli or id_cli.lower() in ("nan", "none", ""):
                continue
            rows.append({"ID_CPF/CNPJ": id_cli, "CPF/CNPJ": id_cli.replace("ss_", ""),
                         "Nome do cliente": None,
                         "Status": "Pacote Mensageria", "_dias": 0, "_reajuste_pct": 0,
                         "_mensalidade_reaj": 0, "_mensalidade_cobr": 0,
                         "_ativacao": 0, "_excedente": 0, "_multa": 0, "_sms": 0,
                         "_total": VALOR_MENS})
        return pd.DataFrame(rows)
