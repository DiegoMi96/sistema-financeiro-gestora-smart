from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, timezone
import io
import os
import tempfile
import uuid
import json

from app.database import get_db
from app.models import (
    User, BillingCycle, BillingLine, BillingClientSummary,
    BillingAdjustment, AuditLog, BillingStatus, AdjustmentType
)
from app.routers.auth import get_current_user
from app.core.permissions import require_permission, get_permission

router = APIRouter(prefix="/billing", tags=["Faturamento"])


def _task_state_path(task_id: str) -> str:
    return f"/tmp/gs_excel_{task_id}.json"


def _task_xlsx_path(task_id: str) -> str:
    return f"/tmp/gs_excel_{task_id}.xlsx"


def _prebuilt_excel_path(cycle_id: int) -> str:
    """Path estável do Excel pré-gerado na aprovação do ciclo."""
    return f"/tmp/gs_prebuilt_{cycle_id}.xlsx"


def _write_task(task_id: str, state: dict) -> None:
    with open(_task_state_path(task_id), "w") as f:
        json.dump(state, f)


def _read_task(task_id: str) -> dict | None:
    try:
        with open(_task_state_path(task_id)) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def _bg_excel_export(task_id: str, cycle_id: int) -> None:
    """Background task: generates Excel, writes state to /tmp (shared across workers)."""
    from app.database import SessionLocal
    from app.services.excel_generator import generate_faturamento_excel

    # Preserva user_id do estado inicial para manter a verificação de acesso
    initial = _read_task(task_id) or {}
    user_id = initial.get("user_id")

    try:
        db = SessionLocal()
        try:
            cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
            cursor = db.execute(
                text("""
                    SELECT
                        bl.id_smart, bl.iccid, bl.msisdn, bl.operadora, bl.status,
                        bl.nome_pedido, bl.bloqueio_automatico, bl.nome_contrato, bl.fornecedor,
                        bl.bloqueio_imei, bl.imsi, bl.status_bloqueio_rede,
                        bl.apelido, bl.observacao, bl.tipo_compartilhamento,
                        bl.operadora_especifica, bl.elegivel_suspensao,
                        bl.ultima_apn, bl.imei, bl.ultima_conexao,
                        bl.status_rede, bl.operadora_conectada,
                        bl.data_ativacao, bl.data_cancelamento,
                        bl.data_inicio_bloqueio, bl.data_fim_bloqueio_rede,
                        bl.data_inicio_suspensao, bl.data_fim_suspensao, bl.data_fim_pre_ativacao,
                        bl.mensalidade_base, bl.preco_ativacao, bl.preco_mb_excedente,
                        bl.credito_simcard_kb, bl.credito_contrato,
                        bl.tipo_fidelidade, bl.multa_contrato,
                        bl.dias_pre_ativacao, bl.porcentagem_consumo, bl.consumo_total_kb,
                        bl.reajuste_pct, bl.mensalidade_reaj, bl.dias,
                        bl.mensalidade_cobrada, bl.ativacao_cobrada, bl.excedente_cobrado,
                        bl.multa_cobrada, bl.sms_cobrado, bl.total_linha,
                        c.nome AS client_nome
                    FROM billing_lines bl
                    LEFT JOIN clients c ON c.id_smart = bl.id_smart
                    WHERE bl.cycle_id = :cid
                    ORDER BY bl.id_smart
                """).execution_options(stream_results=True),
                {"cid": cycle_id},
            )
            buf = generate_faturamento_excel(cycle, cursor.yield_per(2000))
        finally:
            db.close()

        xlsx_path = _task_xlsx_path(task_id)
        with open(xlsx_path, "wb") as f:
            f.write(buf.getvalue())

        _write_task(task_id, {"status": "ready", "cycle_id": cycle_id, "file": xlsx_path, "user_id": user_id})
        # Mantém cópia estável para servir requisições futuras instantaneamente
        import shutil
        shutil.copy(xlsx_path, _prebuilt_excel_path(cycle_id))
    except Exception as exc:
        _write_task(task_id, {"status": "error", "cycle_id": cycle_id, "error": str(exc), "user_id": user_id})


def _bg_excel_pregenerate(cycle_id: int) -> None:
    """Pré-gera Excel do ciclo na aprovação — salva em path estável sem vínculo de task_id."""
    from app.database import SessionLocal
    from app.services.excel_generator import generate_faturamento_excel

    prebuilt = _prebuilt_excel_path(cycle_id)
    print(f"[pre-gen] Iniciando pré-geração Excel ciclo {cycle_id}", flush=True)
    try:
        db = SessionLocal()
        try:
            cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
            cursor = db.execute(
                text("""
                    SELECT
                        bl.id_smart, bl.iccid, bl.msisdn, bl.operadora, bl.status,
                        bl.nome_pedido, bl.bloqueio_automatico, bl.nome_contrato, bl.fornecedor,
                        bl.bloqueio_imei, bl.imsi, bl.status_bloqueio_rede,
                        bl.apelido, bl.observacao, bl.tipo_compartilhamento,
                        bl.operadora_especifica, bl.elegivel_suspensao,
                        bl.ultima_apn, bl.imei, bl.ultima_conexao,
                        bl.status_rede, bl.operadora_conectada,
                        bl.data_ativacao, bl.data_cancelamento,
                        bl.data_inicio_bloqueio, bl.data_fim_bloqueio_rede,
                        bl.data_inicio_suspensao, bl.data_fim_suspensao, bl.data_fim_pre_ativacao,
                        bl.mensalidade_base, bl.preco_ativacao, bl.preco_mb_excedente,
                        bl.credito_simcard_kb, bl.credito_contrato,
                        bl.tipo_fidelidade, bl.multa_contrato,
                        bl.dias_pre_ativacao, bl.porcentagem_consumo, bl.consumo_total_kb,
                        bl.reajuste_pct, bl.mensalidade_reaj, bl.dias,
                        bl.mensalidade_cobrada, bl.ativacao_cobrada, bl.excedente_cobrado,
                        bl.multa_cobrada, bl.sms_cobrado, bl.total_linha,
                        c.nome AS client_nome
                    FROM billing_lines bl
                    LEFT JOIN clients c ON c.id_smart = bl.id_smart
                    WHERE bl.cycle_id = :cid
                    ORDER BY bl.id_smart
                """).execution_options(stream_results=True),
                {"cid": cycle_id},
            )
            buf = generate_faturamento_excel(cycle, cursor.yield_per(2000))
        finally:
            db.close()

        with open(prebuilt, "wb") as f:
            f.write(buf.getvalue())
        print(f"[pre-gen] Excel ciclo {cycle_id} pré-gerado em {prebuilt}", flush=True)
    except Exception as exc:
        print(f"[pre-gen] Erro pré-geração ciclo {cycle_id}: {exc}", flush=True)


# ─────────────────────────────────────────────
# SCHEMAS
# ─────────────────────────────────────────────

class AdjustmentCreate(BaseModel):
    id_smart:        str
    type:            AdjustmentType
    component:       Optional[str] = None
    valor_original:  float
    valor_ajustado:  float
    justificativa:   str
    observacao:      Optional[str] = None
    analista:        Optional[str] = None
    consultor:       Optional[str] = None
    num_fatura:      Optional[str] = None
    data_vencimento: Optional[str] = None   # ISO date string YYYY-MM-DD
    ofensor:         Optional[str] = None


class AdjustmentApprove(BaseModel):
    approved: bool
    observacao: Optional[str] = None


# ─────────────────────────────────────────────
# CICLOS
# ─────────────────────────────────────────────

@router.get("/cycles")
def list_cycles(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Lista todos os ciclos de faturamento."""
    cycles = db.query(BillingCycle).order_by(
        BillingCycle.year.desc(), BillingCycle.month.desc()
    ).all()
    return [_cycle_to_dict(c) for c in cycles]


@router.delete("/cycles/{cycle_id}", status_code=200)
def delete_cycle(
    cycle_id: int,
    current_user: User = Depends(require_permission("can_upload_files")),
    db: Session = Depends(get_db),
):
    """Remove um ciclo em rascunho (falhou no processamento)."""
    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")
    if cycle.status not in ("rascunho", "revisao"):
        raise HTTPException(status_code=400, detail=f"Somente ciclos em rascunho ou revisão podem ser excluídos. Status atual: {cycle.status}")

    # Apaga filhos primeiro (FK sem CASCADE)
    db.query(BillingLine).filter(BillingLine.cycle_id == cycle_id).delete(synchronize_session=False)
    db.query(BillingClientSummary).filter(BillingClientSummary.cycle_id == cycle_id).delete(synchronize_session=False)
    db.query(BillingAdjustment).filter(BillingAdjustment.cycle_id == cycle_id).delete(synchronize_session=False)
    db.delete(cycle)
    db.commit()
    return {"message": "Ciclo excluído"}


@router.get("/cycles/{cycle_id}")
def get_cycle(
    cycle_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from sqlalchemy import func as sqlfunc
    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")
    d = _cycle_to_dict(cycle)
    agg = db.query(
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_mensalidade),   0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_ativacao),      0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_excedente),     0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_multa),         0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_sms),           0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_frete),         0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_mensageria),    0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_cancelamento),  0),
        sqlfunc.coalesce(sqlfunc.sum(BillingClientSummary.total_final),         0),
    ).filter(BillingClientSummary.cycle_id == cycle_id).one()
    d.update({
        "total_mensalidade":  round(float(agg[0]), 2),
        "total_ativacao":     round(float(agg[1]), 2),
        "total_excedente":    round(float(agg[2]), 2),
        "total_multa":        round(float(agg[3]), 2),
        "total_sms":          round(float(agg[4]), 2),
        "total_frete":        round(float(agg[5]), 2),
        "total_mensageria":   round(float(agg[6]), 2),
        "total_cancelamento": round(float(agg[7]), 2),
        "total_final":        round(float(agg[8]), 2),
    })
    return d


@router.post("/cycles/{cycle_id}/approve")
def approve_cycle(
    cycle_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_permission("can_approve_billing")),
    db: Session = Depends(get_db),
):
    """Aprova um ciclo de faturamento — transição Revisão → Aprovado."""
    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")
    if cycle.status != BillingStatus.REVISAO:
        raise HTTPException(status_code=400, detail=f"Ciclo está em status '{cycle.status}', não pode ser aprovado agora")

    from datetime import datetime
    cycle.status      = BillingStatus.APROVADO
    cycle.approved_at = datetime.now(timezone.utc)
    cycle.approved_by = current_user.id

    db.add(AuditLog(
        user_id=current_user.id, action="billing.approve",
        entity="billing_cycle", entity_id=cycle_id,
        details={"month": cycle.month, "year": cycle.year}
    ))
    db.commit()

    # Pré-gera Excel em background para servir instantaneamente quando solicitado
    background_tasks.add_task(_bg_excel_pregenerate, cycle_id)

    return {"message": "Ciclo aprovado", "cycle_id": cycle_id}


# ─────────────────────────────────────────────
# UPLOAD E PROCESSAMENTO
# ─────────────────────────────────────────────

@router.post("/cycles/process")
async def process_billing(
    background_tasks: BackgroundTasks,
    year:  int = 2026,
    month: int = 5,
    base_file:          UploadFile = File(...),
    cancelamentos_file: UploadFile = File(...),
    fretes_file:        UploadFile = File(None),
    vencimentos_file:   UploadFile = File(...),
    atencao_file:       UploadFile = File(None),
    reajuste_file:      UploadFile = File(...),
    sms_file:           UploadFile = File(...),
    mensageria_file:    UploadFile = File(...),
    current_user: User = Depends(require_permission("can_upload_files")),
    db: Session = Depends(get_db),
):
    """
    Recebe os arquivos de entrada, processa o faturamento em background
    e retorna o ID do ciclo criado.
    """
    # Verifica se já existe ciclo para este mês
    existing = db.query(BillingCycle).filter(
        BillingCycle.year == year, BillingCycle.month == month
    ).first()
    if existing and existing.status not in (BillingStatus.RASCUNHO,):
        raise HTTPException(
            status_code=400,
            detail=f"Já existe um ciclo para {month:02d}/{year} com status '{existing.status}'"
        )

    # Cria ou recria o ciclo
    if existing:
        db.delete(existing)
        db.flush()

    cycle = BillingCycle(
        year=year, month=month,
        status=BillingStatus.RASCUNHO,
        created_by=current_user.id,
        base_filename=base_file.filename,
        cancelamentos_filename=cancelamentos_file.filename,
        fretes_filename=fretes_file.filename if fretes_file else None,
    )
    db.add(cycle)
    db.commit()
    db.refresh(cycle)

    tmp_dir = tempfile.mkdtemp(prefix=f"billing_{cycle.id}_")

    async def _save(upload: UploadFile | None, key: str) -> str | None:
        if not upload:
            return None
        path = os.path.join(tmp_dir, f"{key}.xlsx")
        content = await upload.read()
        with open(path, "wb") as f:
            f.write(content)
        return path

    # Salva todos os arquivos em disco — base é salva diretamente (sem manter bytes na RAM)
    file_paths = {
        "base":          await _save(base_file,          "base"),
        "cancelamentos": await _save(cancelamentos_file, "cancelamentos"),
        "fretes":        await _save(fretes_file,        "fretes"),
        "vencimentos":   await _save(vencimentos_file,   "vencimentos"),
        "atencao":       await _save(atencao_file,       "atencao"),
        "reajuste":      await _save(reajuste_file,      "reajuste"),
        "sms":           await _save(sms_file,           "sms"),
        "mensageria":    await _save(mensageria_file,    "mensageria"),
    }

    background_tasks.add_task(
        _run_billing_engine,
        cycle_id=cycle.id,
        year=year,
        month=month,
        file_paths=file_paths,
        tmp_dir=tmp_dir,
    )

    return {
        "message": "Processamento iniciado",
        "cycle_id": cycle.id,
        "status": "rascunho"
    }


def _safe_date(val):
    """Converte Timestamp/date para date, retorna None se NaT/None/NaN."""
    try:
        import pandas as pd
        if val is None or (hasattr(val, '__class__') and pd.isna(val)):
            return None
        if hasattr(val, "date"):
            return val.date()
        return None
    except Exception:
        return None


def _run_billing_engine(cycle_id: int, year: int, month: int, file_paths: dict, tmp_dir: str, base_bytes: bytes | None = None):
    """
    Executa o motor de faturamento em background com processamento em chunks.
    Pico de RAM: ~830 MB por ~45 s durante leitura da base; depois ~60 MB/chunk.
    """
    import gc as _gc
    import os as _os
    import pandas as pd
    from collections import defaultdict
    from app.services.billing_engine import BillingEngineService
    from app.database import SessionLocal
    from sqlalchemy import text as _text
    import shutil, re as _re

    BATCH_SIZE  = 5_000
    CHUNK_SIZE  = 50_000

    db = SessionLocal()
    csv_path = None
    try:
        # ── Ler configurações do banco ────────────────────────────────────────
        try:
            _cnpj_raw = db.execute(_text("SELECT value FROM system_settings WHERE key='cnpj_excluidos'")).scalar()
            _mens_raw = db.execute(_text("SELECT value FROM system_settings WHERE key='mensageria_valor'")).scalar()
        except Exception:
            _cnpj_raw = _mens_raw = None

        _cnpj_excluidos = {
            _re.sub(r"\D", "", line).strip()
            for line in (_cnpj_raw or "").splitlines()
            if _re.sub(r"\D", "", line.strip())
        } or None
        try:
            _mensageria_valor = float((_mens_raw or "").replace(",", ".")) if _mens_raw else None
        except (ValueError, TypeError):
            _mensageria_valor = None

        print(f"⏳ Iniciando motor — ciclo {cycle_id} ({month:02d}/{year})", flush=True)
        engine = BillingEngineService(year=year, month=month,
                                      cnpj_excluidos=_cnpj_excluidos,
                                      mensageria_valor=_mensageria_valor)

        # ── Setup: carrega refs + converte base Excel → CSV (pico 830 MB, depois 0) ──
        ref = engine.setup(file_paths, base_bytes=base_bytes)
        del base_bytes
        _gc.collect()
        csv_path = ref["csv_path"]

        cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
        if not cycle:
            return

        # ── Helpers de conversão de linha para ORM ──────────────────────────

        def _s(row, col):
            v = row.get(col)
            return str(v).strip() if v is not None and str(v).strip() not in ("", "nan", "None") else None

        def _f(row, col):
            try: return float(row.get(col) or 0)
            except: return 0.0

        def _i(row, col):
            try: return int(row.get(col) or 0)
            except: return 0

        def _fpct(row, col):
            v = row.get(col)
            if v is None: return None
            try: return float(str(v).strip().rstrip('%').replace(',', '.'))
            except: return None

        def _inv_line(row):
            return BillingLine(
                cycle_id=cycle_id,
                id_smart=row.get("ID_CPF/CNPJ"),
                iccid=str(row.get("ICCID", "") or ""),
                msisdn=str(row.get("MSISDN", "") or ""),
                operadora=_s(row, "Operadora"),
                status=row.get("Status"),
                nome_pedido=_s(row, "Nome do pedido"),
                id_pedido=_s(row, "ID do pedido"),
                nome_contrato=_s(row, "Nome do contrato"),
                id_contrato=_s(row, "ID do contrato"),
                bloqueio_automatico=_s(row, "Bloqueio automático"),
                fornecedor=_s(row, "Fornecedor"),
                bloqueio_imei=_s(row, "Bloqueio de IMEI"),
                imsi=_s(row, "IMSI"),
                status_bloqueio_rede=_s(row, "Status do bloqueio de rede"),
                apelido=_s(row, "Apelido") or _s(row, "Nome do cliente"),
                observacao=_s(row, "Observação"),
                tipo_compartilhamento=_s(row, "Tipo de compartilhamento"),
                operadora_especifica=_s(row, "Operadora específica"),
                elegivel_suspensao=_s(row, "Elegível à suspensão"),
                ultima_apn=_s(row, "Última APN Conectada"),
                imei=_s(row, "IMEI"),
                ultima_conexao=_s(row, "Última conexão"),
                status_rede=_s(row, "Status de rede"),
                operadora_conectada=_s(row, "Operadora conectada"),
                data_ativacao=_safe_date(row.get("Data de ativação")),
                data_cancelamento=_safe_date(row.get("Data de cancelamento")),
                data_inicio_bloqueio=_safe_date(row.get("Data de início do bloqueio de rede")),
                data_fim_bloqueio_rede=_safe_date(row.get("Data de término do bloqueio de rede")),
                data_inicio_suspensao=_safe_date(row.get("Data de início da suspensão")),
                data_fim_suspensao=_safe_date(row.get("Data de término da suspensão")),
                data_fim_pre_ativacao=_safe_date(row.get("Data fim da pré-ativação")),
                mensalidade_base=_f(row, "Mensalidade"),
                preco_ativacao=_f(row, "Preço de ativação"),
                preco_mb_excedente=_f(row, "Preço do MB Excedente"),
                credito_simcard_kb=_f(row, "Crédito adicionado no Simcard"),
                franquia_mb=_f(row, "Franquia (MB)"),
                credito_contrato=_f(row, "Crédito adicionado no contrato"),
                tipo_fidelidade=_s(row, "Tipo de fidelidade"),
                multa_contrato=_f(row, "Multa"),
                dias_pre_ativacao=_i(row, "Dias de pré-ativação"),
                porcentagem_consumo=_fpct(row, "Porcentagem de consumo"),
                consumo_total_kb=_f(row, "Consumo total (KB)"),
                reajuste_pct=_f(row, "_reajuste_pct"),
                mensalidade_reaj=_f(row, "_mensalidade_reaj"),
                dias=_i(row, "_dias"),
                mensalidade_cobrada=_f(row, "_mensalidade_cobr"),
                ativacao_cobrada=_f(row, "_ativacao"),
                excedente_cobrado=_f(row, "_excedente"),
                multa_cobrada=_f(row, "_multa"),
                sms_cobrado=_f(row, "_sms"),
                total_linha=_f(row, "_total"),
            )

        def _extra_line(row):
            return BillingLine(
                cycle_id=cycle_id,
                id_smart=row.get("ID_CPF/CNPJ"),
                iccid="",
                msisdn=str(row.get("MSISDN", "") or ""),
                operadora=None,
                status=row.get("Status"),
                data_ativacao=_safe_date(row.get("_data_ativacao")),
                data_cancelamento=_safe_date(row.get("_data_cancelamento")),
                dias=int(row.get("_dias", 0) or 0),
                reajuste_pct=float(row.get("_reajuste_pct", 0) or 0),
                mensalidade_reaj=float(row.get("_mensalidade_reaj", 0) or 0),
                mensalidade_cobrada=float(row.get("_mensalidade_cobr", 0) or 0),
                multa_cobrada=float(row.get("_multa", 0) or 0),
                sms_cobrado=float(row.get("_sms", 0) or 0),
                total_linha=float(row.get("_total", 0) or 0),
            )

        # ── Loop em chunks — insere no banco e acumula agregações ──────────────
        # Acumuladores (apenas totais por cliente — minúsculos vs 830 MB do df)
        acc_mens   = defaultdict(float)   # mensalidade (Ativo+Suspenso+Pré-ativo)
        acc_ativ   = defaultdict(float)   # ativação
        acc_exc    = defaultdict(float)   # excedente
        acc_sms    = defaultdict(float)   # SMS
        acc_total  = defaultdict(float)   # total linha (para boletos)
        acc_qtd_at = defaultdict(int)     # qtd Ativo
        acc_qtd_av = defaultdict(int)     # qtd com ativação
        acc_qtd_su = defaultdict(int)     # qtd Suspenso
        nome_map: dict = {}               # id_smart → nome cliente
        da_map:   dict = {}               # ICCID → data ativação (para cancelamentos)

        inv_count     = 0
        chunk_num     = 0

        print(f"🔄 Processando CSV em chunks de {CHUNK_SIZE} linhas...", flush=True)
        for raw_chunk in pd.read_csv(csv_path, chunksize=CHUNK_SIZE, dtype=str, low_memory=False):
            chunk_num += 1

            # Extrai da_map ANTES de process_chunk (precisa de colunas ainda strings)
            if "ICCID" in raw_chunk.columns and "Data de ativação" in raw_chunk.columns:
                _icc = raw_chunk["ICCID"].astype(str).str.strip()
                _da  = pd.to_datetime(raw_chunk["Data de ativação"], errors="coerce", dayfirst=True)
                for icc, da in zip(_icc, _da):
                    if icc and icc not in ("", "nan") and pd.notna(da):
                        da_map[icc] = da

            processed = engine.process_chunk(raw_chunk, ref)
            del raw_chunk
            _gc.collect()

            if processed.empty:
                continue

            chunk_size_actual = len(processed)
            inv_count += chunk_size_actual

            # Insere chunk no banco em lotes
            for i in range(0, chunk_size_actual, BATCH_SIZE):
                batch_recs = processed.iloc[i:i + BATCH_SIZE].to_dict("records")
                db.bulk_save_objects([_inv_line(r) for r in batch_recs])
                db.flush()

            print(f"  📦 chunk {chunk_num}: {chunk_size_actual} linhas — total={inv_count}", flush=True)

            # Atualiza acumuladores
            for id_smart, grp in processed.groupby("ID_CPF/CNPJ"):
                if id_smart is None:
                    continue
                status_grp = grp["Status"]
                acc_mens[id_smart]   += float(grp.loc[status_grp.isin(["Ativo","Suspenso","Pré-ativo"]), "_mensalidade_cobr"].sum())
                acc_ativ[id_smart]   += float(grp["_ativacao"].sum())
                acc_exc[id_smart]    += float(grp["_excedente"].sum())
                acc_sms[id_smart]    += float(grp["_sms"].sum())
                acc_total[id_smart]  += float(grp["_total"].sum())
                acc_qtd_at[id_smart] += int((status_grp == "Ativo").sum())
                acc_qtd_av[id_smart] += int((grp["_ativacao"] > 0).sum())
                acc_qtd_su[id_smart] += int((status_grp == "Suspenso").sum())

            # Mapa de nomes (primeiro valor não-nulo por cliente)
            for col in ["Nome do cliente", "Apelido"]:
                if col in processed.columns:
                    sub = processed[processed[col].notna() & (processed[col].astype(str).str.strip() != "")]
                    for cli_id, val in sub.groupby("ID_CPF/CNPJ")[col].first().items():
                        if cli_id not in nome_map:
                            nome_map[cli_id] = str(val)

            del processed
            _gc.collect()

        print(f"✅ Loop concluído — {inv_count} linhas de inventário gravadas", flush=True)

        # ── Fretes ───────────────────────────────────────────────────────────
        df_fretes = ref["df_fretes"]
        frete_map: dict = {}
        if not df_fretes.empty:
            for i in range(0, len(df_fretes), BATCH_SIZE):
                chunk_recs = df_fretes.iloc[i:i + BATCH_SIZE].to_dict("records")
                db.bulk_save_objects([_extra_line(r) for r in chunk_recs])
                db.flush()
            frete_map = df_fretes.groupby("ID_CPF/CNPJ")["_total"].sum().to_dict()
            print(f"  ✅ {len(df_fretes)} linhas de fretes gravadas", flush=True)

        # ── Mensageria ────────────────────────────────────────────────────────
        df_men = ref["df_mensageria"]
        men_map: dict = {}
        if not df_men.empty:
            for i in range(0, len(df_men), BATCH_SIZE):
                chunk_recs = df_men.iloc[i:i + BATCH_SIZE].to_dict("records")
                db.bulk_save_objects([_extra_line(r) for r in chunk_recs])
                db.flush()
            men_map = df_men.groupby("ID_CPF/CNPJ")["_total"].sum().to_dict()
            print(f"  ✅ {len(df_men)} linhas de mensageria gravadas", flush=True)

        # ── Cancelamentos do arquivo de cancelamentos (df_cancel_rows) ────────
        # Agora que da_map está completo, monta as linhas de cancelamento com fallback correto
        df_cr = engine._montar_cancelamentos(
            ref["df_cancel"], ref["reajuste_map"], ref["cancel_prop"],
            ref["sms_map"], da_map
        )
        cancel_mens_map:  dict = {}
        cancel_multa_map: dict = {}
        sms_cr:           dict = {}
        qtd_cancel:       dict = {}
        if not df_cr.empty:
            for i in range(0, len(df_cr), BATCH_SIZE):
                chunk_recs = df_cr.iloc[i:i + BATCH_SIZE].to_dict("records")
                db.bulk_save_objects([_extra_line(r) for r in chunk_recs])
                db.flush()
            cancel_mens_map  = df_cr.groupby("ID_CPF/CNPJ")["_mensalidade_cobr"].sum().to_dict()
            cancel_multa_map = df_cr.groupby("ID_CPF/CNPJ")["_multa"].sum().to_dict()
            sms_cr           = df_cr.groupby("ID_CPF/CNPJ")["_sms"].sum().to_dict()
            qtd_cancel       = df_cr.groupby("ID_CPF/CNPJ").size().to_dict()
            print(f"  ✅ {len(df_cr)} linhas de cancelamento gravadas", flush=True)

        # ── Agrega nomes via Asaas ────────────────────────────────────────────
        if nome_map:
            id_smart_list = list(nome_map.keys())
            ext_rows = db.execute(
                _text("""
                    SELECT DISTINCT ON (external_reference) external_reference, name
                    FROM asaas_customers_sync
                    WHERE external_reference = ANY(:ids) AND name IS NOT NULL AND name <> ''
                    ORDER BY external_reference, synced_at DESC
                """),
                {"ids": id_smart_list}
            ).fetchall()
            ext_to_name = {r.external_reference: r.name for r in ext_rows}

            missing = [c for c in nome_map if c not in ext_to_name]
            cpf_to_name: dict = {}
            if missing:
                cpfs = [c.replace("ss_", "").replace("SS_", "") for c in missing]
                cpf_rows = db.execute(
                    _text("""
                        SELECT DISTINCT ON (cpf_cnpj) cpf_cnpj, name
                        FROM asaas_customers_sync
                        WHERE cpf_cnpj = ANY(:cpfs) AND name IS NOT NULL AND name <> ''
                        ORDER BY cpf_cnpj, synced_at DESC
                    """),
                    {"cpfs": cpfs}
                ).fetchall()
                cpf_to_name = {r.cpf_cnpj: r.name for r in cpf_rows}

            for cli_id in list(nome_map.keys()):
                if cli_id in ext_to_name:
                    nome_map[cli_id] = ext_to_name[cli_id]
                else:
                    cpf = cli_id.replace("ss_", "").replace("SS_", "")
                    if cpf in cpf_to_name:
                        nome_map[cli_id] = cpf_to_name[cpf]

        # ── Mapas finais de summary ───────────────────────────────────────────
        cancel_map = {cli: round(float(v), 2) for cli, v in cancel_mens_map.items()}
        multa_map  = cancel_multa_map
        sms_map_s  = {cli: acc_sms.get(cli, 0) + sms_cr.get(cli, 0)
                      for cli in set(list(acc_sms.keys()) + list(sms_cr.keys()))}
        venc_map   = ref["venc_map"]

        # ── Boletos: todos os clientes com total > 0 ──────────────────────────
        # total por cliente = inventário + fretes + mensageria
        all_clients = set(acc_total.keys()) | set(frete_map.keys()) | set(men_map.keys())
        boleto_rows = []
        for cli in all_clients:
            tot = (acc_total.get(cli, 0) + frete_map.get(cli, 0) + men_map.get(cli, 0))
            if round(tot, 2) <= 0:
                continue
            boleto_rows.append({
                "Cliente":    cli,
                "valor":      round(tot, 2),
                "vencimento": venc_map.get(cli),
            })
        boleto_rows.sort(key=lambda r: r["valor"], reverse=True)

        # ── Summaries principais ──────────────────────────────────────────────
        summaries = []
        boleto_clients = set()
        for row in boleto_rows:
            cli = row["Cliente"]
            boleto_clients.add(cli)
            summaries.append(BillingClientSummary(
                cycle_id=cycle_id,
                id_smart=cli,
                nome_cliente=nome_map.get(cli),
                total_mensalidade=round(float(acc_mens.get(cli, 0)), 2),
                total_ativacao=round(float(acc_ativ.get(cli, 0)), 2),
                total_excedente=round(float(acc_exc.get(cli, 0)), 2),
                total_multa=round(float(multa_map.get(cli, 0)), 2),
                total_sms=round(float(sms_map_s.get(cli, 0)), 2),
                total_frete=round(float(frete_map.get(cli, 0)), 6),
                total_mensageria=round(float(men_map.get(cli, 0)), 2),
                total_cancelamento=round(float(cancel_map.get(cli, 0)), 2),
                total_final=round(
                    float(acc_mens.get(cli, 0)) + float(acc_ativ.get(cli, 0)) +
                    float(acc_exc.get(cli, 0))  + float(multa_map.get(cli, 0)) +
                    float(sms_map_s.get(cli, 0)) + float(frete_map.get(cli, 0)) +
                    float(men_map.get(cli, 0))  + float(cancel_map.get(cli, 0)),
                    2
                ),
                due_date=_safe_date(row.get("vencimento")),
                qtd_linhas_ativas=int(acc_qtd_at.get(cli, 0)),
                qtd_ativacoes=int(acc_qtd_av.get(cli, 0)),
                qtd_cancelamentos=int(qtd_cancel.get(cli, 0)),
                qtd_suspensoes=int(acc_qtd_su.get(cli, 0)),
            ))
        db.bulk_save_objects(summaries)

        # ── Summaries cancel-only (sem linhas Ativo no inventário) ────────────
        all_cancel_clients = set(cancel_map.keys()) | set(multa_map.keys()) | set(sms_cr.keys())
        cancel_only_clients = all_cancel_clients - boleto_clients
        co_summaries = []
        for cli in cancel_only_clients:
            total_cancel_co = round(float(cancel_map.get(cli, 0)), 2)
            total_multa_co  = round(float(multa_map.get(cli, 0)), 2)
            total_sms_co    = round(float(sms_map_s.get(cli, 0)), 2)
            total_final_co  = round(total_cancel_co + total_multa_co + total_sms_co, 2)
            co_summaries.append(BillingClientSummary(
                cycle_id=cycle_id,
                id_smart=cli,
                nome_cliente=nome_map.get(cli),
                total_mensalidade=0, total_ativacao=0, total_excedente=0,
                total_multa=total_multa_co, total_sms=total_sms_co,
                total_frete=0, total_mensageria=0,
                total_cancelamento=total_cancel_co,
                total_final=total_final_co,
                due_date=None, qtd_linhas_ativas=0, qtd_ativacoes=0,
                qtd_cancelamentos=int(qtd_cancel.get(cli, 0)), qtd_suspensoes=0,
            ))
        if co_summaries:
            db.bulk_save_objects(co_summaries)
            print(f"  ✅ {len(co_summaries)} summaries cancel-only gravados", flush=True)

        # ── Fecha ciclo ───────────────────────────────────────────────────────
        total_value = round(
            sum(s.total_final for s in summaries) +
            sum(s.total_final for s in co_summaries),
            2
        )
        cycle.status        = BillingStatus.REVISAO
        cycle.total_lines   = inv_count
        cycle.total_value   = total_value
        cycle.total_boletos = len(boleto_rows)

        db.commit()
        print(f"🎉 Ciclo {cycle_id} finalizado — {inv_count} linhas, {len(boleto_rows)} boletos, R$ {total_value:,.2f}", flush=True)

    except BaseException as e:
        import traceback, sys
        print(f"❌ ERRO no motor de faturamento (cycle_id={cycle_id}): {type(e).__name__}: {e}", flush=True)
        traceback.print_exc()
        sys.stdout.flush()
        sys.stderr.flush()
        db.rollback()
        try:
            cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
            if cycle:
                cycle.status = BillingStatus.RASCUNHO
                db.commit()
        except Exception:
            pass
    finally:
        db.close()
        if csv_path and _os.path.exists(csv_path):
            try:
                _os.unlink(csv_path)
            except Exception:
                pass
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception:
            pass


# ─────────────────────────────────────────────
# CLIENTES NO CICLO
# ─────────────────────────────────────────────

@router.get("/cycles/{cycle_id}/clients")
def list_cycle_clients(
    cycle_id: int,
    search: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Lista clientes de um ciclo com totais e status de boleto."""
    from sqlalchemy import text as _text

    search_filter = f"%{search}%" if search else None

    # COUNT: inclui JOIN com asaas quando há busca (para buscar pelo nome oficial)
    # Prioridade do JOIN: external_reference (id_smart exato) > cpf_cnpj (fallback)
    _asaas_join = """
        LEFT JOIN (
            SELECT DISTINCT ON (external_reference) external_reference, cpf_cnpj, name
            FROM asaas_customers_sync
            WHERE name IS NOT NULL AND name <> '' AND external_reference IS NOT NULL
            ORDER BY external_reference, synced_at DESC
        ) acs_ext ON acs_ext.external_reference = bcs.id_smart
        LEFT JOIN (
            SELECT DISTINCT ON (cpf_cnpj) cpf_cnpj, name
            FROM asaas_customers_sync
            WHERE name IS NOT NULL AND name <> ''
            ORDER BY cpf_cnpj, synced_at DESC
        ) acs_cpf ON acs_cpf.cpf_cnpj = REPLACE(REPLACE(bcs.id_smart, 'ss_', ''), 'SS_', '')
              AND acs_ext.external_reference IS NULL
    """
    if search_filter:
        count_sql = _text(f"""
            SELECT COUNT(*) FROM billing_client_summaries bcs
            {_asaas_join}
            WHERE bcs.cycle_id = :cid
            AND (bcs.id_smart ILIKE :search OR bcs.nome_cliente ILIKE :search
                 OR acs_ext.name ILIKE :search OR acs_cpf.name ILIKE :search)
        """)
        total = db.execute(count_sql, {"cid": cycle_id, "search": search_filter}).scalar()
    else:
        total = db.execute(
            _text("SELECT COUNT(*) FROM billing_client_summaries WHERE cycle_id = :cid"),
            {"cid": cycle_id}
        ).scalar()

    rows = db.execute(_text(f"""
        SELECT bcs.*,
               COALESCE(acs_ext.name, acs_cpf.name, bcs.nome_cliente, bcs.id_smart) AS client_nome
        FROM billing_client_summaries bcs
        {_asaas_join}
        WHERE bcs.cycle_id = :cid
        AND (:search IS NULL OR bcs.id_smart ILIKE :search OR bcs.nome_cliente ILIKE :search
             OR acs_ext.name ILIKE :search OR acs_cpf.name ILIKE :search)
        ORDER BY bcs.total_final DESC
        LIMIT :lim OFFSET :off
    """), {"cid": cycle_id, "search": search_filter,
           "lim": per_page, "off": (page - 1) * per_page}).mappings().all()

    def _fmt_cnpj(id_smart: str) -> str:
        raw = id_smart.replace("ss_", "").replace("SS_", "").strip()
        d = raw.replace(".", "").replace("/", "").replace("-", "")
        if len(d) == 14:
            return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
        if len(d) == 11:
            return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
        return raw

    items = []
    for r in rows:
        d = dict(r)
        nome = d.get("client_nome") or _fmt_cnpj(d.get("id_smart", ""))
        items.append({
            "id_smart":          d.get("id_smart"),
            "client_nome":       nome,
            "total_mensalidade":  round(d.get("total_mensalidade") or 0, 2),
            "total_ativacao":     round(d.get("total_ativacao") or 0, 2),
            "total_excedente":    round(d.get("total_excedente") or 0, 2),
            "total_multa":        round(d.get("total_multa") or 0, 2),
            "total_sms":          round(d.get("total_sms") or 0, 2),
            "total_frete":        round(d.get("total_frete") or 0, 2),
            "total_mensageria":   round(d.get("total_mensageria") or 0, 2),
            "total_ajustes":      round(d.get("total_ajustes") or 0, 2),
            "total_cancelamento": round(d.get("total_cancelamento") or 0, 2),
            "total_final":        round(d.get("total_final") or 0, 2),
            "due_date":          d["due_date"].isoformat() if d.get("due_date") else None,
            "boleto_url":        d.get("boleto_url"),
            "boleto_status":     d.get("boleto_status"),
            "qtd_linhas_ativas": d.get("qtd_linhas_ativas"),
            "qtd_ativacoes":     d.get("qtd_ativacoes"),
            "qtd_cancelamentos": d.get("qtd_cancelamentos"),
        })

    return {"total": total, "page": page, "per_page": per_page, "items": items}


@router.get("/cycles/{cycle_id}/clients/{id_smart}/summary")
def get_client_summary(
    cycle_id: int,
    id_smart: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Retorna os totais consolidados de um cliente no ciclo (rápido — sem carregar as linhas)."""
    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id,
        BillingClientSummary.id_smart == id_smart,
    ).first()
    if not summary:
        raise HTTPException(status_code=404, detail="Cliente não encontrado no ciclo")
    return _summary_to_dict(summary)


@router.get("/cycles/{cycle_id}/clients/{id_smart}/lines")
def get_client_lines(
    cycle_id: int,
    id_smart: str,
    page: int = 1,
    per_page: int = 200,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Detalha linhas de um cliente no ciclo (paginado, 200 por página)."""
    q = db.query(BillingLine).filter(
        BillingLine.cycle_id == cycle_id,
        BillingLine.id_smart == id_smart,
    )
    total = q.count()
    items = q.order_by(BillingLine.id).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_line_to_dict(l) for l in items],
    }


# ─────────────────────────────────────────────
# AJUSTES
# ─────────────────────────────────────────────

@router.get("/adjustments")
def list_all_adjustments(
    month: int = None,
    year:  int = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Lista todos os ajustes, com filtro opcional por mês/ano do ciclo."""
    # Nome do cliente: o cadastro fica em client_profiles (a tabela legada
    # `clients` está vazia, por isso o nome vinha nulo e a tela repetia o
    # id_smart). Fallback: nome_cliente do resumo do próprio ciclo.
    from app.models import ClientProfile
    from sqlalchemy.orm import aliased
    from sqlalchemy import and_
    ClientAlias  = aliased(ClientProfile)
    SummaryAlias = aliased(BillingClientSummary)
    q = (
        db.query(BillingAdjustment, BillingCycle, ClientAlias, SummaryAlias)
        .join(BillingCycle, BillingAdjustment.cycle_id == BillingCycle.id)
        .outerjoin(ClientAlias, ClientAlias.id_smart == BillingAdjustment.id_smart)
        .outerjoin(SummaryAlias, and_(
            SummaryAlias.cycle_id == BillingAdjustment.cycle_id,
            SummaryAlias.id_smart == BillingAdjustment.id_smart,
        ))
    )
    if month:
        q = q.filter(BillingCycle.month == month)
    if year:
        q = q.filter(BillingCycle.year == year)
    rows = q.order_by(BillingAdjustment.created_at.desc()).all()
    result = []
    for a, cycle, client, summary in rows:
        d = _adjustment_to_dict(a)
        d["cycle_month"]  = cycle.month if cycle else None
        d["cycle_year"]   = cycle.year  if cycle else None
        d["client_nome"]  = (
            (client.nome if client and client.nome else None)
            or (summary.nome_cliente if summary and summary.nome_cliente else None)
        )
        result.append(d)
    return result


@router.get("/cycles/{cycle_id}/adjustments")
def list_adjustments(
    cycle_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    adjs = db.query(BillingAdjustment).filter(
        BillingAdjustment.cycle_id == cycle_id
    ).order_by(BillingAdjustment.created_at.desc()).all()

    # Nome do cliente: cadastro fica em client_profiles; fallback no resumo
    # do próprio ciclo (mesma lógica de list_all_adjustments).
    from app.models import ClientProfile
    id_smarts = {a.id_smart for a in adjs if a.id_smart}
    nomes_profile = {}
    nomes_summary = {}
    if id_smarts:
        nomes_profile = dict(
            db.query(ClientProfile.id_smart, ClientProfile.nome)
            .filter(ClientProfile.id_smart.in_(id_smarts)).all()
        )
        nomes_summary = dict(
            db.query(BillingClientSummary.id_smart, BillingClientSummary.nome_cliente)
            .filter(BillingClientSummary.cycle_id == cycle_id,
                    BillingClientSummary.id_smart.in_(id_smarts)).all()
        )

    result = []
    for a in adjs:
        d = _adjustment_to_dict(a)
        d["client_nome"] = nomes_profile.get(a.id_smart) or nomes_summary.get(a.id_smart)
        result.append(d)
    return result


_COMPONENT_FIELD_MAP = {
    'mensalidade':        'total_mensalidade',
    'ativacao':           'total_ativacao',
    'excedente':          'total_excedente',
    'multa':              'total_multa',
    'multa_cancelamento': 'total_multa',
    'sms':                'total_sms',
    'frete':              'total_frete',
    'mensageria':         'total_mensageria',
    # ativo, pre_ativo, cancelamento, suspenso: sem campo monetário dedicado
    # — ajuste reflete só em total_ajustes e total_final
}


@router.post("/cycles/{cycle_id}/adjustments", status_code=201)
def create_adjustment(
    cycle_id: int,
    data: AdjustmentCreate,
    current_user: User = Depends(require_permission("can_create_adjustment")),
    db: Session = Depends(get_db),
):
    """Cria um ajuste (desconto/acréscimo) para um cliente no ciclo."""
    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")
    if cycle.status == BillingStatus.FECHADO:
        raise HTTPException(status_code=400, detail="Ciclo fechado — ajustes não permitidos")

    # Ajuste acima de R$3.000 requer aprovação do gestor/admin
    diferenca = abs(data.valor_ajustado - data.valor_original)
    requires_approval = diferenca > 3000

    # Analista sempre é o usuário logado para role contas_receber
    analista_name = (
        current_user.name
        if current_user.role in ("contas_receber",)
        else (data.analista or current_user.name)
    )

    from datetime import datetime as _dt
    adj = BillingAdjustment(
        cycle_id=cycle_id,
        id_smart=data.id_smart,
        type=data.type,
        component=data.component,
        valor_original=data.valor_original,
        valor_ajustado=data.valor_ajustado,
        valor_diferenca=data.valor_ajustado - data.valor_original,
        justificativa=data.justificativa,
        observacao=data.observacao,
        analista=analista_name,
        consultor=data.consultor,
        num_fatura=data.num_fatura,
        data_vencimento=_dt.strptime(data.data_vencimento, '%Y-%m-%d').date() if data.data_vencimento else None,
        ofensor=data.ofensor,
        created_by_id=current_user.id,
        requires_approval=requires_approval,
    )
    db.add(adj)

    # Atualiza total do resumo do cliente
    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id,
        BillingClientSummary.id_smart == data.id_smart,
    ).first()
    if summary:
        valor_diferenca = data.valor_ajustado - data.valor_original
        summary.total_ajustes += valor_diferenca
        summary.total_final   += valor_diferenca
        if data.component and data.component in _COMPONENT_FIELD_MAP:
            comp_field = _COMPONENT_FIELD_MAP[data.component]
            setattr(summary, comp_field, round((getattr(summary, comp_field) or 0) + valor_diferenca, 2))

    db.add(AuditLog(
        user_id=current_user.id,
        action="adjustment.create",
        entity="billing_adjustment",
        details={
            "id_smart": data.id_smart, "type": data.type,
            "diferenca": round(diferenca, 2), "justificativa": data.justificativa
        }
    ))
    db.commit()
    db.refresh(adj)
    return _adjustment_to_dict(adj)


@router.put("/cycles/{cycle_id}/adjustments/{adj_id}")
def update_adjustment(
    cycle_id: int,
    adj_id: int,
    data: AdjustmentCreate,
    current_user: User = Depends(require_permission("can_create_adjustment")),
    db: Session = Depends(get_db),
):
    """Edita um ajuste existente (somente se não aprovado e ciclo não fechado)."""
    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")
    if cycle.status == BillingStatus.FECHADO:
        raise HTTPException(status_code=400, detail="Ciclo fechado — edição não permitida")

    adj = db.query(BillingAdjustment).filter(
        BillingAdjustment.id == adj_id,
        BillingAdjustment.cycle_id == cycle_id,
    ).first()
    if not adj:
        raise HTTPException(status_code=404, detail="Ajuste não encontrado")
    if adj.approved_at is not None:
        raise HTTPException(status_code=409, detail="Ajuste já aprovado — edição não permitida")

    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id,
        BillingClientSummary.id_smart == adj.id_smart,
    ).first()

    # Reverte impacto antigo no summary
    if summary:
        summary.total_ajustes -= adj.valor_diferenca
        summary.total_final   -= adj.valor_diferenca
        if adj.component and adj.component in _COMPONENT_FIELD_MAP:
            comp_field = _COMPONENT_FIELD_MAP[adj.component]
            setattr(summary, comp_field, round((getattr(summary, comp_field) or 0) - adj.valor_diferenca, 2))

    nova_diferenca = data.valor_ajustado - data.valor_original
    nova_requires_approval = abs(nova_diferenca) > 3000

    from datetime import datetime as _dt
    adj.type              = data.type
    adj.component         = data.component
    adj.valor_original    = data.valor_original
    adj.valor_ajustado    = data.valor_ajustado
    adj.valor_diferenca   = nova_diferenca
    adj.justificativa     = data.justificativa
    adj.observacao        = data.observacao
    adj.consultor         = data.consultor
    adj.num_fatura        = data.num_fatura
    adj.ofensor           = data.ofensor
    adj.requires_approval = nova_requires_approval
    adj.data_vencimento   = _dt.strptime(data.data_vencimento, '%Y-%m-%d').date() if data.data_vencimento else None

    # Aplica novo impacto no summary
    if summary:
        summary.total_ajustes += nova_diferenca
        summary.total_final   += nova_diferenca
        if data.component and data.component in _COMPONENT_FIELD_MAP:
            comp_field = _COMPONENT_FIELD_MAP[data.component]
            setattr(summary, comp_field, round((getattr(summary, comp_field) or 0) + nova_diferenca, 2))

    db.commit()
    db.refresh(adj)
    return _adjustment_to_dict(adj)


@router.put("/cycles/{cycle_id}/adjustments/{adj_id}/approve")
def approve_adjustment(
    cycle_id: int,
    adj_id: int,
    data: AdjustmentApprove,
    current_user: User = Depends(require_permission("can_approve_adjustment")),
    db: Session = Depends(get_db),
):
    adj = db.query(BillingAdjustment).filter(
        BillingAdjustment.id == adj_id,
        BillingAdjustment.cycle_id == cycle_id,
    ).first()
    if not adj:
        raise HTTPException(status_code=404, detail="Ajuste não encontrado")

    if adj.approved_at is not None:
        raise HTTPException(status_code=409, detail="Ajuste já aprovado — operação não permitida")

    from datetime import datetime
    try:
        if data.approved:
            adj.approved_at    = datetime.now(timezone.utc)
            adj.approved_by_id = current_user.id
        else:
            summary = db.query(BillingClientSummary).filter(
                BillingClientSummary.cycle_id == cycle_id,
                BillingClientSummary.id_smart == adj.id_smart,
            ).first()
            if summary is None:
                raise HTTPException(status_code=409, detail="Summary do cliente não encontrado — reversão não é possível")
            summary.total_ajustes -= adj.valor_diferenca
            summary.total_final   -= adj.valor_diferenca
            if adj.component and adj.component in _COMPONENT_FIELD_MAP:
                comp_field = _COMPONENT_FIELD_MAP[adj.component]
                setattr(summary, comp_field, round((getattr(summary, comp_field) or 0) - adj.valor_diferenca, 2))
            db.delete(adj)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Erro ao processar ajuste")
    return {"message": "Aprovado" if data.approved else "Rejeitado"}


@router.delete("/cycles/{cycle_id}/adjustments/{adj_id}", status_code=204)
def delete_adjustment(
    cycle_id: int,
    adj_id: int,
    current_user: User = Depends(require_permission("can_create_adjustment")),
    db: Session = Depends(get_db),
):
    adj = db.query(BillingAdjustment).filter(
        BillingAdjustment.id == adj_id,
        BillingAdjustment.cycle_id == cycle_id,
    ).first()
    if not adj:
        raise HTTPException(status_code=404, detail="Ajuste não encontrado")

    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if cycle and cycle.status == BillingStatus.FECHADO:
        raise HTTPException(status_code=400, detail="Ciclo fechado — não é possível remover ajuste")

    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id,
        BillingClientSummary.id_smart == adj.id_smart,
    ).first()
    if summary:
        summary.total_ajustes -= adj.valor_diferenca
        summary.total_final   -= adj.valor_diferenca
        if adj.component and adj.component in _COMPONENT_FIELD_MAP:
            comp_field = _COMPONENT_FIELD_MAP[adj.component]
            setattr(summary, comp_field, round((getattr(summary, comp_field) or 0) - adj.valor_diferenca, 2))

    db.delete(adj)
    db.commit()


# ─────────────────────────────────────────────
# EXPORTAÇÕES
# ─────────────────────────────────────────────

@router.post("/cycles/{cycle_id}/export/excel/start")
def start_excel_export(
    cycle_id: int,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")

    task_id = str(uuid.uuid4())
    prebuilt = _prebuilt_excel_path(cycle_id)

    if os.path.exists(prebuilt):
        # Excel já pré-gerado na aprovação — copia para task path e marca pronto imediatamente
        xlsx_path = _task_xlsx_path(task_id)
        import shutil
        shutil.copy(prebuilt, xlsx_path)
        _write_task(task_id, {"status": "ready", "cycle_id": cycle_id, "file": xlsx_path, "user_id": current_user.id})
        print(f"[excel] Ciclo {cycle_id} servido do cache pré-gerado (task {task_id})", flush=True)
    else:
        _write_task(task_id, {"status": "pending", "cycle_id": cycle_id, "user_id": current_user.id})
        background_tasks.add_task(_bg_excel_export, task_id, cycle_id)

    return {"task_id": task_id}


@router.get("/cycles/{cycle_id}/export/excel/status")
def excel_export_status(
    cycle_id: int,
    task_id: str,
    current_user: User = Depends(get_current_user),
):
    task = _read_task(task_id)
    if not task or task.get("cycle_id") != cycle_id:
        raise HTTPException(status_code=404, detail="Tarefa não encontrada")
    if task.get("user_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Acesso negado")
    return {"status": task["status"], "error": task.get("error")}


@router.get("/cycles/{cycle_id}/export/excel/file")
def excel_export_file(
    cycle_id: int,
    task_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from fastapi.responses import Response as FastResponse

    task = _read_task(task_id)
    if not task or task.get("cycle_id") != cycle_id or task.get("status") != "ready":
        raise HTTPException(status_code=404, detail="Arquivo não disponível")
    if task.get("user_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Acesso negado")

    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    filename = f"Faturamento_{cycle.month:02d}_{cycle.year}.xlsx"

    xlsx_path = _task_xlsx_path(task_id)
    try:
        with open(xlsx_path, "rb") as f:
            content = f.read()
    except OSError:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado no disco")

    for path in (xlsx_path, _task_state_path(task_id)):
        try:
            os.unlink(path)
        except OSError:
            pass

    return FastResponse(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/cycles/{cycle_id}/export/excel")
def export_excel(
    cycle_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    raise HTTPException(
        status_code=410,
        detail="Use o endpoint assíncrono: POST /cycles/{id}/export/excel/start"
    )


@router.get("/cycles/{cycle_id}/breakdown")
def get_cycle_breakdown(
    cycle_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Resumo do ciclo agrupado por status — igual à tabela de pivot do Excel."""
    from sqlalchemy import func

    rows = (
        db.query(
            BillingLine.status,
            func.count().label("qtd"),
            func.sum(BillingLine.total_linha).label("total"),
        )
        .filter(BillingLine.cycle_id == cycle_id)
        .group_by(BillingLine.status)
        .all()
    )

    result = []
    soma_linhas = 0.0
    soma_qtd    = 0
    for r in rows:
        val = round(float(r.total or 0), 2)
        soma_linhas += val
        soma_qtd    += r.qtd
        result.append({"status": r.status, "qtd": r.qtd, "total": val})

    result.append({"status": "Total Geral", "qtd": soma_qtd, "total": round(soma_linhas, 2)})
    return result


@router.post("/cycles/{cycle_id}/export/remessa")
async def export_remessa(
    cycle_id: int,
    vencimento_padrao: str = Form(""),
    descricao: str = Form(""),
    forma_pagamento: str = Form("Boleto"),
    juros: float = Form(1.0),
    multa_valor: float = Form(2.0),
    multa_tipo: str = Form("PORCENTAGEM"),
    planilha: UploadFile = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Gera planilha Excel de remessa no formato exato do Asaas."""
    from fastapi.responses import StreamingResponse
    import openpyxl, io as _io, csv as _csv
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date, datetime

    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")

    # Parse vencimento_padrao (YYYY-MM-DD vindo do input date HTML) — opcional
    due_default = None
    if vencimento_padrao.strip():
        try:
            due_default = datetime.strptime(vencimento_padrao.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=422, detail="Formato de vencimento inválido (use YYYY-MM-DD)")

    # Carrega overrides da planilha enviada (id_smart -> date)
    due_overrides: dict[str, date] = {}
    if planilha and planilha.filename:
        content = await planilha.read()
        fname = planilha.filename.lower()
        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            tmp_wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
            tmp_ws = tmp_wb.active
            for row in tmp_ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                id_val = str(row[0]).strip()
                due_val = row[1]
                if isinstance(due_val, datetime):
                    due_overrides[id_val] = due_val.date()
                elif isinstance(due_val, date):
                    due_overrides[id_val] = due_val
                elif isinstance(due_val, str):
                    for fmt_str in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                        try:
                            due_overrides[id_val] = datetime.strptime(due_val.strip(), fmt_str).date()
                            break
                        except ValueError:
                            continue
        elif fname.endswith(".csv"):
            reader = _csv.reader(_io.StringIO(content.decode("utf-8-sig")))
            next(reader, None)  # skip header
            for row in reader:
                if len(row) < 2 or not row[0].strip():
                    continue
                id_val = row[0].strip()
                due_str = row[1].strip()
                for fmt_str in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        due_overrides[id_val] = datetime.strptime(due_str, fmt_str).date()
                        break
                    except ValueError:
                        continue

    summaries = (
        db.query(BillingClientSummary)
        .filter(BillingClientSummary.cycle_id == cycle_id, BillingClientSummary.total_final > 0)
        .order_by(BillingClientSummary.total_final.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    # ── Linha 1: cabeçalho real ───────────────────────────────────
    header = [
        "Identificador único do CLIENTE no sistema de origem (obrigatório)",
        "Vencimento da próxima cobrança (formato dia/mês/ano, ex: 30/01/2016) (obrigatório)",
        "Valor (obrigatório)",
        "Forma de pagamento (Boleto, Cartão) (obrigatório)",
        "Descrição (Opcional)",
        "Parcelas (caso seja parcelado, senão, deixe em branco)",
        "% Juros ao mês",
        "Valor Multa",
        "Tipo da Multa (Porcentagem ou Fixo)",
        "Valor Desconto",
        "Tipo do desconto (Porcentagem ou Fixo)",
    ]
    for col, val in enumerate(header, 1):
        ws.cell(1, col, val)
    hdr_row = 1
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    hdr_font = Font(bold=True)
    for col in range(1, 12):
        cell = ws.cell(hdr_row, col)
        cell.fill = hdr_fill
        cell.font = hdr_font

    # ── Linhas 9+: dados ─────────────────────────────────────────
    for s in summaries:
        due = due_overrides.get(s.id_smart) or due_default or None
        ws.append([
            s.id_smart,
            due,
            round(s.total_final, 2),
            forma_pagamento,
            descricao,
            None,
            juros,
            multa_valor,
            multa_tipo,
            None,
            None,
        ])
        # Formata data como dd/mm/yyyy
        ws.cell(ws.max_row, 2).number_format = "DD/MM/YYYY"

    # Ajusta largura das colunas
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 60

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    mes_str = f"{cycle.month:02d}"
    filename = f"Remessa_Asaas_{mes_str}_{cycle.year}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/cycles/{cycle_id}/export/remessa-template")
def export_remessa_template(
    cycle_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Gera planilha modelo com exemplo de preenchimento para upload de vencimentos."""
    from fastapi.responses import StreamingResponse
    import openpyxl, io as _io
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from datetime import date as _date

    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vencimentos"

    # Cabeçalho
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    for col, h in enumerate(["id_smart", "vencimento"], 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.cell(1, 3, "⚠ Preencha id_smart e vencimento (DD/MM/AAAA) e salve como .xlsx ou .csv")
    ws.cell(1, 3).font = Font(italic=True, color="888888")

    # Linha 2 — exemplo destacado em amarelo
    exemplo_fill = PatternFill("solid", fgColor="FFF2CC")
    exemplo_font = Font(italic=True, color="888888")
    ws.cell(2, 1, "ss_00000000000000").fill = exemplo_fill
    ws.cell(2, 1).font = exemplo_font
    ws.cell(2, 2, "10/07/2026").fill = exemplo_fill
    ws.cell(2, 2).font = exemplo_font
    ws.cell(2, 3, "← EXEMPLO — apague esta linha antes de enviar").fill = exemplo_fill
    ws.cell(2, 3).font = Font(italic=True, color="C00000")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Modelo_Vencimentos_{cycle.month:02d}_{cycle.year}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/cycles/{cycle_id}/clients/{id_smart}/pdf")
def export_client_pdf(
    cycle_id: int,
    id_smart: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from fastapi.responses import StreamingResponse
    from app.services.pdf_generator import generate_client_invoice_pdf
    from app.models import Client

    cycle   = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id,
        BillingClientSummary.id_smart == id_smart
    ).first()
    adjs    = db.query(BillingAdjustment).filter(
        BillingAdjustment.cycle_id == cycle_id,
        BillingAdjustment.id_smart == id_smart
    ).all()
    client  = db.query(Client).filter(Client.id_smart == id_smart).first()

    if not summary:
        raise HTTPException(status_code=404, detail="Cliente não encontrado no ciclo")

    cpf_cnpj_raw = id_smart.replace("ss_", "").replace("SS_", "")

    # SQL GROUP BY — substitui ORM + iteração Python sobre potencialmente milhares de linhas.
    # Retorna ~10 linhas (uma por grupo×operadora) em vez de N linhas individuais.
    agg_rows = db.execute(
        text("""
            SELECT
                CASE
                    WHEN status IN ('Ativo', 'Suspenso') OR status LIKE 'Aguardando%%' THEN 'ativo'
                    WHEN status = 'Cancelamento'  THEN 'cancelamento'
                    WHEN status = 'Pré-ativo'     THEN 'pre_ativo'
                    WHEN lower(status) = 'frete'  THEN 'frete'
                    WHEN lower(status) = 'pacote mensageria' THEN 'mensageria'
                    ELSE 'outros'
                END AS grupo,
                upper(operadora) AS operadora,
                COUNT(*)                                                             AS qtd,
                COALESCE(SUM(mensalidade_cobrada), 0)                               AS mens,
                COALESCE(SUM(ativacao_cobrada), 0)                                  AS atv,
                COALESCE(SUM(CASE WHEN COALESCE(excedente_cobrado,0) > 0
                                  THEN COALESCE(credito_simcard_kb, 0)
                                  ELSE 0 END), 0) / 1024.0                          AS mb_exc,
                COUNT(*) FILTER (WHERE COALESCE(sms_cobrado, 0) > 0)                AS qtd_sms
            FROM billing_lines
            WHERE cycle_id = :cid AND id_smart = :smart
            GROUP BY 1, 2
        """),
        {"cid": cycle_id, "smart": id_smart}
    ).fetchall()

    # Dados do Asaas: external_reference (id_smart exato) > cpf_cnpj (fallback)
    asaas_cust = None
    try:
        asaas_cust = db.execute(
            text("""
                SELECT name, email, phone, mobile_phone,
                       address, address_number, complement, province, city, state, postal_code
                FROM asaas_customers_sync
                WHERE external_reference = :id_smart
                ORDER BY synced_at DESC LIMIT 1
            """),
            {"id_smart": id_smart}
        ).fetchone()
        if not asaas_cust:
            asaas_cust = db.execute(
                text("""
                    SELECT name, email, phone, mobile_phone,
                           address, address_number, complement, province, city, state, postal_code
                    FROM asaas_customers_sync
                    WHERE cpf_cnpj = :cpf
                    ORDER BY synced_at DESC LIMIT 1
                """),
                {"cpf": cpf_cnpj_raw}
            ).fetchone()
    except Exception:
        pass

    nome_fallback = getattr(summary, "nome_cliente", None) or None

    output = generate_client_invoice_pdf(
        cycle, agg_rows, summary, adjs,
        client=client,
        asaas_cust=asaas_cust,
        nome_override=nome_fallback,
        cpf_cnpj_override=cpf_cnpj_raw,
    )

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Fatura_{id_smart}_{cycle.month:02d}{cycle.year}.pdf"'}
    )


@router.get("/cycles/{cycle_id}/clients/{id_smart}/excel")
def export_client_excel(
    cycle_id: int,
    id_smart: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Excel linha a linha do cliente — formato 32 colunas, geração rápida."""
    from fastapi.responses import StreamingResponse
    from sqlalchemy import text

    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")

    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id,
        BillingClientSummary.id_smart == id_smart
    ).first()
    if not summary:
        raise HTTPException(status_code=404, detail="Cliente não encontrado no ciclo")

    # Linhas do cliente com todos os campos e nome via JOIN
    rows = db.execute(
        text("""
            SELECT
                bl.id_smart, bl.iccid, bl.msisdn, bl.operadora, bl.status,
                bl.nome_pedido, bl.id_pedido, bl.nome_contrato, bl.id_contrato,
                bl.bloqueio_automatico, bl.fornecedor,
                bl.bloqueio_imei, bl.imsi, bl.status_bloqueio_rede,
                bl.apelido, bl.observacao, bl.tipo_compartilhamento,
                bl.operadora_especifica, bl.elegivel_suspensao,
                bl.ultima_apn, bl.imei, bl.ultima_conexao,
                bl.status_rede, bl.operadora_conectada,
                bl.data_ativacao, bl.data_cancelamento,
                bl.data_inicio_bloqueio, bl.data_fim_bloqueio_rede,
                bl.data_inicio_suspensao, bl.data_fim_suspensao, bl.data_fim_pre_ativacao,
                bl.mensalidade_base, bl.preco_ativacao, bl.preco_mb_excedente,
                bl.credito_simcard_kb, bl.franquia_mb, bl.credito_contrato,
                bl.tipo_fidelidade, bl.multa_contrato,
                bl.dias_pre_ativacao, bl.porcentagem_consumo, bl.consumo_total_kb,
                bl.reajuste_pct, bl.mensalidade_reaj, bl.dias,
                bl.mensalidade_cobrada, bl.ativacao_cobrada, bl.excedente_cobrado,
                bl.multa_cobrada, bl.sms_cobrado, bl.total_linha,
                c.nome AS client_nome
            FROM billing_lines bl
            LEFT JOIN clients c ON c.id_smart = bl.id_smart
            WHERE bl.cycle_id = :cid AND bl.id_smart = :smart
            ORDER BY bl.iccid
        """),
        {"cid": cycle_id, "smart": id_smart}
    ).fetchall()

    from app.services.excel_generator import generate_client_excel_fast
    output   = generate_client_excel_fast(cycle, rows)
    filename = f"Detalhamento_{id_smart}_{cycle.month:02d}{cycle.year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/cycles/{cycle_id}/export/vencimentos")
def export_vencimentos(
    cycle_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Excel com ID Smart, CNPJ, Nome (Asaas), Vencimento de todos os clientes do ciclo."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io, datetime

    cycle = db.query(BillingCycle).filter(BillingCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo não encontrado")

    summaries = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle_id
    ).order_by(BillingClientSummary.id_smart).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vencimentos"

    VERDE    = "FF3CB54A"
    CINZA    = "FF595959"
    BRANCO   = "FFFFFFFF"
    CINZA_LN = "FFf2f2f2"

    def fill(h): return PatternFill("solid", fgColor=h)
    def bdr():
        s = Side(style="thin", color="FFCCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = f"GESTORA SMART — Vencimentos {cycle.month:02d}/{cycle.year}"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color=BRANCO)
    ws["A1"].fill      = fill(CINZA)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Gerado em {datetime.date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font      = Font(name="Calibri", size=9, color=BRANCO)
    ws["A2"].fill      = fill(VERDE)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    # Cabeçalhos
    HEADERS = ["ID SMART", "CNPJ/CPF", "RAZÃO SOCIAL", "VENCIMENTO", "VALOR TOTAL"]
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(name="Calibri", bold=True, size=9, color=BRANCO)
        c.fill      = fill(CINZA)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    ws.row_dimensions[4].height = 16

    MESES = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

    for i, s in enumerate(summaries):
        row  = 5 + i
        bg   = fill(CINZA_LN) if i % 2 == 1 else fill(BRANCO)
        cnpj = s.id_smart.replace("ss_", "").replace("SS_", "")

        # Formata vencimento: dia do cliente × mês/ano do ciclo
        due_date = s.due_date
        if due_date:
            venc_str = due_date.strftime("%d/%m/%Y") if hasattr(due_date, 'strftime') else str(due_date)
        else:
            venc_str = ""

        vals = [s.id_smart, cnpj, "", venc_str, s.total_final or 0]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=9)
            c.fill      = bg
            c.border    = bdr()
            c.alignment = Alignment(horizontal="right" if col in (4,5) else "left",
                                    vertical="center")
            if col == 5:
                c.number_format = 'R$ #,##0.00'
        ws.row_dimensions[row].height = 14

    # Larguras
    for col, w in enumerate([20, 18, 40, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Vencimentos_{cycle.month:02d}{cycle.year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _cycle_to_dict(c: BillingCycle) -> dict:
    return {
        "id":           c.id,
        "year":         c.year,
        "month":        c.month,
        "period":       f"{c.month:02d}/{c.year}",
        "status":       c.status,
        "total_lines":  c.total_lines,
        "total_value":  round(c.total_value or 0, 2),
        "total_boletos": c.total_boletos,
        "base_filename": c.base_filename,
        "created_at":   c.created_at.isoformat() if c.created_at else None,
        "approved_at":  c.approved_at.isoformat() if c.approved_at else None,
    }


def _summary_to_dict(s: BillingClientSummary) -> dict:
    return {
        "id_smart":            s.id_smart,
        "total_mensalidade":   round(s.total_mensalidade or 0, 2),
        "total_ativacao":      round(s.total_ativacao or 0, 2),
        "total_excedente":     round(s.total_excedente or 0, 2),
        "total_cancelamento":  round(s.total_cancelamento or 0, 2),
        "total_multa":         round(s.total_multa or 0, 2),
        "total_sms":           round(s.total_sms or 0, 2),
        "total_frete":         round(s.total_frete or 0, 2),
        "total_mensageria":    round(s.total_mensageria or 0, 2),
        "total_ajustes":       round(s.total_ajustes or 0, 2),
        "total_final":         round(s.total_final or 0, 2),
        "due_date":            s.due_date.isoformat() if s.due_date else None,
        "boleto_url":          s.boleto_url,
        "boleto_status":       s.boleto_status,
        "qtd_linhas_ativas":   s.qtd_linhas_ativas,
        "qtd_ativacoes":       s.qtd_ativacoes,
        "qtd_cancelamentos":   s.qtd_cancelamentos,
    }


def _line_to_dict(l: BillingLine) -> dict:
    return {
        "id":                   l.id,
        "iccid":                l.iccid,
        "msisdn":               l.msisdn,
        "operadora":            l.operadora,
        "status":               l.status,
        "data_ativacao":        l.data_ativacao.isoformat() if l.data_ativacao else None,
        "data_cancelamento":    l.data_cancelamento.isoformat() if l.data_cancelamento else None,
        "mensalidade_base":     l.mensalidade_base,
        "reajuste_pct":         l.reajuste_pct,
        "mensalidade_reaj":     round(l.mensalidade_reaj or 0, 4),
        "dias":                 l.dias,
        "mensalidade_cobrada":  round(l.mensalidade_cobrada or 0, 2),
        "ativacao_cobrada":     round(l.ativacao_cobrada or 0, 2),
        "excedente_cobrado":    round(l.excedente_cobrado or 0, 2),
        "multa_cobrada":        round(l.multa_cobrada or 0, 2),
        "sms_cobrado":          round(l.sms_cobrado or 0, 2),
        "total_linha":          round(l.total_linha or 0, 2),
    }


def _adjustment_to_dict(a: BillingAdjustment) -> dict:
    return {
        "id":               a.id,
        "id_smart":         a.id_smart,
        "type":             a.type,
        "component":        a.component,
        "valor_original":   round(a.valor_original, 2),
        "valor_ajustado":   round(a.valor_ajustado, 2),
        "valor_diferenca":  round(a.valor_diferenca or 0, 2),
        "justificativa":    a.justificativa,
        "observacao":       a.observacao,
        "analista":         a.analista,
        "consultor":        a.consultor,
        "num_fatura":       a.num_fatura,
        "data_vencimento":  a.data_vencimento.isoformat() if a.data_vencimento else None,
        "ofensor":          a.ofensor,
        "requires_approval": a.requires_approval,
        "approved_at":      a.approved_at.isoformat() if a.approved_at else None,
        "created_at":       a.created_at.isoformat() if a.created_at else None,
        "created_by_id":    a.created_by_id,
    }
