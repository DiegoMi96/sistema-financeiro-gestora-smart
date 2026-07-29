"""
Router do Dashboard do Analista — Gestora Smart
Visão operacional: agenda semanal, alertas, projeção de recebimentos.
"""
import asyncio
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, text
from pydantic import BaseModel
from typing import Optional
from datetime import date, timedelta, datetime
import io

from app.database import get_db, SessionLocal
from app.models import (
    BillingCycle, BillingClientSummary, BillingAdjustment,
    BillingStatus, User
)
from app.models.extra import PaymentRecord, ItauBoleto
from app.routers.auth import get_current_user

router = APIRouter(prefix="/analyst", tags=["Dashboard Analista"])


# ─────────────────────────────────────────────
# SCHEMAS
# ─────────────────────────────────────────────

class PaymentCreate(BaseModel):
    cycle_id:       int
    id_smart:       str
    valor_recebido: float
    data_pagamento: date
    forma:          str   # PIX | boleto | deposito | negociacao
    observacao:     Optional[str] = None


class DueDateUpdate(BaseModel):
    cycle_id:   int
    id_smart:   str
    nova_data:  date
    motivo:     Optional[str] = None


# ─────────────────────────────────────────────
# UPLOAD PLANILHA ITAÚ
# ─────────────────────────────────────────────

@router.post("/itau/upload")
async def upload_itau_boletos(
    month: int = Query(..., ge=1, le=12),
    year:  int = Query(..., ge=2020),
    file:  UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Recebe a planilha xlsx do Itaú (relatório de consulta de boletos).
    Estrutura esperada: metadados nas linhas 1-5, cabeçalho na linha 7, dados a partir da linha 8.
    Colunas: Carteira | Pagador | CPF/CNPJ | Tipo | Nosso Número | Seu Número |
             Data Emissão | Data Vencimento | Data Pagamento | Data Baixa |
             Valor Título | Valor Pago | Status
    Faz upsert pelo campo Nosso Número (chave única).
    """
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um .xlsx do relatório Itaú.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Header esperado na linha 7 (índice 6)
    header_row = rows[6] if len(rows) > 6 else None
    if not header_row or "Pagador" not in str(header_row):
        raise HTTPException(status_code=400, detail="Formato não reconhecido. Verifique se é o relatório correto.")

    def _parse_date(v):
        if not v:
            return None
        if hasattr(v, 'date'):
            return v.date()
        try:
            from datetime import datetime as _dt
            return _dt.strptime(str(v).strip()[:10], "%d/%m/%Y").date()
        except Exception:
            try:
                return date.fromisoformat(str(v).strip()[:10])
            except Exception:
                return None

    # Detecta índice da coluna de descrição pelo cabeçalho (flexível)
    _DESC_NAMES = {
        'histórico', 'historico', 'descrição', 'descricao',
        'mensagem', 'instrução', 'instrucao', 'observação', 'observacao',
        'referência', 'referencia', 'identificação', 'identificacao',
        'especificação', 'especificacao', 'especificaçao',
        'instrução 1', 'instrucao 1', 'instrução1', 'instrucao1',
        'informação', 'informacao', 'informações', 'informacoes',
        'sacador', 'sacador/avalista', 'avalista',
        'campo livre', 'campol ivre',
    }
    _DESC_FRAGMENTS = ('descri', 'instru', 'histor', 'mensa', 'observ', 'ident', 'especif', 'inform')
    desc_col_idx = None
    header_names_detected = [str(c).strip() if c else '' for c in header_row]
    for col_idx, cell in enumerate(header_row):
        if not cell:
            continue
        normalized = str(cell).strip().lower()
        if normalized in _DESC_NAMES or any(normalized.startswith(f) for f in _DESC_FRAGMENTS):
            # ignora colunas fixas conhecidas (0-12)
            if col_idx > 12:
                desc_col_idx = col_idx
                break
            # aceita se for exatamente um dos nomes conhecidos mesmo em posição fixa
            if normalized in _DESC_NAMES:
                desc_col_idx = col_idx
                break

    upload_ref = f"{year}-{month:02d}"
    inseridos = 0
    atualizados = 0

    # Coleta nossos_numeros presentes no arquivo para remoção de registros obsoletos
    nossos_numeros_arquivo: set[str] = set()
    for row in rows[7:]:
        if not row or not row[0]:
            continue
        raw_num = row[4]
        try:
            f = float(raw_num)
            nn = str(int(f)) if f == int(f) else str(f)
        except (TypeError, ValueError):
            nn = str(raw_num).strip() if raw_num is not None else None
        if nn and nn not in ("None", "nan"):
            nossos_numeros_arquivo.add(nn)

    # Remove boletos do upload_ref que não estão mais no arquivo (ex: vencimento em outro mês)
    if nossos_numeros_arquivo:
        db.execute(
            text("DELETE FROM itau_boletos WHERE upload_ref = :ref AND nosso_numero NOT IN :nums"),
            {"ref": upload_ref, "nums": tuple(nossos_numeros_arquivo)},
        )

    for row in rows[7:]:
        if not row or not row[0]:
            continue
        raw_num = row[4]
        # Normaliza nosso_numero: Excel pode enviar inteiros como float (288 → 288.0)
        # Se for numérico sem parte decimal relevante, converte para inteiro string ("288.0" → "288")
        try:
            f = float(raw_num)
            nosso_num = str(int(f)) if f == int(f) else str(f)
        except (TypeError, ValueError):
            nosso_num = str(raw_num).strip() if raw_num is not None else None
        if not nosso_num or nosso_num in ("None", "nan"):
            continue

        status_raw = str(row[12] or "").strip().lower()

        desc_val = None
        if desc_col_idx is not None and desc_col_idx < len(row):
            raw_desc = row[desc_col_idx]
            desc_val = str(raw_desc).strip() if raw_desc is not None else None
            if not desc_val:
                desc_val = None

        valores = {
            "carteira":        str(row[0]).strip() if row[0] else None,
            "pagador":         str(row[1]).strip()[:300] if row[1] else None,
            "cpf_cnpj":        str(row[2]).strip()[:20] if row[2] else None,
            "tipo":            str(row[3]).strip()[:50] if row[3] else None,
            "seu_numero":      str(row[5]).strip()[:50] if row[5] else None,
            "data_emissao":    _parse_date(row[6]),
            "data_vencimento": _parse_date(row[7]),
            "data_pagamento":  _parse_date(row[8]),
            "data_baixa":      _parse_date(row[9]),
            "valor_titulo":    float(row[10]) if row[10] is not None else None,
            "valor_pago":      float(row[11]) if row[11] is not None else None,
            "status":          status_raw,
            "description":     desc_val,
            "upload_ref":      upload_ref,
            # Atualiza o carimbo em TODA importação (insert e update) — antes só
            # o server_default do insert preenchia, então reenviar a mesma planilha
            # não avançava o "última importação" do Resumo por banco (ficava travado).
            "uploaded_at":     func.now(),
        }

        existing = db.query(ItauBoleto).filter(ItauBoleto.nosso_numero == nosso_num).first()
        if existing:
            for k, v in valores.items():
                setattr(existing, k, v)
            atualizados += 1
        else:
            db.add(ItauBoleto(nosso_numero=nosso_num, **valores))
            inseridos += 1

    db.commit()
    return {
        "ok": True,
        "inseridos": inseridos,
        "desc_col_detectada": header_names_detected[desc_col_idx] if desc_col_idx is not None else None,
        "cabecalhos": header_names_detected,
        "atualizados": atualizados,
        "upload_ref": upload_ref,
    }


@router.get("/itau/status")
def itau_upload_status(
    month: int = Query(..., ge=1, le=12),
    year:  int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    upload_ref = f"{year}-{month:02d}"
    count = db.query(func.count(ItauBoleto.id)).filter(ItauBoleto.upload_ref == upload_ref).scalar()
    return {"upload_ref": upload_ref, "registros": count}


# ─────────────────────────────────────────────
# AGENDA SEMANAL
# ─────────────────────────────────────────────

@router.get("/weekly-agenda")
def get_weekly_agenda(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Retorna a agenda da semana atual:
    - Boletos vencendo nos próximos 7 dias
    - Boletos já vencidos (em aberto)
    - Projeção de recebimentos por dia
    """
    hoje      = date.today()
    fim_semana = hoje + timedelta(days=7)

    # Último ciclo aprovado ou em revisão
    cycle = db.query(BillingCycle).filter(
        BillingCycle.status.in_([BillingStatus.REVISAO, BillingStatus.APROVADO])
    ).order_by(BillingCycle.year.desc(), BillingCycle.month.desc()).first()

    if not cycle:
        return {"vencendo_semana": [], "vencidos": [], "projecao_diaria": [], "totais": {}}

    summaries = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == cycle.id
    ).all()

    # Separa por categoria
    vencendo = []
    vencidos  = []
    projecao  = {}   # {data: valor}

    for s in summaries:
        if not s.due_date:
            continue

        # Já pago via Asaas → ignora
        if s.boleto_status in ("RECEIVED", "CONFIRMED"):
            continue

        total = round(s.total_final or 0, 2)
        item  = {
            "id_smart":      s.id_smart,
            "valor":         total,
            "due_date":      s.due_date.isoformat(),
            "boleto_status": s.boleto_status,
            "boleto_url":    s.boleto_url,
            "dias_atraso":   (hoje - s.due_date).days if s.due_date < hoje else 0,
        }

        if s.due_date < hoje:
            vencidos.append(item)
        elif s.due_date <= fim_semana:
            vencendo.append(item)
            # Acumula projeção por dia
            d = s.due_date.isoformat()
            projecao[d] = round(projecao.get(d, 0) + total, 2)

    # Ordena vencidos por atraso (maior primeiro)
    vencidos.sort(key=lambda x: x["dias_atraso"], reverse=True)
    vencendo.sort(key=lambda x: x["due_date"])

    # Projeção diária como lista
    projecao_lista = [
        {"data": k, "valor": v, "qtd": sum(1 for s in summaries if s.due_date and s.due_date.isoformat() == k)}
        for k, v in sorted(projecao.items())
    ]

    return {
        "cycle_id":        cycle.id,
        "periodo":         f"{cycle.month:02d}/{cycle.year}",
        "vencendo_semana": vencendo,
        "vencidos":        vencidos,
        "projecao_diaria": projecao_lista,
        "totais": {
            "qtd_vencendo":      len(vencendo),
            "valor_vencendo":    round(sum(x["valor"] for x in vencendo), 2),
            "qtd_vencidos":      len(vencidos),
            "valor_vencidos":    round(sum(x["valor"] for x in vencidos), 2),
            "valor_semana":      round(sum(x["valor"] for x in projecao_lista), 2),
        }
    }


# ─────────────────────────────────────────────
# ALERTAS
# ─────────────────────────────────────────────

@router.get("/alerts")
def get_alerts(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Alertas prioritários para o analista agir hoje."""
    hoje = date.today()

    cycle = db.query(BillingCycle).filter(
        BillingCycle.status.in_([BillingStatus.REVISAO, BillingStatus.APROVADO])
    ).order_by(BillingCycle.year.desc(), BillingCycle.month.desc()).first()

    alerts = []

    if cycle:
        summaries = db.query(BillingClientSummary).filter(
            BillingClientSummary.cycle_id == cycle.id
        ).all()

        # Boletos vencidos há mais de 5 dias
        vencidos_criticos = [
            s for s in summaries
            if s.due_date and s.due_date < (hoje - timedelta(days=5))
            and s.boleto_status not in ("RECEIVED", "CONFIRMED")
        ]
        if vencidos_criticos:
            alerts.append({
                "level":   "critical",
                "icon":    "AlertTriangle",
                "title":   f"{len(vencidos_criticos)} boletos vencidos há mais de 5 dias",
                "value":   f"R$ {sum(s.total_final or 0 for s in vencidos_criticos):,.2f}",
                "action":  "Ver vencidos",
                "link":    "/analitico/vencidos",
            })

        # Vencendo hoje
        vencendo_hoje = [
            s for s in summaries
            if s.due_date == hoje
            and s.boleto_status not in ("RECEIVED", "CONFIRMED")
        ]
        if vencendo_hoje:
            alerts.append({
                "level":   "warning",
                "icon":    "Clock",
                "title":   f"{len(vencendo_hoje)} boletos vencem hoje",
                "value":   f"R$ {sum(s.total_final or 0 for s in vencendo_hoje):,.2f}",
                "action":  "Ver agenda",
                "link":    "/analitico",
            })

        # Ajustes pendentes de aprovação
        adj_pendentes = db.query(BillingAdjustment).filter(
            BillingAdjustment.cycle_id == cycle.id,
            BillingAdjustment.requires_approval == True,
            BillingAdjustment.approved_at.is_(None),
        ).count()
        if adj_pendentes:
            alerts.append({
                "level":   "info",
                "icon":    "Edit",
                "title":   f"{adj_pendentes} ajustes aguardando aprovação",
                "value":   None,
                "action":  "Ver ajustes",
                "link":    f"/faturamento/{cycle.id}",
            })

    return {"alerts": alerts, "total": len(alerts)}


# ─────────────────────────────────────────────
# REGISTRAR PAGAMENTO MANUAL
# ─────────────────────────────────────────────

@router.post("/payments", status_code=201)
def register_payment(
    data: PaymentCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Registra um pagamento recebido fora do Asaas."""
    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == data.cycle_id,
        BillingClientSummary.id_smart == data.id_smart,
    ).first()

    if not summary:
        raise HTTPException(status_code=404, detail="Cliente não encontrado no ciclo")

    payment = PaymentRecord(
        cycle_id=data.cycle_id,
        id_smart=data.id_smart,
        valor_fatura=summary.total_final or 0,
        valor_recebido=data.valor_recebido,
        data_pagamento=data.data_pagamento,
        forma=data.forma,
        observacao=data.observacao,
        created_by_id=current_user.id,
    )
    db.add(payment)

    # Atualiza status do boleto no summary
    if data.valor_recebido >= (summary.total_final or 0) * 0.99:
        summary.boleto_status = "RECEIVED"
    else:
        summary.boleto_status = "PARTIAL"

    db.commit()
    return {
        "message":  "Pagamento registrado",
        "id_smart": data.id_smart,
        "valor":    data.valor_recebido,
        "status":   summary.boleto_status,
    }


@router.get("/payments/{cycle_id}/{id_smart}")
def get_payments(
    cycle_id: int,
    id_smart: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Histórico de pagamentos de um cliente no ciclo."""
    payments = db.query(PaymentRecord).filter(
        PaymentRecord.cycle_id == cycle_id,
        PaymentRecord.id_smart == id_smart,
    ).order_by(PaymentRecord.data_pagamento.desc()).all()

    return [
        {
            "id":             p.id,
            "valor_fatura":   p.valor_fatura,
            "valor_recebido": p.valor_recebido,
            "data_pagamento": p.data_pagamento.isoformat(),
            "forma":          p.forma,
            "observacao":     p.observacao,
            "created_at":     p.created_at.isoformat(),
        }
        for p in payments
    ]


# ─────────────────────────────────────────────
# ATUALIZAR VENCIMENTO
# ─────────────────────────────────────────────

@router.put("/due-date")
def update_due_date(
    data: DueDateUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Atualiza a data de vencimento de um boleto no ciclo atual."""
    summary = db.query(BillingClientSummary).filter(
        BillingClientSummary.cycle_id == data.cycle_id,
        BillingClientSummary.id_smart == data.id_smart,
    ).first()

    if not summary:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    data_anterior      = summary.due_date
    summary.due_date   = data.nova_data
    db.commit()

    return {
        "message":        "Vencimento atualizado",
        "data_anterior":  data_anterior.isoformat() if data_anterior else None,
        "nova_data":      data.nova_data.isoformat(),
    }


# ─────────────────────────────────────────────
# RESUMO OPERACIONAL — novo layout
# ─────────────────────────────────────────────

@router.get("/operational-summary")
async def get_operational_summary(
    month: Optional[int] = Query(None),
    year:  Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Indicadores operacionais 100% Asaas — sem vínculo com ciclos de faturamento.
    KPIs, donut, forma×situação, comportamento de pagamento, aging,
    por semana e acumulado diário.
    """
    import calendar as cal_mod

    hoje = date.today()

    # ── Período selecionado (default = mês atual) ──────────────
    sel_year  = year  or hoje.year
    sel_month = month or hoje.month

    # ── Dropdown: meses com dados reais no banco ──────────────
    MONTHS_PT = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                 "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    try:
        sync_months = db.execute(text("""
            SELECT DISTINCT
                EXTRACT(YEAR  FROM due_date)::int AS year,
                EXTRACT(MONTH FROM due_date)::int AS month
            FROM asaas_payments_sync
            WHERE due_date IS NOT NULL
            ORDER BY 1 DESC, 2 DESC
        """)).fetchall()
        available_cycles = [
            {"id": None, "month": r.month, "year": r.year,
             "label": f"{MONTHS_PT[r.month]} {r.year}"}
            for r in sync_months
        ]
        available_cycles.reverse()
    except Exception:
        available_cycles = []
        y, m = hoje.year, hoje.month
        for _ in range(12):
            available_cycles.append({"id": None, "month": m, "year": y,
                                      "label": f"{MONTHS_PT[m]} {y}"})
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        available_cycles.reverse()

    # ── Busca dados do sync local (banco) ─────────────────────
    try:
        from app.database import SessionLocal
        _db2 = SessionLocal()
        try:
            rows = _db2.execute(text("""
                SELECT asaas_id, customer_id, customer_name, customer_cpf_cnpj,
                       value, value_original, net_value, due_date, payment_date, credit_date,
                       status, billing_type, description, external_reference
                FROM asaas_payments_sync
                WHERE EXTRACT(YEAR  FROM due_date) = :y
                  AND EXTRACT(MONTH FROM due_date) = :m
            """), {"y": sel_year, "m": sel_month}).fetchall()
        finally:
            _db2.close()

        def _row_to_dict(r):
            return {
                "id":                r.asaas_id,
                "customer":          r.customer_id,
                "name":              r.customer_name,
                "cpfCnpj":           r.customer_cpf_cnpj,
                "value":             r.value or 0,
                "originalValue":     r.value_original,
                "netValue":          r.net_value,
                "dueDate":           r.due_date.isoformat() if r.due_date else None,
                "paymentDate":       r.payment_date.isoformat() if r.payment_date else None,
                "creditDate":        r.credit_date.isoformat() if r.credit_date else None,
                "status":            r.status,
                "billingType":       r.billing_type,
                "description":       r.description,
                "externalReference": r.external_reference,
            }

        payments = [_row_to_dict(r) for r in rows]
    except Exception as _e:
        print(f"⚠️  operational-summary DB read: {_e}")
        payments = []

    # ── Carrega boletos Itaú do mês selecionado ───────────────
    upload_ref_itau = f"{sel_year}-{sel_month:02d}"
    try:
        _dbi0 = SessionLocal()
        try:
            itau_rows = _dbi0.execute(text("""
                SELECT nosso_numero, pagador, cpf_cnpj, valor_titulo, valor_pago,
                       data_vencimento, data_pagamento, status
                FROM itau_boletos
                WHERE upload_ref = :ref
                  AND status != 'cancelada'
                  AND EXTRACT(MONTH FROM data_vencimento) = :m
                  AND EXTRACT(YEAR  FROM data_vencimento) = :y
            """), {"ref": upload_ref_itau, "m": sel_month, "y": sel_year}).fetchall()
        finally:
            _dbi0.close()
    except Exception:
        itau_rows = []

    # converte para dicts no mesmo formato dos pagamentos Asaas
    itau_payments = []
    for r in itau_rows:
        itau_payments.append({
            "id":          r.nosso_numero,
            "customer":    None,
            "name":        r.pagador,
            "cpfCnpj":     r.cpf_cnpj,
            "value":       r.valor_titulo or 0,
            "netValue":    r.valor_pago or 0,
            "dueDate":     r.data_vencimento.isoformat() if r.data_vencimento else None,
            "paymentDate": r.data_pagamento.isoformat() if r.data_pagamento else None,
            "creditDate":  r.data_pagamento.isoformat() if r.data_pagamento else None,
            # normaliza status: paga→RECEIVED, a vencer→PENDING, vencida→OVERDUE
            "status":      {"paga": "RECEIVED", "a vencer": "PENDING", "vencida": "OVERDUE"}.get(
                               (r.status or "").lower(), r.status or ""),
            "billingType": "BOLETO",
            "banco":       "Itaú",
        })

    all_payments = payments + itau_payments

    received_statuses = ("RECEIVED", "CONFIRMED")
    vencidos  = [p for p in all_payments if p.get("status") == "OVERDUE"]
    pendentes = [p for p in all_payments if p.get("status") == "PENDING"]

    total_emitido    = sum(p.get("value", 0) for p in all_payments)
    total_vencido    = sum(p.get("value", 0) for p in vencidos)
    total_confirmado = sum(p.get("value", 0) for p in pendentes)
    qtd_total        = len(all_payments)
    qtd_vencido      = len(vencidos)

    # Recebidos: cash basis — o que entrou no caixa neste mês (credit_date/data_pagamento)
    # independe do mês de vencimento (inclui boletos de meses anteriores pagos neste mês)
    recebidos: list = []
    try:
        _db_rec = SessionLocal()
        _rec_rows = _db_rec.execute(text("""
            SELECT asaas_id, customer_id, customer_name, customer_cpf_cnpj,
                   value, value_original, net_value, due_date, payment_date, credit_date,
                   status, billing_type, description, external_reference
            FROM asaas_payments_sync
            WHERE status IN ('RECEIVED', 'CONFIRMED')
              AND credit_date IS NOT NULL
              AND EXTRACT(YEAR  FROM credit_date) = :y
              AND EXTRACT(MONTH FROM credit_date) = :m
        """), {"y": sel_year, "m": sel_month}).fetchall()
        _db_rec.close()
        recebidos = [_row_to_dict(r) for r in _rec_rows]
    except Exception:
        recebidos = [p for p in all_payments if p.get("status") in received_statuses]

    # Itaú pagos neste mês (data_pagamento), independente do upload_ref
    try:
        _dbi_rec = SessionLocal()
        _itau_rec = _dbi_rec.execute(text("""
            SELECT nosso_numero, pagador, cpf_cnpj, valor_titulo, valor_pago,
                   data_vencimento, data_pagamento, status
            FROM itau_boletos
            WHERE status = 'paga'
              AND data_pagamento IS NOT NULL
              AND EXTRACT(YEAR  FROM data_pagamento) = :y
              AND EXTRACT(MONTH FROM data_pagamento) = :m
        """), {"y": sel_year, "m": sel_month}).fetchall()
        _dbi_rec.close()
        for r in _itau_rec:
            recebidos.append({
                "id":            r.nosso_numero,
                "customer":      None,
                "name":          r.pagador,
                "cpfCnpj":       r.cpf_cnpj,
                "value":         r.valor_titulo or 0,
                "netValue":      r.valor_pago or r.valor_titulo or 0,
                "originalValue": None,
                "dueDate":       r.data_vencimento.isoformat() if r.data_vencimento else None,
                "paymentDate":   r.data_pagamento.isoformat() if r.data_pagamento else None,
                "creditDate":    r.data_pagamento.isoformat() if r.data_pagamento else None,
                "status":        "RECEIVED",
                "billingType":   "BOLETO",
                "banco":         "Itaú",
            })
    except Exception:
        pass

    total_recebido = sum(p.get("netValue") or p.get("value", 0) for p in recebidos)
    qtd_recebido   = len(recebidos)

    # Juros Asaas: value - originalValue quando originalValue preenchido (boleto pago com acréscimo)
    # Juros Itaú:  netValue (valor_pago) - value (valor_titulo) quando positivo
    total_juros = 0.0
    for p in recebidos:
        orig = p.get("originalValue")
        if orig and orig > 0:
            total_juros += max(0, (p.get("value") or 0) - orig)
        elif p.get("banco") == "Itaú":
            total_juros += max(0, (p.get("netValue") or 0) - (p.get("value") or 0))

    # ── Comportamento (creditDate vs dueDate) ──────────────────
    antes_qtd = nodia_qtd = apos_qtd = 0
    dias_list = []
    for p in recebidos:
        credit_str = p.get("creditDate") or p.get("paymentDate") or ""
        due_str    = p.get("dueDate") or ""
        if not credit_str or not due_str:
            continue
        try:
            cd = date.fromisoformat(credit_str[:10])
            dd = date.fromisoformat(due_str[:10])
            diff = (dd - cd).days
            dias_list.append(diff)
            if diff > 0:    antes_qtd += 1
            elif diff == 0: nodia_qtd += 1
            else:            apos_qtd += 1
        except Exception:
            continue
    total_pag   = (antes_qtd + nodia_qtd + apos_qtd) or 1
    prazo_medio = round(sum(dias_list) / len(dias_list), 1) if dias_list else None
    comportamento = {
        "antes_pct": round(antes_qtd / total_pag * 100, 1),
        "nodia_pct": round(nodia_qtd / total_pag * 100, 1),
        "apos_pct":  round(apos_qtd  / total_pag * 100, 1),
        "antes_qtd": antes_qtd, "nodia_qtd": nodia_qtd, "apos_qtd": apos_qtd,
    }

    # ── Por instrumento ────────────────────────────────────────
    def billing_type(p):
        bt = p.get("billingType") or "UNDEFINED"
        return "PIX" if bt in ("UNDEFINED", None, "") else bt

    instr_map: dict = {}
    for p in all_payments:
        bt = billing_type(p)
        if bt not in instr_map:
            instr_map[bt] = {"Recebida": 0.0, "Vencida": 0.0}
        if p.get("status") in received_statuses:
            instr_map[bt]["Recebida"] += p.get("netValue") or p.get("value", 0)
        elif p.get("status") == "OVERDUE":
            instr_map[bt]["Vencida"] += p.get("value", 0)
    por_instrumento = [
        {"name": k, "Recebida": round(v["Recebida"], 2), "Vencida": round(v["Vencida"], 2)}
        for k, v in instr_map.items()
    ]

    # ── Aging — todos OVERDUE com vencimento ≤ último dia do mês selecionado ──
    # (acumulado histórico, excluindo meses futuros ao mês selecionado)
    import calendar as _cal
    _last_day_sel = _cal.monthrange(sel_year, sel_month)[1]
    _aging_limit  = date(sel_year, sel_month, _last_day_sel)

    _AGING_SQL = text("""
        WITH combined AS (
            SELECT value, due_date FROM asaas_payments_sync
            WHERE status = 'OVERDUE' AND due_date >= :inicio AND due_date <= :lim
            UNION ALL
            SELECT valor_titulo AS value, data_vencimento AS due_date FROM itau_boletos
            WHERE status = 'vencida' AND data_vencimento >= :inicio AND data_vencimento <= :lim
        ),
        bucketed AS (
            SELECT
                CASE
                    WHEN (CURRENT_DATE - due_date) BETWEEN 1  AND 4    THEN '1–4 dias'
                    WHEN (CURRENT_DATE - due_date) BETWEEN 5  AND 7    THEN '5–7 dias'
                    WHEN (CURRENT_DATE - due_date) BETWEEN 8  AND 15   THEN '8–15 dias'
                    WHEN (CURRENT_DATE - due_date) BETWEEN 16 AND 30   THEN '15–30 dias'
                    ELSE '> 31 dias'
                END AS bucket,
                value
            FROM combined WHERE due_date IS NOT NULL
        ),
        totals AS (SELECT SUM(value) AS grand_total FROM bucketed)
        SELECT b.bucket, COUNT(*) AS qtd, SUM(b.value) AS valor,
               ROUND((SUM(b.value) / NULLIF(t.grand_total, 0) * 100)::numeric, 1) AS pct
        FROM bucketed b CROSS JOIN totals t
        GROUP BY b.bucket, t.grand_total
    """)
    _BUCKET_META = {
        "1–4 dias":   "Disparo de régua automática",
        "5–7 dias":   "Bloqueio de serviço",
        "8–15 dias":  "Contato direto por telefone",
        "15–30 dias": "Negociação e parcelamento",
        "> 31 dias":  "Escalonamento jurídico",
    }
    _BUCKET_ORDER = list(_BUCKET_META.keys())

    try:
        _ano_inicio = date(sel_year, 1, 1)
        aging_sql_rows = db.execute(_AGING_SQL, {"inicio": _ano_inicio, "lim": _aging_limit}).fetchall()
        aging_map = {r.bucket: {"qtd": r.qtd, "valor": float(r.valor or 0), "pct": float(r.pct or 0)} for r in aging_sql_rows}
    except Exception:
        aging_map = {}

    aging = [
        {
            "bucket": label,
            "qtd":   aging_map.get(label, {}).get("qtd", 0),
            "valor": round(aging_map.get(label, {}).get("valor", 0), 2),
            "pct":   aging_map.get(label, {}).get("pct", 0),
            "acao":  acao,
        }
        for label, acao in _BUCKET_META.items()
    ]

    # ── Donut ──────────────────────────────────────────────────
    donut = [
        {"name": "Recebida", "value": round(total_recebido,   2), "color": "#1E9B6B"},
        {"name": "Vencida",  "value": round(total_vencido,    2), "color": "#ef4444"},
        {"name": "Pendente", "value": round(total_confirmado, 2), "color": "#f59e0b"},
    ]

    # ── Por semana + acumulado diário ──────────────────────────
    last_day = cal_mod.monthrange(sel_year, sel_month)[1]

    def week_of_day(d: int) -> str:
        if d <= 7:  return "W1"
        if d <= 14: return "W2"
        if d <= 21: return "W3"
        return "W4"

    por_semana: dict = {
        "W1": {"planejado": 0.0, "realizado": 0.0, "label": "01–07"},
        "W2": {"planejado": 0.0, "realizado": 0.0, "label": "08–14"},
        "W3": {"planejado": 0.0, "realizado": 0.0, "label": "15–21"},
        "W4": {"planejado": 0.0, "realizado": 0.0, "label": f"22–{last_day:02d}"},
    }
    plan_por_dia: dict = {}
    real_por_dia: dict = {}

    prefix = f"{sel_year}-{sel_month:02d}"
    for p in all_payments:
        due_str = p.get("dueDate") or ""
        if due_str.startswith(prefix):
            try:
                d = int(due_str[8:10])
                por_semana[week_of_day(d)]["planejado"] += p.get("value", 0)
                plan_por_dia[d] = plan_por_dia.get(d, 0) + p.get("value", 0)
            except Exception:
                pass
    for p in recebidos:
        cd = p.get("creditDate") or p.get("paymentDate") or ""
        if cd.startswith(prefix):
            try:
                d = int(cd[8:10])
                por_semana[week_of_day(d)]["realizado"] += p.get("netValue") or p.get("value", 0)
                real_por_dia[d] = real_por_dia.get(d, 0) + (p.get("netValue") or p.get("value", 0))
            except Exception:
                pass
    for wk in por_semana.values():
        wk["planejado"] = round(wk["planejado"], 2)
        wk["realizado"] = round(wk["realizado"], 2)

    acumulado_diario = []
    cum_plan = cum_real = 0.0
    for d in range(1, last_day + 1):
        cum_plan += plan_por_dia.get(d, 0)
        cum_real += real_por_dia.get(d, 0)
        acumulado_diario.append({
            "dia": f"{d:02d}",
            "Planejado": round(cum_plan, 2),
            "Realizado": round(cum_real, 2),
        })

    # ── KPIs derivados ─────────────────────────────────────────
    taxa_inadimplencia = round(total_vencido / total_emitido * 100, 1) if total_emitido else 0
    ticket_recebido    = round(total_recebido / qtd_recebido, 2) if qtd_recebido else 0
    ticket_vencido     = round(total_vencido  / qtd_vencido,  2) if qtd_vencido  else 0

    alerta_aging = None
    if aging and aging[0]["qtd"] > 0 and aging[0]["qtd"] == qtd_vencido:
        alerta_aging = (
            f"Todas as {qtd_vencido} cobranças vencidas têm entre 1 e 3 dias de atraso. "
            "Janela crítica de ação."
        )

    # ── Curva de caixa (histograma dia a dia) ──────────────────
    curva_caixa = [
        {
            "dia":       f"{d:02d}",
            "previsto":  round(plan_por_dia.get(d, 0), 2),
            "realizado": round(real_por_dia.get(d, 0), 2),
        }
        for d in range(1, last_day + 1)
        if plan_por_dia.get(d, 0) > 0 or real_por_dia.get(d, 0) > 0
    ]

    # ── Taxa histórica de conversão — filtrada pelo mês selecionado ──
    try:
        from app.database import SessionLocal as _SL
        _db3 = _SL()
        conv = _db3.execute(text("""
            SELECT
                COUNT(DISTINCT customer_id) FILTER (
                    WHERE status IN ('RECEIVED','CONFIRMED')
                ) AS pagaram,
                COUNT(DISTINCT customer_id) AS com_historico
            FROM asaas_payments_sync
            WHERE customer_id IS NOT NULL
              AND EXTRACT(YEAR  FROM due_date) = :y
              AND EXTRACT(MONTH FROM due_date) = :m
        """), {"y": sel_year, "m": sel_month}).fetchone()
        _db3.close()
        taxa_conversao_historica = {
            "pct":     round(conv.pagaram / conv.com_historico * 100, 1) if (conv and conv.com_historico) else 0,
            "base":    conv.com_historico or 0 if conv else 0,
            "pagaram": conv.pagaram or 0 if conv else 0,
        }
    except Exception:
        taxa_conversao_historica = {"pct": 0, "base": 0, "pagaram": 0}

    # ── Taxa de conversão da régua ────────────────────────────
    # Dos clientes com OVERDUE no mês anterior (M-1), quantos %
    # pagaram (credit_date) a partir do 1º dia de M-1 até hoje.
    try:
        from app.database import SessionLocal as _SL2
        import calendar as _cal2
        _pm = sel_month - 1 if sel_month > 1 else 12
        _py = sel_year if sel_month > 1 else sel_year - 1
        _prev_first = date(_py, _pm, 1)
        from app.database import SessionLocal as _SL2
        _db4 = _SL2()
        regua = _db4.execute(text("""
            WITH overdue_prev AS (
                SELECT DISTINCT customer_id
                FROM asaas_payments_sync
                WHERE status = 'OVERDUE'
                  AND customer_id IS NOT NULL
                  AND EXTRACT(YEAR  FROM due_date) = :py
                  AND EXTRACT(MONTH FROM due_date) = :pm
            ),
            regularizados AS (
                SELECT DISTINCT p.customer_id
                FROM asaas_payments_sync p
                JOIN overdue_prev o ON o.customer_id = p.customer_id
                WHERE p.status IN ('RECEIVED', 'CONFIRMED')
                  AND p.credit_date IS NOT NULL
                  AND p.credit_date >= :from_date
            )
            SELECT
                COUNT(DISTINCT o.customer_id) AS total_overdue,
                COUNT(DISTINCT r.customer_id) AS regularizados,
                COALESCE(SUM(aps.value), 0) AS valor_total_overdue
            FROM overdue_prev o
            LEFT JOIN regularizados r ON r.customer_id = o.customer_id
            LEFT JOIN asaas_payments_sync aps
                   ON aps.customer_id = o.customer_id
                  AND aps.status = 'OVERDUE'
                  AND EXTRACT(YEAR  FROM aps.due_date) = :py
                  AND EXTRACT(MONTH FROM aps.due_date) = :pm
        """), {"py": _py, "pm": _pm, "from_date": _prev_first}).fetchone()
        _db4.close()
        pct_regua = round(regua.regularizados / regua.total_overdue * 100, 1) if (regua and regua.total_overdue) else 0
        taxa_conversao_regua = {
            "pct":                pct_regua,
            "total_overdue":      regua.total_overdue or 0 if regua else 0,
            "regularizados":      regua.regularizados or 0 if regua else 0,
            "valor_total_overdue": round(float(regua.valor_total_overdue or 0), 2) if regua else 0,
            "mes_ref":            f"{_pm:02d}/{_py}",
        }
    except Exception:
        taxa_conversao_regua = {"pct": 0, "total_overdue": 0, "regularizados": 0, "valor_total_overdue": 0, "mes_ref": ""}

    # ── Resumo por banco (Asaas = asaas_payments_sync; Itaú = billing_client_summaries) ──
    try:
        import calendar as _cal_b
        _last_day_b = _cal_b.monthrange(sel_year, sel_month)[1]

        # Asaas — do banco local (mês selecionado)
        _dbb = SessionLocal()
        asaas_rows = _dbb.execute(text("""
            SELECT
                SUM(value)                                                  AS faturado,
                SUM(CASE WHEN status = 'PENDING'
                          AND due_date BETWEEN CURRENT_DATE AND CURRENT_DATE + 30
                         THEN value ELSE 0 END)                           AS a_vencer_30d,
                SUM(CASE WHEN status = 'OVERDUE' THEN value ELSE 0 END)   AS vencido_ativo,
                COUNT(CASE WHEN status = 'OVERDUE' THEN 1 END)            AS qtd_vencido
            FROM asaas_payments_sync
            WHERE EXTRACT(YEAR  FROM due_date) = :y
              AND EXTRACT(MONTH FROM due_date) = :m
        """), {"y": sel_year, "m": sel_month}).fetchone()
        _dbb.close()

        asaas_fat = round(asaas_rows.faturado or 0, 2)
        asaas_av = round(asaas_rows.a_vencer_30d or 0, 2)
        asaas_vv = round(asaas_rows.vencido_ativo or 0, 2)
        asaas_qt = int(asaas_rows.qtd_vencido or 0)
        asaas_tot = asaas_av + asaas_vv
        asaas_inadimp = round(asaas_vv / asaas_tot * 100, 1) if asaas_tot else 0

        # Itaú — da planilha importada (itau_boletos)
        upload_ref = f"{sel_year}-{sel_month:02d}"
        _dbi = SessionLocal()
        itau_row = _dbi.execute(text("""
            SELECT
                SUM(valor_titulo)                                            AS faturado,
                SUM(CASE WHEN status = 'a vencer'
                          AND data_vencimento BETWEEN CURRENT_DATE AND CURRENT_DATE + 30
                         THEN valor_titulo ELSE 0 END)       AS a_vencer_30d,
                SUM(CASE WHEN status = 'vencida'
                         THEN valor_titulo ELSE 0 END)       AS vencido_ativo,
                COUNT(CASE WHEN status = 'vencida' THEN 1 END) AS qtd_vencido,
                MAX(uploaded_at)                                             AS ultima_importacao
            FROM itau_boletos
            WHERE upload_ref = :ref
              AND status != 'cancelada'
              AND EXTRACT(MONTH FROM data_vencimento) = :m
              AND EXTRACT(YEAR  FROM data_vencimento) = :y
        """), {"ref": upload_ref, "m": sel_month, "y": sel_year}).fetchone()
        _dbi.close()

        itau_fat = round(itau_row.faturado or 0, 2)
        itau_av = round(itau_row.a_vencer_30d or 0, 2)
        itau_vv = round(itau_row.vencido_ativo or 0, 2)
        itau_qt = int(itau_row.qtd_vencido or 0)

        resumo_banco = []
        if asaas_av > 0 or asaas_vv > 0 or asaas_fat > 0:
            resumo_banco.append({
                "banco": "Asaas",
                "faturado": asaas_fat,
                "a_vencer_30d": asaas_av,
                "vencido_ativo": asaas_vv,
                "inadimplencia_pct": asaas_inadimp,
                "qtd_vencido": asaas_qt,
            })
        if itau_av > 0 or itau_vv > 0 or itau_fat > 0:
            itau_tot = itau_av + itau_vv
            itau_ultima = itau_row.ultima_importacao
            resumo_banco.append({
                "banco": "Itaú",
                "faturado": itau_fat,
                "a_vencer_30d": itau_av,
                "vencido_ativo": itau_vv,
                "inadimplencia_pct": round(itau_vv / itau_tot * 100, 1) if itau_tot else 0,
                "qtd_vencido": itau_qt,
                "ultima_importacao": itau_ultima.isoformat() if itau_ultima else None,
            })
        resumo_banco.sort(key=lambda x: x["banco"])

        insight_banco = None
        asaas_b = next((b for b in resumo_banco if b["banco"] == "Asaas"), None)
        itau_b  = next((b for b in resumo_banco if b["banco"] == "Itaú"),  None)
        total_av_total = sum(b["a_vencer_30d"] for b in resumo_banco)
        if itau_b and asaas_b and total_av_total:
            itau_pct = round(itau_b["a_vencer_30d"] / total_av_total * 100, 1)
            insight_banco = (
                f"O Itaú representa {itau_pct}% do risco dos próximos 30 dias "
                f"mas tem {itau_b['inadimplencia_pct']}% de inadimplência registrada nesta base. "
                f"O risco Asaas está concentrado nos {asaas_b['qtd_vencido']} boletos vencidos."
                if asaas_b and asaas_b["qtd_vencido"] > 0 else
                f"O Itaú representa {itau_pct}% do volume a vencer nos próximos 30 dias."
            )
        elif asaas_b and total_av_total:
            insight_banco = f"Asaas concentra todo o volume do mês. {asaas_b['qtd_vencido']} boleto(s) vencido(s) em aberto."
    except Exception as _e:
        resumo_banco = []
        insight_banco = None

    # ── Lista de vencidos — mesma fonte do aging (DB) ────────────
    from collections import defaultdict

    def _fmt_cnpj(raw: str) -> str:
        d = (raw or "").replace(".", "").replace("/", "").replace("-", "").strip()
        if len(d) == 14:
            return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
        if len(d) == 11:
            return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
        return raw or ""

    try:
        lv_rows = db.execute(text("""
            WITH asaas_venc AS (
                SELECT aps.customer_cpf_cnpj AS cnpj, aps.customer_name AS nome,
                       'Asaas' AS banco, aps.value, aps.due_date,
                       aps.description,
                       aps.invoice_number AS num_boleto,
                       acs.email,
                       CASE
                           WHEN (CURRENT_DATE - aps.due_date) BETWEEN 1  AND 4  THEN '1–4 dias'
                           WHEN (CURRENT_DATE - aps.due_date) BETWEEN 5  AND 7  THEN '5–7 dias'
                           WHEN (CURRENT_DATE - aps.due_date) BETWEEN 8  AND 15 THEN '8–15 dias'
                           WHEN (CURRENT_DATE - aps.due_date) BETWEEN 16 AND 30 THEN '15–30 dias'
                           ELSE '> 31 dias'
                       END AS bucket
                FROM asaas_payments_sync aps
                LEFT JOIN (
                    SELECT DISTINCT ON (cpf_cnpj) cpf_cnpj, email
                    FROM asaas_customers_sync
                    ORDER BY cpf_cnpj, id
                ) acs ON acs.cpf_cnpj = aps.customer_cpf_cnpj
                WHERE aps.status = 'OVERDUE' AND aps.due_date >= :inicio AND aps.due_date <= :lim
            ),
            itau_venc AS (
                SELECT ib.cpf_cnpj AS cnpj, ib.pagador AS nome,
                       'Itaú' AS banco, ib.valor_titulo AS value, ib.data_vencimento AS due_date,
                       ib.description,
                       ib.nosso_numero AS num_boleto,
                       acs.email,
                       CASE
                           WHEN (CURRENT_DATE - ib.data_vencimento) BETWEEN 1  AND 4  THEN '1–4 dias'
                           WHEN (CURRENT_DATE - ib.data_vencimento) BETWEEN 5  AND 7  THEN '5–7 dias'
                           WHEN (CURRENT_DATE - ib.data_vencimento) BETWEEN 8  AND 15 THEN '8–15 dias'
                           WHEN (CURRENT_DATE - ib.data_vencimento) BETWEEN 16 AND 30 THEN '15–30 dias'
                           ELSE '> 31 dias'
                       END AS bucket
                FROM itau_boletos ib
                LEFT JOIN (
                    SELECT DISTINCT ON (cpf_cnpj) cpf_cnpj, email
                    FROM asaas_customers_sync
                    ORDER BY cpf_cnpj, id
                ) acs ON acs.cpf_cnpj = REGEXP_REPLACE(ib.cpf_cnpj, '[^0-9]', '', 'g')
                WHERE ib.status = 'vencida' AND ib.data_vencimento >= :inicio AND ib.data_vencimento <= :lim
            ),
            combined AS (SELECT * FROM asaas_venc UNION ALL SELECT * FROM itau_venc)
            SELECT
                cnpj, MAX(nome) AS nome, banco, bucket,
                SUM(value) AS valor, COUNT(*) AS qtd,
                MIN(due_date) AS vencimento_orig,
                (CURRENT_DATE - MIN(due_date))::int AS dias,
                MAX(email) AS email,
                MAX(description) AS description,
                STRING_AGG(DISTINCT num_boleto::text, ', ') AS num_boleto
            FROM combined
            WHERE due_date IS NOT NULL
            GROUP BY cnpj, banco, bucket
            ORDER BY valor DESC
        """), {"inicio": date(sel_year, sel_month, 1), "lim": _aging_limit}).fetchall()

        lista_vencidos = [
            {
                "customer_id":    None,
                "nome":           r.nome or "Não identificado",
                "cnpj":           _fmt_cnpj(r.cnpj or ""),
                "valor":          round(float(r.valor or 0), 2),
                "qtd":            int(r.qtd),
                "vencimento_orig": str(r.vencimento_orig) if r.vencimento_orig else None,
                "dias":           int(r.dias or 0),
                "bucket":         r.bucket,
                "banco":          r.banco,
                "email":          r.email,
                "description":    r.description,
                "num_boleto":     r.num_boleto,
                "score":          "novo",
                "historico_rec":  0,
                "historico_tot":  int(r.qtd),
            }
            for r in lv_rows
        ]
    except Exception:
        lista_vencidos = []

    # Descrição dos boletos Itaú (o relatório do Itaú não traz descrição):
    # gera a partir do resumo de faturamento (qtd de linhas do ciclo mais recente),
    # cruzando pelo CNPJ. Ex.: "Cobrança referente a 45 linhas — Ref. Jul/2026".
    try:
        import re as _re_desc
        _MESES_ABREV = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                        'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        _tem_itau_sem_desc = any(v["banco"] == "Itaú" and not v.get("description") for v in lista_vencidos)
        if _tem_itau_sem_desc:
            _last_cycle = db.query(BillingCycle).order_by(BillingCycle.id.desc()).first()
            if _last_cycle:
                _qtd_por_cnpj = {}
                for _s in db.execute(text("""
                    SELECT id_smart, qtd_linhas_ativas
                    FROM billing_client_summaries WHERE cycle_id = :cid
                """), {"cid": _last_cycle.id}).fetchall():
                    _dig = _re_desc.sub(r'\D', '', _s.id_smart or '')
                    if _dig:
                        _qtd_por_cnpj[_dig] = _s.qtd_linhas_ativas or 0
                for _row in lista_vencidos:
                    if _row["banco"] != "Itaú" or _row.get("description"):
                        continue
                    _dig = _re_desc.sub(r'\D', '', _row.get("cnpj") or '')
                    _qtd = _qtd_por_cnpj.get(_dig)
                    if not _qtd:
                        continue
                    _ref = ''
                    _vo = _row.get("vencimento_orig")
                    if _vo and len(str(_vo)) >= 7:
                        try:
                            _mm = int(str(_vo)[5:7])
                            _ref = f" — Ref. {_MESES_ABREV[_mm]}/{str(_vo)[0:4]}"
                        except Exception:
                            _ref = ''
                    _row["description"] = f"Cobrança referente a {_qtd} linhas{_ref}"
    except Exception:
        pass

    # ── Clientes a vencer nos próximos 15 dias ─────────────────
    quinze_dias = hoje + timedelta(days=15)
    _INSTRUMENTO = {
        "BOLETO": "Boleto", "PIX": "PIX", "UNDEFINED": "PIX",
        "CREDIT_CARD": "Cartão", "DEBIT_CARD": "Cartão Déb",
    }

    # — Asaas PENDING next 15d —
    pend_15_asaas = []
    for p in pendentes:
        due_str = p.get("dueDate") or ""
        try:
            dd = date.fromisoformat(due_str[:10])
        except Exception:
            continue
        if hoje <= dd <= quinze_dias:
            pend_15_asaas.append(p)

    asaas_ids_15 = list({p["customer"] for p in pend_15_asaas if p.get("customer")})
    if asaas_ids_15:
        try:
            name_rows_15 = db.execute(
                text("SELECT asaas_id, nome FROM clients WHERE asaas_id = ANY(:ids)"),
                {"ids": asaas_ids_15},
            ).fetchall()
            id_to_nome_15 = {r.asaas_id: r.nome for r in name_rows_15}
        except Exception:
            id_to_nome_15 = {}
    else:
        id_to_nome_15 = {}

    # group by customer (Asaas)
    map_15: dict = defaultdict(lambda: {"valor": 0.0, "qtd": 0, "due_date": None, "billing_type": "BOLETO"})
    for p in pend_15_asaas:
        cid = p.get("customer") or ""
        map_15[cid]["valor"]    += p.get("value", 0)
        map_15[cid]["qtd"]      += 1
        if not map_15[cid]["due_date"]:
            map_15[cid]["due_date"]     = p.get("dueDate")
            map_15[cid]["billing_type"] = p.get("billingType", "BOLETO")

    # — Itaú PENDING next 15d (from DB) —
    try:
        _sel_cycle = db.query(BillingCycle).filter(
            BillingCycle.month == sel_month,
            BillingCycle.year  == sel_year,
        ).first()
        itau_rows = db.execute(text("""
            SELECT bcs.id_smart, c.nome, c.asaas_id,
                   bcs.total_final, bcs.due_date
            FROM billing_client_summaries bcs
            LEFT JOIN clients c ON c.id_smart = bcs.id_smart
            WHERE bcs.asaas_boleto_id IS NULL
              AND bcs.boleto_status = 'PENDING'
              AND bcs.due_date BETWEEN :hoje AND :lim
              AND bcs.cycle_id = :cid
        """), {"hoje": hoje, "lim": quinze_dias, "cid": _sel_cycle.id}).fetchall() if _sel_cycle else []
    except Exception:
        itau_rows = []

    a_vencer_15d = []
    for cid, info in map_15.items():
        dd = date.fromisoformat(info["due_date"][:10]) if info["due_date"] else None
        a_vencer_15d.append({
            "nome":        id_to_nome_15.get(cid, cid),
            "valor":       round(info["valor"], 2),
            "qtd":         info["qtd"],
            "due_date":    info["due_date"],
            "dias":        (dd - hoje).days if dd else None,
            "banco":       "Asaas",
            "instrumento": _INSTRUMENTO.get(info["billing_type"], "Boleto"),
        })
    for r in itau_rows:
        dd = r.due_date
        a_vencer_15d.append({
            "nome":        r.nome or r.id_smart or "",
            "valor":       round(r.total_final or 0, 2),
            "qtd":         1,
            "due_date":    dd.isoformat() if dd else None,
            "dias":        (dd - hoje).days if dd else None,
            "banco":       "Itaú",
            "instrumento": "Boleto",
        })
    a_vencer_15d.sort(key=lambda x: x["valor"], reverse=True)

    return {
        "periodo":          f"{sel_month:02d}/{sel_year}",
        "available_cycles": available_cycles,
        "qtd_total":        qtd_total,
        "emitido":          round(total_emitido, 2),
        "recebido":         round(total_recebido, 2),
        "juros":            round(total_juros, 2),
        "vencido":          round(total_vencido, 2),
        "confirmado":       round(total_confirmado, 2),
        "qtd_recebido":     qtd_recebido,
        "qtd_vencido":      qtd_vencido,
        "qtd_pendente":     len(pendentes),
        "recebido_pct":     round(total_recebido / total_emitido * 100, 1) if total_emitido else 0,
        "taxa_inadimplencia":  taxa_inadimplencia,
        "prazo_medio":         prazo_medio,
        "ticket_recebido":     ticket_recebido,
        "ticket_vencido":      ticket_vencido,
        "comportamento":       comportamento,
        "por_instrumento":     por_instrumento,
        "aging":               aging,
        "alerta_aging":        alerta_aging,
        "donut":               donut,
        "por_semana":               por_semana,
        "acumulado_diario":         acumulado_diario,
        "calendario":               {str(d): round(v, 2) for d, v in plan_por_dia.items()},
        "curva_caixa":              curva_caixa,
        "taxa_conversao_historica": taxa_conversao_historica,
        "taxa_conversao_regua":     taxa_conversao_regua,
        "resumo_banco":             resumo_banco,
        "insight_banco":            insight_banco,
        "lista_vencidos":           lista_vencidos,
        "a_vencer_15d":             a_vencer_15d,
    }


@router.get("/payment-planning")
async def payment_planning(
    month: int = Query(..., ge=1, le=12),
    year:  int = Query(..., ge=2020),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Planejamento comportamental: previsão de recebimento baseada nos últimos
    3 meses de comportamento por CNPJ/CPF (credit_date vs due_date).
    """
    import calendar as cal_mod
    from datetime import date as _date

    last_day    = cal_mod.monthrange(year, month)[1]
    target_start = _date(year, month, 1)

    hm, hy = month - 3, year
    while hm <= 0:
        hm += 12
        hy -= 1
    hist_start = _date(hy, hm, 1)

    # Overrides de vencimento planejado vindos da Lista de Vencidos do mês
    import re as _re
    from app.models.extra import VencidoNota as _VencidoNota

    def _digits(s):
        return _re.sub(r'\D', '', s or '')

    _notas = db.query(_VencidoNota).filter(
        _VencidoNota.mes == month,
        _VencidoNota.ano == year,
        _VencidoNota.vencimento_planejado.isnot(None),
    ).all()
    # chave: apenas dígitos do CNPJ/CPF → dia do mês planejado
    planned_overrides = {_digits(n.cnpj): n.vencimento_planejado.day for n in _notas if n.vencimento_planejado}

    try:
        _db2 = SessionLocal()
        try:
            hist_rows = _db2.execute(text("""
                SELECT
                    customer_cpf_cnpj,
                    ROUND(AVG(
                        EXTRACT(DAY FROM credit_date) - EXTRACT(DAY FROM due_date)
                    )::numeric, 1) AS avg_offset
                FROM asaas_payments_sync
                WHERE status IN ('RECEIVED', 'CONFIRMED')
                  AND credit_date IS NOT NULL
                  AND due_date IS NOT NULL
                  AND due_date >= :hist_start
                  AND due_date < :target_start
                  AND customer_cpf_cnpj IS NOT NULL
                GROUP BY customer_cpf_cnpj
            """), {"hist_start": hist_start, "target_start": target_start}).fetchall()

            hist_offsets = {r.customer_cpf_cnpj: float(r.avg_offset or 0) for r in hist_rows}

            curr_rows = _db2.execute(text("""
                SELECT customer_cpf_cnpj, value, net_value,
                       due_date, credit_date, status
                FROM asaas_payments_sync
                WHERE EXTRACT(YEAR  FROM due_date) = :y
                  AND EXTRACT(MONTH FROM due_date) = :m
            """), {"y": year, "m": month}).fetchall()
        finally:
            _db2.close()
    except Exception as e:
        print(f"⚠️  payment-planning: {e}")
        return {
            "planejado_por_dia": [], "por_semana": {},
            "acumulado": [], "cobertura_pct": 0,
            "clientes_com_historico": 0, "clientes_sem_historico": 0,
        }

    plan_por_dia: dict = {}
    real_por_dia: dict = {}
    clientes_com: set = set()
    clientes_sem: set = set()

    # Asaas — planejado
    # Prioridade: 1) venc. planejado da Lista de Vencidos  2) offset histórico  3) due_date
    for r in curr_rows:
        if not r.due_date:
            continue
        cpf_cnpj  = r.customer_cpf_cnpj
        valor     = r.net_value or r.value or 0
        cnpj_key  = _digits(cpf_cnpj)

        if cnpj_key and cnpj_key in planned_overrides:
            pred_day = max(1, min(planned_overrides[cnpj_key], last_day))
            if cpf_cnpj:
                clientes_com.add(cpf_cnpj)
        elif cpf_cnpj and cpf_cnpj in hist_offsets:
            clientes_com.add(cpf_cnpj)
            offset   = round(hist_offsets[cpf_cnpj])
            pred_day = max(1, min(r.due_date.day + offset, last_day))
        else:
            if cpf_cnpj:
                clientes_sem.add(cpf_cnpj)
            pred_day = r.due_date.day

        plan_por_dia[pred_day] = plan_por_dia.get(pred_day, 0) + valor

    # Asaas — realizado (cash basis: credit_date no mês selecionado, independe do due_date)
    try:
        _db_rec = SessionLocal()
        _rec_rows = _db_rec.execute(text("""
            SELECT net_value, value, credit_date
            FROM asaas_payments_sync
            WHERE status IN ('RECEIVED', 'CONFIRMED')
              AND credit_date IS NOT NULL
              AND EXTRACT(YEAR  FROM credit_date) = :y
              AND EXTRACT(MONTH FROM credit_date) = :m
        """), {"y": year, "m": month}).fetchall()
        _db_rec.close()
        for r in _rec_rows:
            if r.credit_date:
                rd = r.credit_date.day
                real_por_dia[rd] = real_por_dia.get(rd, 0) + (r.net_value or r.value or 0)
    except Exception as _e:
        print(f"⚠️  payment-planning Asaas realizado: {_e}")

    # Itaú — planejado (data_vencimento do mês) + realizado (data_pagamento no mês)
    try:
        _db_itau = SessionLocal()
        _itau_plan = _db_itau.execute(text("""
            SELECT valor_titulo, data_vencimento, cpf_cnpj
            FROM itau_boletos
            WHERE upload_ref = :ref
              AND status != 'cancelada'
              AND EXTRACT(MONTH FROM data_vencimento) = :m
              AND EXTRACT(YEAR  FROM data_vencimento) = :y
        """), {"ref": f"{year}-{month:02d}", "m": month, "y": year}).fetchall()
        for r in _itau_plan:
            if r.data_vencimento:
                cnpj_key = _digits(r.cpf_cnpj)
                if cnpj_key and cnpj_key in planned_overrides:
                    d = max(1, min(planned_overrides[cnpj_key], last_day))
                else:
                    d = max(1, min(r.data_vencimento.day, last_day))
                plan_por_dia[d] = plan_por_dia.get(d, 0) + (r.valor_titulo or 0)

        _itau_real = _db_itau.execute(text("""
            SELECT valor_pago, valor_titulo, data_pagamento
            FROM itau_boletos
            WHERE status = 'paga'
              AND data_pagamento IS NOT NULL
              AND EXTRACT(MONTH FROM data_pagamento) = :m
              AND EXTRACT(YEAR  FROM data_pagamento) = :y
        """), {"m": month, "y": year}).fetchall()
        _db_itau.close()
        for r in _itau_real:
            if r.data_pagamento:
                rd = r.data_pagamento.day
                real_por_dia[rd] = real_por_dia.get(rd, 0) + (r.valor_pago or r.valor_titulo or 0)
    except Exception as _e:
        print(f"⚠️  payment-planning Itaú: {_e}")

    def _wk(d: int) -> str:
        if d <= 7:  return "W1"
        if d <= 14: return "W2"
        if d <= 21: return "W3"
        return "W4"

    por_semana = {
        "W1": {"planejado": 0.0, "label": "01–07"},
        "W2": {"planejado": 0.0, "label": "08–14"},
        "W3": {"planejado": 0.0, "label": "15–21"},
        "W4": {"planejado": 0.0, "label": f"22–{last_day:02d}"},
    }

    planejado_list = []
    acumulado      = []
    cum_plan = cum_real = 0.0

    for d in range(1, last_day + 1):
        vp = round(plan_por_dia.get(d, 0), 2)
        vr = round(real_por_dia.get(d, 0), 2)
        planejado_list.append({"dia": f"{d:02d}", "planejado": vp, "realizado": vr})
        por_semana[_wk(d)]["planejado"] += vp
        cum_plan += vp
        cum_real += vr
        acumulado.append({
            "dia": f"{d:02d}",
            "Planejado": round(cum_plan, 2),
            "Realizado": round(cum_real, 2),
        })

    for wk in por_semana.values():
        wk["planejado"] = round(wk["planejado"], 2)

    total     = len(clientes_com) + len(clientes_sem)
    cobertura = round(len(clientes_com) / total * 100, 1) if total else 0

    return {
        "planejado_por_dia":      planejado_list,
        "por_semana":             por_semana,
        "acumulado":              acumulado,
        "clientes_com_historico": len(clientes_com),
        "clientes_sem_historico": len(clientes_sem),
        "cobertura_pct":          cobertura,
    }


@router.get("/acumulado-detail")
async def acumulado_detail(
    month: int = Query(..., ge=1, le=12),
    year:  int = Query(..., ge=2020),
    current_user: User = Depends(get_current_user),
):
    """Detalhe diário para exportação do Acumulado: cliente, vencimento, pagamento, valor."""
    rows = []
    try:
        _db = SessionLocal()
        asaas_rows = _db.execute(text("""
            SELECT customer_name, due_date, credit_date,
                   net_value, value
            FROM asaas_payments_sync
            WHERE status IN ('RECEIVED', 'CONFIRMED')
              AND credit_date IS NOT NULL
              AND EXTRACT(YEAR  FROM credit_date) = :y
              AND EXTRACT(MONTH FROM credit_date) = :m
            ORDER BY credit_date, customer_name
        """), {"y": year, "m": month}).fetchall()
        for r in asaas_rows:
            rows.append({
                "nome":       r.customer_name or "—",
                "vencimento": r.due_date.isoformat() if r.due_date else None,
                "pagamento":  r.credit_date.isoformat() if r.credit_date else None,
                "valor":      round(float(r.net_value or r.value or 0), 2),
                "banco":      "Asaas",
            })
        itau_rows = _db.execute(text("""
            SELECT pagador, data_vencimento, data_pagamento,
                   valor_pago, valor_titulo
            FROM itau_boletos
            WHERE status = 'paga'
              AND data_pagamento IS NOT NULL
              AND EXTRACT(YEAR  FROM data_pagamento) = :y
              AND EXTRACT(MONTH FROM data_pagamento) = :m
            ORDER BY data_pagamento, pagador
        """), {"y": year, "m": month}).fetchall()
        _db.close()
        for r in itau_rows:
            rows.append({
                "nome":       r.pagador or "—",
                "vencimento": r.data_vencimento.isoformat() if r.data_vencimento else None,
                "pagamento":  r.data_pagamento.isoformat() if r.data_pagamento else None,
                "valor":      round(float(r.valor_pago or r.valor_titulo or 0), 2),
                "banco":      "Itaú",
            })
    except Exception:
        pass
    rows.sort(key=lambda r: (r["pagamento"] or "", r["nome"]))
    return {"rows": rows}


@router.get("/regua-clientes")
def regua_clientes(
    month: int = Query(default=None),
    year:  int = Query(default=None),
    current_user: User = Depends(get_current_user),
):
    """Clientes com OVERDUE no mês anterior (M-1) — quem pagou e quem não pagou."""
    hoje = date.today()
    m = month or hoje.month
    y = year  or hoje.year
    pm = m - 1 if m > 1 else 12
    py = y   if m > 1 else y - 1
    prev_first = date(py, pm, 1)

    db2 = SessionLocal()
    try:
        rows = db2.execute(text("""
            WITH overdue_prev AS (
                SELECT DISTINCT customer_id, customer_name, customer_cpf_cnpj,
                       SUM(value) AS valor_vencido
                FROM asaas_payments_sync
                WHERE status = 'OVERDUE'
                  AND customer_id IS NOT NULL
                  AND EXTRACT(YEAR  FROM due_date) = :py
                  AND EXTRACT(MONTH FROM due_date) = :pm
                GROUP BY customer_id, customer_name, customer_cpf_cnpj
            ),
            regularizados AS (
                SELECT DISTINCT p.customer_id
                FROM asaas_payments_sync p
                JOIN overdue_prev o ON o.customer_id = p.customer_id
                WHERE p.status IN ('RECEIVED', 'CONFIRMED')
                  AND p.credit_date IS NOT NULL
                  AND p.credit_date >= :from_date
            )
            SELECT
                o.customer_id, o.customer_name AS nome, o.customer_cpf_cnpj AS cnpj,
                o.valor_vencido,
                CASE WHEN r.customer_id IS NOT NULL THEN true ELSE false END AS pagou
            FROM overdue_prev o
            LEFT JOIN regularizados r ON r.customer_id = o.customer_id
            ORDER BY pagou DESC, o.valor_vencido DESC
        """), {"py": py, "pm": pm, "from_date": prev_first}).fetchall()
        db2.close()
    except Exception:
        db2.close()
        return {"mes_ref": f"{pm:02d}/{py}", "clientes": []}

    def _fmt(raw):
        d = (raw or "").replace(".", "").replace("/", "").replace("-", "").strip()
        if len(d) == 14:
            return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
        if len(d) == 11:
            return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
        return raw or ""

    return {
        "mes_ref": f"{pm:02d}/{py}",
        "clientes": [
            {
                "nome":         r.nome or "Não identificado",
                "cnpj":         _fmt(r.cnpj),
                "valor_vencido": round(float(r.valor_vencido or 0), 2),
                "pagou":        bool(r.pagou),
            }
            for r in rows
        ],
    }


@router.get("/sync-status")
async def get_sync_status(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Retorna quando foi o último sync do Asaas e quantos registros há no banco local."""
    try:
        row = db.execute(text("""
            SELECT MAX(synced_at) AS last_sync, COUNT(*) AS total
            FROM asaas_payments_sync
        """)).fetchone()
        return {
            "last_sync": row.last_sync.isoformat() if row and row.last_sync else None,
            "total_payments": row.total if row else 0,
            "sync_interval_minutes": 20,
        }
    except Exception:
        return {"last_sync": None, "total_payments": 0, "sync_interval_minutes": 20}


@router.post("/sync-now")
async def trigger_sync(
    current_user: User = Depends(get_current_user),
):
    """Dispara um sync manual imediato (admin)."""
    from app.services.asaas_sync import run_sync
    result = await run_sync()
    return result


# ─────────────────────────────────────────────
# NOTAS DE VENCIDOS (vencimento planejado + observação)
# ─────────────────────────────────────────────

class VencidoNotaUpsert(BaseModel):
    vencimento_planejado: Optional[str] = None  # YYYY-MM-DD
    observacao: Optional[str] = None


@router.get("/vencidos-notas")
def get_vencidos_notas(
    mes: int = Query(..., ge=1, le=12),
    ano: int = Query(..., ge=2020),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from app.models.extra import VencidoNota
    notas = db.query(VencidoNota).filter(
        VencidoNota.mes == mes,
        VencidoNota.ano == ano,
    ).all()
    return {
        n.cnpj: {
            "vencimento_planejado": str(n.vencimento_planejado) if n.vencimento_planejado else None,
            "observacao": n.observacao,
        }
        for n in notas
    }


@router.put("/vencidos-notas/{cnpj}")
def upsert_vencido_nota(
    cnpj: str,
    mes: int = Query(..., ge=1, le=12),
    ano: int = Query(..., ge=2020),
    data: VencidoNotaUpsert = Body(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from app.models.extra import VencidoNota
    from datetime import date as _date

    nota = db.query(VencidoNota).filter(
        VencidoNota.cnpj == cnpj,
        VencidoNota.mes == mes,
        VencidoNota.ano == ano,
    ).first()

    if not nota:
        nota = VencidoNota(cnpj=cnpj, mes=mes, ano=ano)
        db.add(nota)

    if data.vencimento_planejado is not None:
        try:
            nota.vencimento_planejado = _date.fromisoformat(data.vencimento_planejado) if data.vencimento_planejado else None
        except ValueError:
            nota.vencimento_planejado = None
    else:
        nota.vencimento_planejado = None

    nota.observacao = data.observacao
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────
# TEMPLATE XLSX + UPLOAD EM LOTE
# ─────────────────────────────────────────────

@router.get("/vencidos-template")
async def download_vencidos_template(
    mes: int = Query(..., ge=1, le=12),
    ano: int = Query(..., ge=2020),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Gera planilha xlsx com a lista de vencidos do mês + notas existentes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import Response as FastResponse
    from app.models.extra import VencidoNota
    from datetime import date as _date
    import calendar as _cal

    _last_day = _cal.monthrange(ano, mes)[1]
    _inicio   = _date(ano, mes, 1)
    _lim      = _date(ano, mes, _last_day)

    try:
        rows = db.execute(text("""
            WITH asaas_venc AS (
                SELECT
                    REGEXP_REPLACE(aps.customer_cpf_cnpj, '[^0-9]', '', 'g') AS cnpj_raw,
                    aps.customer_name AS nome,
                    aps.value
                FROM asaas_payments_sync aps
                WHERE aps.status = 'OVERDUE'
                  AND aps.due_date >= :inicio AND aps.due_date <= :lim
            ),
            itau_venc AS (
                SELECT
                    REGEXP_REPLACE(ib.cpf_cnpj, '[^0-9]', '', 'g') AS cnpj_raw,
                    ib.pagador AS nome,
                    ib.valor_titulo AS value
                FROM itau_boletos ib
                WHERE ib.status = 'vencida'
                  AND ib.data_vencimento >= :inicio AND ib.data_vencimento <= :lim
            ),
            combined AS (SELECT * FROM asaas_venc UNION ALL SELECT * FROM itau_venc),
            venc_by_cnpj AS (
                SELECT
                    cnpj_raw,
                    MAX(nome) AS nome,
                    SUM(value) AS valor
                FROM combined
                WHERE cnpj_raw IS NOT NULL AND cnpj_raw <> ''
                GROUP BY cnpj_raw
            )
            SELECT
                v.cnpj_raw,
                v.nome,
                v.valor,
                COALESCE(c.id_smart, '') AS id_smart
            FROM venc_by_cnpj v
            LEFT JOIN clients c ON REGEXP_REPLACE(c.cpf_cnpj, '[^0-9]', '', 'g') = v.cnpj_raw
            ORDER BY v.valor DESC
        """), {"inicio": _inicio, "lim": _lim}).fetchall()
    except Exception as _e:
        print(f"⚠️  vencidos-template query: {_e}")
        rows = []

    notas = {
        n.cnpj: n
        for n in db.query(VencidoNota).filter(
            VencidoNota.mes == mes,
            VencidoNota.ano == ano,
        ).all()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Vencidos {mes:02d}-{ano}"

    header_fill = PatternFill("solid", fgColor="3CB54A")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    lock_fill   = PatternFill("solid", fgColor="F3F4F6")
    lock_font   = Font(color="9CA3AF", size=9)

    headers = [
        ("ID_Smart",                     16),
        ("CNPJ",                         18),
        ("Nome",                         35),
        ("Venc. Planejado (AAAA-MM-DD)", 22),
        ("Observação",                   45),
    ]

    for col, (h, width) in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + col)].width = width

    ws.row_dimensions[1].height = 22

    def _fmt_cnpj(raw: str) -> str:
        d = (raw or "").strip()
        if len(d) == 14:
            return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
        if len(d) == 11:
            return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
        return raw or ""

    for i, r in enumerate(rows, 2):
        cnpj_raw = r.cnpj_raw or ""
        nota     = notas.get(cnpj_raw)
        vp_val   = str(nota.vencimento_planejado) if nota and nota.vencimento_planejado else ""
        obs_val  = nota.observacao if nota and nota.observacao else ""

        c1 = ws.cell(row=i, column=1, value=r.id_smart or "")
        c1.fill = lock_fill; c1.font = lock_font
        c1.alignment = Alignment(horizontal="left")

        c2 = ws.cell(row=i, column=2, value=_fmt_cnpj(cnpj_raw))
        c2.fill = lock_fill; c2.font = lock_font
        c2.alignment = Alignment(horizontal="left")

        c3 = ws.cell(row=i, column=3, value=r.nome or "")
        c3.fill = lock_fill; c3.font = lock_font
        c3.alignment = Alignment(horizontal="left")

        c4 = ws.cell(row=i, column=4, value=vp_val)
        c4.font = Font(size=9); c4.alignment = Alignment(horizontal="center")

        c5 = ws.cell(row=i, column=5, value=obs_val)
        c5.font = Font(size=9); c5.alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"vencidos_{ano}_{mes:02d}.xlsx"
    return FastResponse(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/vencidos-upload")
async def upload_vencidos_planilha(
    mes: int = Query(..., ge=1, le=12),
    ano: int = Query(..., ge=2020),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Lê xlsx enviado e faz upsert em lote dos campos Venc. Planejado e Observação."""
    import openpyxl
    from app.models.extra import VencidoNota
    from datetime import date as _date

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um .xlsx.")

    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Planilha vazia ou sem dados.")

    header   = [str(h or "").strip().lower() for h in rows[0]]
    cnpj_idx = next((i for i, h in enumerate(header) if "cnpj" in h), None)
    vp_idx   = next((i for i, h in enumerate(header) if "venc" in h and "plan" in h), None)
    obs_idx  = next((i for i, h in enumerate(header) if "observ" in h), None)

    if cnpj_idx is None or vp_idx is None or obs_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Colunas obrigatórias não encontradas. Use a planilha modelo com colunas: CNPJ, Venc. Planejado, Observação",
        )

    updated = 0
    errors  = []

    for row_num, row in enumerate(rows[1:], 2):
        if not row or len(row) <= max(cnpj_idx, vp_idx, obs_idx):
            continue

        raw_cnpj = str(row[cnpj_idx] or "").replace(".", "").replace("/", "").replace("-", "").strip()
        if not raw_cnpj or raw_cnpj in ("None", "nan"):
            continue

        vp_raw  = str(row[vp_idx]  or "").strip()
        obs_raw = str(row[obs_idx] or "").strip()

        vp_date = None
        if vp_raw and vp_raw not in ("None", "nan"):
            try:
                vp_date = _date.fromisoformat(vp_raw[:10])
            except ValueError:
                try:
                    parts   = vp_raw.split("/")
                    vp_date = _date(int(parts[2]), int(parts[1]), int(parts[0]))
                except Exception:
                    errors.append(f"Linha {row_num}: data inválida '{vp_raw}'")
                    continue

        nota = db.query(VencidoNota).filter(
            VencidoNota.cnpj == raw_cnpj,
            VencidoNota.mes  == mes,
            VencidoNota.ano  == ano,
        ).first()

        if not nota:
            nota = VencidoNota(cnpj=raw_cnpj, mes=mes, ano=ano)
            db.add(nota)

        nota.vencimento_planejado = vp_date
        nota.observacao           = obs_raw or None
        updated += 1

    db.commit()
    return {"updated": updated, "errors": errors}
