"""
Router: Previsibilidade Comportamental
Lê do banco local (asaas_payments_sync) — sem chamadas em tempo real ao Asaas.
"""
import io
import math
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import get_db, SessionLocal
from app.models import User
from app.routers.auth import get_current_user

router = APIRouter(prefix="/previsibilidade", tags=["Previsibilidade"])


# ── helpers ───────────────────────────────────────────────────────────────────

def _parse_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None


def _score(avg_dias: float) -> tuple[str, str]:
    if avg_dias >= 0:
        return "A", "A – Baixo risco"
    if avg_dias >= -5:
        return "B", "B – Médio risco"
    return "C", "C – Alto risco"


def _previsao_padrao(avg_dias: float) -> str:
    if avg_dias > 0:
        return "Paga antes do vencimento"
    if avg_dias == 0:
        return "Paga no vencimento"
    return f"Atrasa em média {abs(int(round(avg_dias)))} dias"


def _std(values: list[float]) -> float:
    n = len(values)
    if n < 2:
        return 0.0
    mean = sum(values) / n
    variance = sum((x - mean) ** 2 for x in values) / (n - 1)
    return math.sqrt(variance)


def _fmt_cnpj(raw: str) -> str:
    d = (raw or "").replace(".", "").replace("/", "").replace("-", "").strip()
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    return raw or ""


def _load_payments(year: int, month: int, months_back: int = 6) -> list[dict]:
    """Carrega pagamentos dos últimos N meses a partir do banco local."""
    ref = date(year, month, 1)
    months = []
    for i in range(months_back):
        m = ref.month - i
        y = ref.year
        while m <= 0:
            m += 12
            y -= 1
        months.append((y, m))

    if not months:
        return []

    # monta cláusula WHERE para os meses
    conditions = " OR ".join(
        f"(EXTRACT(YEAR FROM due_date)={y} AND EXTRACT(MONTH FROM due_date)={m})"
        for y, m in months
    )

    db = SessionLocal()
    try:
        rows = db.execute(text(f"""
            SELECT asaas_id, customer_id, customer_name, customer_cpf_cnpj,
                   value, net_value, due_date, payment_date, credit_date,
                   status, billing_type, external_reference
            FROM asaas_payments_sync
            WHERE {conditions}
        """)).fetchall()

        return [
            {
                "id":               r.asaas_id,
                "customer":         r.customer_id,
                "name":             r.customer_name,
                "cpfCnpj":          r.customer_cpf_cnpj,
                "value":            r.value or 0,
                "netValue":         r.net_value,
                "dueDate":          r.due_date.isoformat() if r.due_date else None,
                "paymentDate":      r.payment_date.isoformat() if r.payment_date else None,
                "creditDate":       r.credit_date.isoformat() if r.credit_date else None,
                "status":           r.status,
                "billingType":      r.billing_type,
                "externalReference": r.external_reference,
            }
            for r in rows
        ]
    finally:
        db.close()


# ── endpoint: dados JSON para visualização ────────────────────────────────────

@router.get("/summary")
async def get_previsibilidade_summary(
    month: int = Query(..., ge=1, le=12),
    year:  int = Query(..., ge=2020),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    all_payments = _load_payments(year, month, months_back=6)

    if not all_payments:
        from fastapi import HTTPException
        raise HTTPException(status_code=503, detail="Sem dados no banco local. Aguarde o próximo sync (20 min).")

    # pagamentos do mês selecionado
    prefix = f"{year}-{month:02d}"

    # 1. calcula diffs por customer_id (apenas RECEIVED/CONFIRMED)
    customer_data: dict[str, dict] = {}
    for p in all_payments:
        if p.get("status") not in ("RECEIVED", "CONFIRMED"):
            continue
        cid = p.get("customer") or ""
        if not cid:
            continue
        due_date    = _parse_date(p.get("dueDate"))
        credit_date = _parse_date(p.get("creditDate") or p.get("paymentDate"))
        if not due_date or not credit_date:
            continue
        diff = (due_date - credit_date).days
        if cid not in customer_data:
            customer_data[cid] = {
                "diffs": [], "name": p.get("name"), "cpfCnpj": p.get("cpfCnpj")
            }
        customer_data[cid]["diffs"].append(diff)
        if not customer_data[cid]["name"] and p.get("name"):
            customer_data[cid]["name"] = p.get("name")

    # 2. cobranças PENDING do mês selecionado
    pending = [
        p for p in all_payments
        if p.get("status") == "PENDING"
        and (p.get("dueDate") or "").startswith(prefix)
    ]

    # 3. monta scores
    scores = []
    for cid, info in customer_data.items():
        diffs = info["diffs"]
        avg   = sum(diffs) / len(diffs)
        sc, cl = _score(avg)
        scores.append({
            "customer_id":     cid,
            "id_smart":        "",
            "nome":            info.get("name") or "Não identificado",
            "cnpj":            _fmt_cnpj(info.get("cpfCnpj") or ""),
            "qtd_pagamentos":  len(diffs),
            "avg_dias":        round(avg, 1),
            "std_dias":        round(_std(diffs), 1),
            "score":           sc,
            "classificacao":   cl,
            "previsao_padrao": _previsao_padrao(avg),
        })

    order = {"C": 0, "B": 1, "A": 2}
    scores.sort(key=lambda x: (order.get(x["score"], 9), x["avg_dias"]))

    # 4. monta cobranças abertas com previsão
    score_map = {s["customer_id"]: s for s in scores}
    pending_list = []
    for p in pending:
        cid      = p.get("customer") or ""
        due_date = _parse_date(p.get("dueDate"))
        sc_info  = score_map.get(cid)
        if sc_info:
            delta = int(round(sc_info["avg_dias"]))
            data_prevista = (due_date - timedelta(days=delta)).isoformat() if due_date else None
            obs = f"Baseado em {sc_info['qtd_pagamentos']} pagamento(s)"
            sc  = sc_info["score"]
        else:
            data_prevista = None
            obs = "Sem histórico"
            sc  = None

        pending_list.append({
            "customer_id":   cid,
            "id_smart":      "",
            "nome":          p.get("name") or "Não identificado",
            "cnpj":          _fmt_cnpj(p.get("cpfCnpj") or ""),
            "valor":         p.get("value", 0),
            "vencimento":    p.get("dueDate"),
            "data_prevista": data_prevista,
            "score":         sc,
            "observacao":    obs,
        })

    # 5. KPIs
    total = len(scores)
    kpis = {
        "total_clientes":  total,
        "pct_score_a":     round(sum(1 for s in scores if s["score"] == "A") / total * 100, 1) if total else 0,
        "pct_score_b":     round(sum(1 for s in scores if s["score"] == "B") / total * 100, 1) if total else 0,
        "pct_score_c":     round(sum(1 for s in scores if s["score"] == "C") / total * 100, 1) if total else 0,
        "total_pending":   len(pending_list),
        "valor_pending":   round(sum(p["valor"] for p in pending_list), 2),
    }

    return {"kpis": kpis, "scores": scores, "pending": pending_list}


# ── endpoint: export Excel ────────────────────────────────────────────────────

@router.get("/export")
async def export_previsibilidade(
    month: int = Query(..., ge=1, le=12),
    year:  int = Query(..., ge=2020),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    all_payments = _load_payments(year, month, months_back=6)
    prefix = f"{year}-{month:02d}"

    customer_data: dict[str, dict] = {}
    for p in all_payments:
        if p.get("status") not in ("RECEIVED", "CONFIRMED"):
            continue
        cid = p.get("customer") or ""
        if not cid:
            continue
        due_date    = _parse_date(p.get("dueDate"))
        credit_date = _parse_date(p.get("creditDate") or p.get("paymentDate"))
        if not due_date or not credit_date:
            continue
        diff = (due_date - credit_date).days
        if cid not in customer_data:
            customer_data[cid] = {"diffs": [], "ultima_ref": f"{month:02d}/{year}",
                                   "name": p.get("name"), "cpfCnpj": p.get("cpfCnpj")}
        customer_data[cid]["diffs"].append(diff)

    scores: dict[str, dict] = {}
    for cid, info in customer_data.items():
        diffs = info["diffs"]
        avg   = sum(diffs) / len(diffs)
        sc, cl = _score(avg)
        scores[cid] = {
            "nome":            info.get("name") or "Não identificado",
            "cnpj":            _fmt_cnpj(info.get("cpfCnpj") or ""),
            "qtd_pagamentos":  len(diffs),
            "avg_dias":        round(avg, 2),
            "std_dias":        round(_std(diffs), 2),
            "score":           sc,
            "classificacao":   cl,
            "previsao_padrao": _previsao_padrao(avg),
            "ultima_ref":      info["ultima_ref"],
        }

    pending = [
        p for p in all_payments
        if p.get("status") == "PENDING"
        and (p.get("dueDate") or "").startswith(prefix)
    ]

    previsao_rows = []
    for p in pending:
        cid      = p.get("customer") or ""
        due_date = _parse_date(p.get("dueDate"))
        valor    = p.get("value", 0)
        if cid in scores:
            sc_info   = scores[cid]
            delta     = int(round(sc_info["avg_dias"]))
            data_prev = (due_date - timedelta(days=delta)) if due_date else None
            obs       = f"Previsão baseada em {sc_info['qtd_pagamentos']} pagamento(s)"
            sc        = sc_info["score"]
            dias_lbl  = f"{delta:+d} dias"
        else:
            data_prev = None
            obs       = "Sem histórico"
            sc        = "—"
            dias_lbl  = "—"

        previsao_rows.append({
            "nome":          p.get("name") or "Não identificado",
            "cnpj":          _fmt_cnpj(p.get("cpfCnpj") or ""),
            "valor":         valor,
            "vencimento":    due_date,
            "data_prevista": data_prev,
            "score":         sc,
            "dias_label":    dias_lbl,
            "obs":           obs,
        })

    # ── gera Excel ────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    VERDE_ESCURO = "3CB54A"
    VERDE_CLARO  = "F0FDF4"
    BRANCO       = "FFFFFF"
    CINZA_CLARO  = "F9FAFB"
    LARANJA      = "F97316"
    VERMELHO_C   = "EF4444"
    VERDE_SCORE  = "16A34A"

    wb  = Workbook()
    thin = Side(style="thin", color="D1D5DB")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr_font():
        return Font(name="Arial", bold=True, color=BRANCO, size=10)
    def _title_font():
        return Font(name="Arial", bold=True, color=VERDE_ESCURO, size=12)
    def _cell_font():
        return Font(name="Arial", size=10)
    def _fill(h):
        return PatternFill("solid", fgColor=h)
    def _score_color(sc):
        return VERDE_SCORE if sc == "A" else (LARANJA if sc == "B" else VERMELHO_C)

    def _write_title(ws, title_text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, title_text)
        c.font = _title_font(); c.fill = _fill(VERDE_CLARO)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

    def _write_headers(ws, row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h)
            c.font = _hdr_font(); c.fill = _fill(VERDE_ESCURO)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
        ws.row_dimensions[row].height = 18

    # Aba 1: Score por Cliente
    ws1 = wb.active
    ws1.title = "Score por Cliente"
    headers1 = ["Nome", "CNPJ", "Qtd Pagamentos", "Média Dias", "Desvio Padrão", "Score", "Classificação", "Últ. Ref."]
    _write_title(ws1, f"Previsibilidade Comportamental — Score por Cliente — {month:02d}/{year}", len(headers1))
    _write_headers(ws1, 2, headers1)

    for i, (cid, info) in enumerate(scores.items()):
        row = i + 3
        fill = _fill(BRANCO) if i % 2 == 0 else _fill(CINZA_CLARO)
        vals = [info["nome"], info["cnpj"], info["qtd_pagamentos"], info["avg_dias"],
                info["std_dias"], info["score"], info["classificacao"], info["ultima_ref"]]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row, col, val)
            c.font = _cell_font(); c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
            if col == 6 and val in ("A", "B", "C"):
                c.font = Font(name="Arial", bold=True, size=10, color=_score_color(val))

    for col, w in zip(range(1, len(headers1)+1), [35, 20, 16, 12, 14, 8, 20, 14]):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Aba 2: Previsão – Cobranças Abertas
    ws2 = wb.create_sheet("Previsão – Cobranças Abertas")
    headers2 = ["Nome", "CNPJ", "Valor (R$)", "Vencimento", "Data Prevista", "Score", "Dias Antes/Após", "Observação"]
    _write_title(ws2, f"Previsibilidade — Cobranças Abertas — {month:02d}/{year}", len(headers2))
    _write_headers(ws2, 2, headers2)

    for i, row_data in enumerate(previsao_rows):
        row = i + 3
        fill = _fill(BRANCO) if i % 2 == 0 else _fill(CINZA_CLARO)
        cells = [
            (1, row_data["nome"],          None),
            (2, row_data["cnpj"],          None),
            (3, row_data["valor"],         'R$ #,##0.00'),
            (4, row_data["vencimento"],    "DD/MM/YYYY"),
            (5, row_data["data_prevista"], "DD/MM/YYYY"),
            (6, row_data["score"],         None),
            (7, row_data["dias_label"],    None),
            (8, row_data["obs"],           None),
        ]
        for col, val, fmt_str in cells:
            c = ws2.cell(row, col, val)
            c.font = _cell_font(); c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
            if fmt_str:
                c.number_format = fmt_str
            if col == 6 and val in ("A", "B", "C"):
                c.font = Font(name="Arial", bold=True, size=10, color=_score_color(val))

    for col, w in zip(range(1, len(headers2)+1), [35, 20, 14, 14, 18, 8, 16, 36]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"previsibilidade_comportamental_{month:02d}_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
